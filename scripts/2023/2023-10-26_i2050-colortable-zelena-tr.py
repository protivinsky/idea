# ========================================
# TABLE COLORING
import string
# Ensure the necessary libraries are installed
# !pip install pandas openpyxl matplotlib seaborn

# import the libraries
from datetime import datetime
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


def to_excel_rgb(rgb):
    """
    Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl.

    Args:
        rgb (tuple): A tuple of RGB values on a 0-1 scale.

    Returns:
        str: RGB value on a 0-255 scale, formatted for use in openpyxl.
    """
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))

# Create dataframe
data = {
    'A': [1, 5, 10],
    'B': [2, 6, 8],
    'C': [3, 7, 1]
}
df = pd.DataFrame(data)

# Save dataframe to excel
df.to_excel('D:/temp/output.xlsx', index=False)

# Load workbook
book = load_workbook('D:/temp/output.xlsx')
writer = pd.ExcelWriter('D:/temp/output.xlsx', engine='openpyxl')
writer.book = book

# normalize data for color mapping
norm = plt.Normalize(df.values.min(), df.values.max())
colors = plt.cm.Reds(norm(df.values))  # replace "Reds" with the color map you prefer

# Get workbook active sheet
sheet = writer.sheets['Sheet1']

for i in range(df.shape[0]):
    for j in range(df.shape[1]):
        cell = sheet.cell(row=i+2, column=j+1)
        color = to_excel_rgb(colors[i,j])
        fill = PatternFill(start_color=color,
                           end_color=color, fill_type="solid")
        cell.fill = fill

writer.save()

# FIXED API
with pd.ExcelWriter('D:/temp/output.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Sheet1')

    sheet = writer.sheets['Sheet1']

    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            cell = sheet.cell(row=i + 2, column=j + 2)
            color = to_excel_rgb(colors[i, j])
            fill = PatternFill(start_color=color,
                               end_color=color, fill_type="solid")
            cell.fill = fill


type(sheet)

# ==============
# DATA LOADING

import os
import numpy as np
import pandas as pd
from omoment import OMean, OMeanVar
import pyreadstat
from libs.extensions import *

root = r'D:\projects\jan-krajhanzl\2023-10-18_barevne-tabulky\zadani'
file = 'Česká (ne)transformace_Umiernení.sav'

df, df_meta = pyreadstat.read_sav(os.path.join(root, file), encoding='utf-8')
df.show()

y_vars = [
    ('vo3_5_rec', 1),
    ('vo3_5_rec', 2),
    ('vo3_7_rec', 1),
    ('vo3_7_rec', 2),
    ('vo3_8_rec', 1),
    ('vo3_8_rec', 2),
    ('VOTE_Count', 0),
]

x_vars_1 = [
    'vzd', 'vel', 'mat',
    'nar1_1', 'nar1_2', 'nar1_3', 'nar1_4', 'nar1_5', 'nar1_6', 'nar1_7', 'nar1_8', 'nar1_9', 'nar1_10', 'nar1_11',
    'nar1_12', 'nar1_13', 'nar1_14', 'nar1_15', 'nar1_16', 'nar1_17', 'nar1_18', 'nar1_19',
    'nar2_1', 'nar2_2', 'nar2_3', 'nar2_4', 'nar2_5', 'nar2_6', 'nar2_7', 'nar2_8', 'nar2_9', 'nar2_10', 'nar2_11',
    'nar2_12', 'nar2_13', 'nar2_14', 'nar2_15', 'nar2_16', 'nar2_17', 'nar2_18', 'nar2_19',
    'urgb', 'transf', 'gdf', 'sem_1', 'sem_2', 'sem_3', 'sem_4', 'sem_5', 'sem_6', 'sem_7', 'sem_8', 'imp_1', 'imp_2',
    'imp_3', 'imp_4', 'imp_5', 'imp_6', 'imp_7', 'imp_8', 'imp_10', 'imp_11', 'imp_12', 'imp_13', 'imp_14',
    'pol_1.x', 'pol_2.x', 'pol_3.x', 'pol_4.x', 'pol_5.x', 'pol_6.x', 'pol_7.x', 'pol_8.x', 'pol_9.x', 'pol_10.x',
    'pol_11.x', 'pol_12.x',
    'ene_1', 'ene_2', 'ene_3', 'ene_4', 'ene_5', 'ene_6', 'ene_7',
    'medt_1.x', 'medt_2.x', 'medt_3.x', 'medt_4.x', 'medt_5.x', 'env',
    'elit_1', 'elit_2', 'elit_3', 'elit_4', 'elit_5', 'elit_6', 'elit_7', 'elit_8', 'elit_9', 'elit_10', 'elit_11',
    'elit_12', 'elit_13', 'elit_14', 'elit_15', 'elit_16', 'elit_17', 'elit_18', 'elit_19',
    'know_1.x', 'know_2.x', 'know_3.x', 'know_4.x', 'know_5.x', 'know_6.x',
]

w = 'w'

# SANITY CHECKS:
#   - do I have all in data?
#   - are the frequencies identical to what they claim to be? --> yes, need to include weights of course!
#   - are all these variables of corresponding types that can be tabulated?

for y, y_value in y_vars:
    freq = (df[y] == y_value).sum()
    wfreq = df[df[y] == y_value][w].sum()
    print(f'{y}: wN = {round(wfreq)}, N = {freq}')

len(x_vars_1)
len(set(x_vars_1))
df[x_vars_1]

# Ok, check spss metadata
# - how to check the type? labels? variable_type?

[x for x in x_vars_1 if x not in df_meta.variable_to_label]  # ok, all row_vars have labels

{df_meta.original_variable_types[x] for x in x_vars_1}  # all are F8.0
set(df_meta.original_variable_types.values())

{df_meta.readstat_variable_types[x] for x in x_vars_1}  # read as double
set(df_meta.readstat_variable_types.values())

{df_meta.variable_measure[x] for x in x_vars_1}  # scale
# -> eh, it would have been better had they been marked as ordinal or nominal
set(df_meta.variable_measure.values())  # nominal, scale

# CREATE THE TABLE IN SMALL SETTING
# - also, I do not have any info about missing -> will have to try to infer it
y, y_value = y_vars[0]
x = x_vars_1[45]

foo = df[df[y] == y_value].groupby(x)[w].sum()
foo / df[df[y] == y_value][w].sum()
x_labels = df_meta.variable_value_labels[x]

x_name = df_meta.column_names_to_labels[x]

df.groupby(x)[w].sum()
df[pd.isna(df[x])][w].sum()
df.groupby(x)[w].sum().sum()
df[w].sum()

np.isnan(df[x]).any()




##############################################################
# TODO:
#   1. Code the processing to prepare data for a single x var
#   2. Figure out how to output it into XLSX
#      - colors
#      - var labels
#      - excel writer
#   3. Handle the missing values correctly
#   4. Additional columns:
#      - sums of positive and negative (manual classification?)
#      - tests
#   ...then we are ready to process to the second part


##############################################################
# 1. Code the processing to prepare data for a single x var
out = pd.DataFrame(index=sorted(df[x].unique()))
for y, y_value in y_vars:
    sum_per_group = df[df[y] == y_value].groupby(x, dropna=False)[w].sum()
    y_total = df[df[y] == y_value][w].sum()
    pct_per_group = sum_per_group / y_total
    out[f'{y}_{y_value}'] = pct_per_group


sum_total = df.groupby(x, dropna=False)[w].sum()
pct_total = sum_total / df[w].sum()
out['total'] = pct_total
out = out.fillna(0.)
out = out.sort_index()


##############################################################
# 2. Figure out how to output it into XLSX (colors, labels, writer)
def temp_output():
    return f'D:/temp/output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

def to_excel_rgb(rgb):
    """ Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl. """
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))

wb = Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# HEADER
sheet.cell(row=1, column=1, value="VARIABLE")
sheet.cell(row=1, column=2, value="NAME")
sheet.cell(row=1, column=3, value="VALUE")
sheet.cell(row=1, column=4, value="LABEL")

for j, (y, y_value) in enumerate(y_vars):
    y_name = df_meta.column_names_to_labels[y]
    y_label = df_meta.variable_value_labels[y][float(y_value)] if y in df_meta.variable_value_labels else 0
    sheet.cell(row=1, column=5 + j, value=f'{y} = {y_value} ["{y_name}" = "{y_label}"]')

sheet.cell(row=1, column=5 + len(y_vars), value='TOTAL')

# PROCEED COLUMN-WISE
next_row = 2

sheet.cell(row=next_row, column=1, value=x)
sheet.cell(row=next_row, column=2, value=x_name)
for i, x_value in enumerate(out.index):
    sheet.cell(row=next_row + i, column=3, value=str(x_value))
    if np.isnan(x_value):
        if len(df_meta.variable_value_labels[x]) == 1:
            x_label = f'MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]'
        else:
            x_label = "MISSING"
    elif x_value in df_meta.variable_value_labels[x]:
        x_label = df_meta.variable_value_labels[x][x_value]
    else:
        x_label = str(x_value)
    sheet.cell(row=next_row + i, column=4, value=x_label)

for j, out_col in enumerate(out.columns):
    y_out = out[out_col]
    norm = plt.Normalize(y_out.values.min(), y_out.values.max())
    colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
    for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
        # lighten colors
        cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
        fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
        cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
        cell.fill = fill
        cell.number_format = '0.0%'

wb.save(temp_output())

# CONDITIONAL FORMATTING
wb = Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# HEADER
sheet.cell(row=1, column=1, value="VARIABLE")
sheet.cell(row=1, column=2, value="NAME")
sheet.column_dimensions[get_column_letter(2)].width = 30
sheet.cell(row=1, column=3, value="VALUE")
sheet.cell(row=1, column=4, value="LABEL")
sheet.column_dimensions[get_column_letter(4)].width = 12

for j, (y, y_value) in enumerate(y_vars):
    y_name = df_meta.column_names_to_labels[y]
    y_label = df_meta.variable_value_labels[y][float(y_value)] if y in df_meta.variable_value_labels else 0
    sheet.cell(row=1, column=5 + j, value=f'{y} = {y_value} ["{y_name}" = "{y_label}"]')
    sheet.column_dimensions[get_column_letter(5 + j)].width = 12

sheet.cell(row=1, column=5 + len(y_vars), value='TOTAL')
sheet.column_dimensions[get_column_letter(5 + len(y_vars))].width = 12

# PROCEED COLUMN-WISE
next_row = 2
color_scale_rule = ColorScaleRule(start_type='min', start_color='FFEF9C', end_type='max', end_color='63BE7B')
center_wrap = Alignment(vertical='center', wrap_text=True)


cell = sheet.cell(row=next_row, column=1, value=x)
cell.alignment = center_wrap
sheet.merge_cells(f'A{next_row}:A{next_row + len(out.index) - 1}')

cell = sheet.cell(row=next_row, column=2, value=x_name)
cell.alignment = center_wrap
sheet.merge_cells(f'B{next_row}:B{next_row + len(out.index) - 1}')

for i, x_value in enumerate(out.index):
    sheet.cell(row=next_row + i, column=3, value=str(x_value))
    if np.isnan(x_value):
        if len(df_meta.variable_value_labels[x]) == 1:
            x_label = f'MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]'
        else:
            x_label = "MISSING"
    elif x_value in df_meta.variable_value_labels[x]:
        x_label = df_meta.variable_value_labels[x][x_value]
    else:
        x_label = str(x_value)
    sheet.cell(row=next_row + i, column=4, value=x_label)

for j, out_col in enumerate(out.columns):
    y_out = out[out_col]
    norm = plt.Normalize(y_out.values.min(), y_out.values.max())
    colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
    for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
        # lighten colors
        cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
        # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
        cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
        # cell.fill = fill
        cell.number_format = '0.0%'

    col_letter = get_column_letter(5 + j)
    rng = f'{col_letter}{next_row}:{col_letter}{next_row + i}'
    sheet.conditional_formatting.add(rng, color_scale_rule)

wb.save(temp_output())




# DO IT ALL NOW
# CONDITIONAL FORMATTING
wb = Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

color_scale_rule = ColorScaleRule(start_type='min', start_color='FFEF9C', end_type='max', end_color='63BE7B')
center_wrap = Alignment(vertical='center', wrap_text=True)
horizontal_center = Alignment(horizontal='center')
horizontal_center_wrap = Alignment(horizontal='center', wrap_text=True)
gray_font = Font(color='808080')
thin_border = Border(bottom=Side(style='thin'))
thick_border = Border(bottom=Side(style='thick'))

# HEADER
sheet.cell(row=1, column=1, value="VARIABLE")
sheet.cell(row=1, column=2, value="NAME")
sheet.column_dimensions[get_column_letter(2)].width = 36
sheet.cell(row=1, column=3, value="VALUE")
sheet.cell(row=1, column=4, value="LABEL")
sheet.column_dimensions[get_column_letter(4)].width = 24
for j in range(1, 5):
    sheet.cell(row=1, column=j).border = thick_border

for j, (y, y_value) in enumerate(y_vars):
    y_name = df_meta.column_names_to_labels[y]
    y_label = df_meta.variable_value_labels[y][float(y_value)] if y in df_meta.variable_value_labels else 0
    cell = sheet.cell(row=1, column=5 + j, value=f'{y} = {y_value} ["{y_name}" = "{y_label}"]')
    cell.alignment = horizontal_center_wrap
    cell.border = thick_border
    sheet.column_dimensions[get_column_letter(5 + j)].width = 18

sheet.cell(row=1, column=5 + len(y_vars), value='TOTAL').border = thick_border
sheet.column_dimensions[get_column_letter(5 + len(y_vars))].width = 12

# PROCEED COLUMN-WISE
next_row = 2
for x in x_vars_1:
    print(f'Processing {x}')
    x_name = df_meta.column_names_to_labels[x]
    out = pd.DataFrame(index=sorted(df[x].unique()))
    for y, y_value in y_vars:
        sum_per_group = df[df[y] == y_value].groupby(x, dropna=False)[w].sum()
        y_total = df[df[y] == y_value][w].sum()
        pct_per_group = sum_per_group / y_total
        out[f'{y}_{y_value}'] = pct_per_group


    sum_total = df.groupby(x, dropna=False)[w].sum()
    pct_total = sum_total / df[w].sum()
    out['total'] = pct_total
    out = out.fillna(0.)
    out = out.sort_index()

    cell = sheet.cell(row=next_row, column=1, value=x)
    cell.alignment = center_wrap
    cell.border = thick_border
    sheet.merge_cells(f'A{next_row}:A{next_row + len(out.index) - 1}')

    cell = sheet.cell(row=next_row, column=2, value=x_name)
    cell.alignment = center_wrap
    cell.border = thick_border
    sheet.merge_cells(f'B{next_row}:B{next_row + len(out.index) - 1}')

    for i, x_value in enumerate(out.index):
        sheet.cell(row=next_row + i, column=3, value=str(x_value)).border = thin_border
        if np.isnan(x_value):
            if len(df_meta.variable_value_labels[x]) == 1:
                x_label = f'MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]'
            else:
                x_label = "MISSING"
        elif x_value in df_meta.variable_value_labels[x]:
            x_label = df_meta.variable_value_labels[x][x_value]
        else:
            x_label = str(x_value)
        sheet.cell(row=next_row + i, column=4, value=x_label).border = thin_border
    sheet.cell(row=next_row + i, column=3).border = thick_border
    sheet.cell(row=next_row + i, column=4).border = thick_border

    for j, out_col in enumerate(out.columns):
        y_out = out[out_col]
        norm = plt.Normalize(y_out.values.min(), y_out.values.max())
        colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
        for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
            # lighten colors
            cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
            # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
            cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
            # cell.fill = fill
            cell.number_format = '0.0%'
            cell.alignment = horizontal_center
            cell.border = thin_border if i < len(y_out) - 1 else thick_border
            if out_col == 'total':
                cell.font = gray_font
        if out_col != 'total':
            col_letter = get_column_letter(5 + j)
            rng = f'{col_letter}{next_row}:{col_letter}{next_row + i}'
            sheet.conditional_formatting.add(rng, color_scale_rule)

    next_row = next_row + i + 1

wb.save(temp_output())






# normalize data for color mapping
norm = plt.Normalize(df.values.min(), df.values.max())
colors = plt.cm.Reds(norm(df.values))  # replace "Reds" with the color map you prefer


# FIXED API
with pd.ExcelWriter('D:/temp/output.xlsx', engine='openpyxl') as writer:

    df.to_excel(writer, sheet_name='Sheet1')
    sheet = writer.sheets['Sheet1']

    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            cell = sheet.cell(row=i + 2, column=j + 2)
            color = to_excel_rgb(colors[i, j])
            fill = PatternFill(start_color=color,
                               end_color=color, fill_type="solid")
            cell.fill = fill



import pandas as pd
import numpy as np

# Sample dataframe
data = {
    'Category': ['A', 'B', np.nan, 'A', np.nan, 'C'],
    'Values': [10, 20, 30, 40, 50, 60]
}

foo = pd.DataFrame(data)

grouped = foo.groupby('Category').sum()
print(grouped)





