import os
from datetime import datetime
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from omoment import OMean
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import pyreadstat
from statsmodels.stats.weightstats import DescrStatsW
from libs.extensions import *


root = r'D:\projects\jan-krajhanzl\2023-10-18_barevne-tabulky\zadani'
file = 'Česká (ne)transformace_Umiernení.sav'

df, df_meta = pyreadstat.read_sav(os.path.join(root, file), encoding='utf-8', user_missing=True)
# df.show()

y_vars_1 = [
    ('vo3_5_rec', 1),
    ('vo3_5_rec', 2),
    ('vo3_7_rec', 1),
    ('vo3_7_rec', 2),
    ('vo3_8_rec', 1),
    ('vo3_8_rec', 2),
    ('VOTE_Count', 0),
]

y_vars_2 = [
    ('vo3_3_rec', 1),
    ('vo3_3_rec', 2),
    ('vo3_6_rec', 1),
    ('vo3_6_rec', 2),
]

# frequencies, Kruskal-Wallis
x_vars_1 = [
    'vzd', 'vel', 'mat',
    'nar1_1', 'nar1_2', 'nar1_3', 'nar1_4', 'nar1_5', 'nar1_6', 'nar1_7', 'nar1_8', 'nar1_9', 'nar1_10', 'nar1_11',
    'nar1_12', 'nar1_13', 'nar1_14', 'nar1_15', 'nar1_16', 'nar1_17', 'nar1_18', 'nar1_19',
    'nar2_1', 'nar2_2', 'nar2_3', 'nar2_4', 'nar2_5', 'nar2_6', 'nar2_7', 'nar2_8', 'nar2_9', 'nar2_10', 'nar2_11',
    'nar2_12', 'nar2_13', 'nar2_14', 'nar2_15', 'nar2_16', 'nar2_17', 'nar2_18', 'nar2_19',
    'urgb', 'transf', 'gdf', 'sem_1', 'sem_2', 'sem_3', 'sem_4', 'sem_5', 'sem_6', 'sem_7', 'sem_8', 'imp_1', 'imp_2',
    'imp_3', 'imp_4', 'imp_5', 'imp_6', 'imp_7', 'imp_8', 'imp_10', 'imp_11', 'imp_12', 'imp_13', 'imp_14',
    'pol_1.x', 'pol_2.x', 'pol_3.x', 'pol_4.x', 'pol_5.x', 'pol_6.x', 'pol_7.x', 'pol_8.x', 'pol_9.x', 'pol_10.x',
    'pol_11.x', 'pol_12.x',
    'ene_1', 'ene_2', 'ene_3', 'ene_4', 'ene_5', 'ene_6', 'ene_7',
    'medt_1.x', 'medt_2.x', 'medt_3.x', 'medt_4.x', 'medt_5.x', 'env',
    'elit_1', 'elit_2', 'elit_3', 'elit_4', 'elit_5', 'elit_6', 'elit_7', 'elit_8', 'elit_9', 'elit_10', 'elit_11',
    'elit_12', 'elit_13', 'elit_14', 'elit_15', 'elit_16', 'elit_17', 'elit_18', 'elit_19',
    'know_1.x', 'know_2.x', 'know_3.x', 'know_4.x', 'know_5.x', 'know_6.x',
]

# frequencies, Chi-square
x_vars_2 = ['poh', 'belief', 'clu#', 'tridy_6r']

# scale variables
x_vars_3 = ['vek', 'scient', 'nation', 'nar_opt', 'nar_dis', 'nar_env', 'nar_anti', 'sem_poz', 'imp_poz', 'imp_env',
            'medt_envi', 'elit_op', 'elit_dis', 'elit_exp', 'elit_pow']

w = 'w'

agg_omitted_vars = ['vzd', 'vel', 'mat', 'urgb', 'gdf', 'transf'] + x_vars_2


# SANITY CHECKS:
#   - do I have all in data?
#   - are the frequencies identical to what they claim to be? --> yes, need to include weights of course!
#   - are all these variables of corresponding types that can be tabulated?

# for y, y_value in y_vars:
#     freq = (df[y] == y_value).sum()
#     wfreq = df[df[y] == y_value][w].sum()
#     print(f'{y}: wN = {round(wfreq)}, N = {freq}')
#
# len(x_vars_1)
# len(set(x_vars_1))
# df[x_vars_1]
#
# # Ok, check spss metadata
# # - how to check the type? labels? variable_type?
#
# [x for x in x_vars_1 if x not in df_meta.variable_to_label]  # ok, all row_vars have labels
#
# {df_meta.original_variable_types[x] for x in x_vars_1}  # all are F8.0
# set(df_meta.original_variable_types.values())
#
# {df_meta.readstat_variable_types[x] for x in x_vars_1}  # read as double
# set(df_meta.readstat_variable_types.values())
#
# {df_meta.variable_measure[x] for x in x_vars_1}  # scale
# # -> eh, it would have been better had they been marked as ordinal or nominal
# set(df_meta.variable_measure.values())  # nominal, scale



# DO IT ALL NOW
# CONDITIONAL FORMATTING

def temp_output():
    return f'D:/temp/output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

def to_excel_rgb(rgb):
    """ Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl. """
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))


wb = Workbook()
sheet = wb.active
sheet.title = 'Sheet1'

# STYLES ETC
color_scale_rule = ColorScaleRule(start_type='min', start_color='FFEF9C', end_type='max', end_color='63BE7B')
color_scale_rule_3 = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percent', mid_value=50,
                                    mid_color='FFEB84', end_type='max', end_color='63BE7B')
color_scale_rule_cont = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percent', mid_value=50,
                                       mid_color='FCFCFF', end_type='max', end_color='5A8AC6')
# 3-graded: red F8696B, yellow FFEB84, green 63BE7B
# 3-graded for scale vars: red F8696B, white FCFCFF, blue 5A8AC6
center_wrap = Alignment(vertical='center', wrap_text=True)
horizontal_center = Alignment(horizontal='center')
both_center = Alignment(horizontal='center', vertical='center')
horizontal_center_wrap = Alignment(horizontal='center', wrap_text=True)
gray_font = Font(color='808080')
border_color = '666666'
thin_side = Side(style='thin', color=border_color)
thick_side = Side(style='medium', color=border_color)
thin_border = Border(bottom=thin_side)
thick_border = Border(bottom=thick_side)
value_cell_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)
value_cell_thick_border = Border(left=thin_side, right=thin_side, bottom=thick_side)
all_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)
all_thin_upper_thick_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thick_side)
top_thick_border = Border(top=thick_side)
left_thin_border = Border(left=thin_side)
right_thin_border = Border(right=thin_side)
right_thin_top_thick_border = Border(right=thin_side, top=thick_side)
left_thin_top_thick_border = Border(left=thin_side, top=thick_side)

for sheet_num, y_vars in enumerate([y_vars_1, y_vars_2]):
    if sheet_num:
        sheet = wb.create_sheet(title=f'Sheet{1 + sheet_num}')

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 36
    sheet.cell(row=1, column=3, value="VALUE").alignment = horizontal_center
    sheet.cell(row=1, column=4, value="LABEL").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(4)].width = 24

    for y_offset in (5, 7 + len(y_vars) + 1):

        for j, (y, y_value) in enumerate(y_vars):
            y_name = df_meta.column_names_to_labels[y]
            y_label = df_meta.variable_value_labels[y][float(y_value)] if y in df_meta.variable_value_labels else 0
            y_name_label = '' if y_name is None else f' [{y_name} = {y_label}]'
            cell = sheet.cell(row=1, column=y_offset + j, value=f'{y} = {y_value}{y_name_label}')
            cell.alignment = horizontal_center_wrap
            sheet.column_dimensions[get_column_letter(y_offset + j)].width = 18

        sheet.cell(row=1, column=y_offset + len(y_vars), value='TOTAL').alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(y_offset + len(y_vars))].width = 12

    sheet.column_dimensions[get_column_letter(6 + len(y_vars))].width = 12
    sheet.column_dimensions[get_column_letter(7 + len(y_vars))].width = 12
    sheet.cell(row=1, column=6 + len(y_vars), value='Kruskal-Wallis test').alignment = horizontal_center_wrap
    sheet.merge_cells(f'{get_column_letter(6 + len(y_vars))}1:{get_column_letter(7 + len(y_vars))}1')
    sheet.cell(row=2, column=6 + len(y_vars), value='Chui-Square').alignment = horizontal_center_wrap
    sheet.cell(row=2, column=7 + len(y_vars), value='Signifikance').alignment = horizontal_center_wrap

    col_freq_desc = 'Sloupcové relativní četnosti (formátování dle sloupců)'
    sheet.cell(row=2, column=5, value=col_freq_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f'E2:{get_column_letter(5 + len(y_vars))}2')

    col_agg_desc = 'Součty kladných a záporných odpovědí (formátování dle řádků)'
    sheet.cell(row=2, column=8 + len(y_vars), value=col_agg_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f'{get_column_letter(8 + len(y_vars))}2:{get_column_letter(8 + 2 * len(y_vars))}2')

    total_columns = 7 + 2 * len(y_vars) + 1
    merged_columns = 2
    for j in range(1, total_columns + 1):
        if j > 4:
            sheet.cell(row=1, column=j).border = all_thin_border
        sheet.cell(row=2, column=j).border = value_cell_thick_border if j > 4 else thick_border

    for j in range(1, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)

    # PROCEED COLUMN-WISE
    next_row = 3
    for x in x_vars_1 + x_vars_2:  # [:7]:
        print(f'Processing {x}')
        x_name = df_meta.column_names_to_labels[x]
        out_index = sorted(set([*df_meta.variable_value_labels[x].keys(), *df[x].unique()]))
        out = pd.DataFrame(index=out_index)
        for y, y_value in y_vars:
            sum_per_group = df[df[y] == y_value].groupby(x, dropna=False)[w].sum()
            y_total = df[df[y] == y_value][w].sum()
            pct_per_group = sum_per_group / y_total
            out[f'{y}_{y_value}'] = pct_per_group

        sum_total = df.groupby(x, dropna=False)[w].sum()
        pct_total = sum_total / df[w].sum()
        out['total'] = pct_total
        out = out.fillna(0.)
        out = out.sort_index()

        cell = sheet.cell(row=next_row, column=1, value=x)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f'A{next_row}:A{next_row + len(out.index) - 1}')

        cell = sheet.cell(row=next_row, column=2, value=x_name)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f'B{next_row}:B{next_row + len(out.index) - 1}')

        for i, x_value in enumerate(out.index):
            sheet.cell(row=next_row + i, column=3, value=str(x_value)).border = thin_border
            if np.isnan(x_value):
                if len(df_meta.variable_value_labels[x]) == 1:
                    x_label = f'MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]'
                else:
                    x_label = "MISSING"
            elif x_value in df_meta.variable_value_labels[x]:
                x_label = df_meta.variable_value_labels[x][x_value]
            else:
                x_label = str(x_value)
            sheet.cell(row=next_row + i, column=4, value=x_label).border = thin_border
        sheet.cell(row=next_row + i, column=3).border = thick_border
        sheet.cell(row=next_row + i, column=4).border = thick_border

        # COLUMN-PERCENTAGES
        for j, out_col in enumerate(out.columns):
            y_out = out[out_col]
            norm = plt.Normalize(y_out.values.min(), y_out.values.max())
            colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
            for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
                # lighten colors
                cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
                # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
                cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
                # cell.fill = fill
                cell.number_format = '0.0%'
                cell.alignment = horizontal_center
                cell.border = value_cell_thin_border if i < len(y_out) - 1 else value_cell_thick_border
                if out_col == 'total':
                    cell.font = gray_font
            if out_col != 'total':
                col_letter = get_column_letter(5 + j)
                rng = f'{col_letter}{next_row}:{col_letter}{next_row + i}'
                sheet.conditional_formatting.add(rng, color_scale_rule)

        # KRUSKAL WALLIS TEST
        for i in range(out.shape[0]):
            sheet.cell(row=next_row + i,
                       column=5 + out.shape[1]).border = left_thin_border if i else left_thin_top_thick_border
            sheet.cell(row=next_row + i,
                       column=6 + out.shape[1]).border = right_thin_border if i else right_thin_top_thick_border

        # POSITIVE-NEGATIVE AGGREGATIONS
        agg_initial_col = 7 + out.shape[1]
        if x not in agg_omitted_vars:
            # this is a bit hacky, but should work
            agg_count = (out.shape[0] - 1) // 2
            agg_start = out.iloc[:agg_count].sum(axis=0)
            agg_end = out.iloc[agg_count + 1: 2 * agg_count + 1].sum(axis=0)
            for j, (sum_start, sum_end) in enumerate(zip(agg_start, agg_end)):
                for offset, sum_value in [(0, sum_start), (agg_count + 1, sum_end)]:
                    cell = sheet.cell(row=next_row + offset, column=agg_initial_col + j, value=sum_value)
                    cell.number_format = '0.0%'
                    cell.alignment = both_center
                    cell.border = all_thin_border if offset else all_thin_upper_thick_border
                    if j == out.shape[1] - 1:
                        cell.font = gray_font

                sheet.cell(row=next_row + offset - 1, column=agg_initial_col + j).border = all_thin_border
                for i in range(2 * agg_count + 1, out.shape[0]):
                    sheet.cell(row=next_row + i, column=agg_initial_col + j).border = all_thin_border

            # conditional formatting
            col_letter_start = get_column_letter(agg_initial_col)
            col_letter_end = get_column_letter(agg_initial_col + out.shape[1] - 2)
            rng_start = f'{col_letter_start}{next_row}:{col_letter_end}{next_row}'
            rng_end = f'{col_letter_start}{next_row + agg_count + 1}:{col_letter_end}{next_row + agg_count + 1}'
            sheet.conditional_formatting.add(rng_start, color_scale_rule_3)
            sheet.conditional_formatting.add(rng_end, color_scale_rule_3)

            # merge if necessary
            if agg_count > 1:
                for j in range(out.shape[1]):
                    col_letter = get_column_letter(agg_initial_col + j)
                    sheet.merge_cells(f'{col_letter}{next_row}:{col_letter}{next_row + agg_count - 1}')
                    sheet.merge_cells(
                        f'{col_letter}{next_row + agg_count + 1}:{col_letter}{next_row + 2 * agg_count}')

        else:  # only fix borders here
            for j in range(out.shape[1]):
                # bd = all_thin_border if i else all_thin_upper_thick_border
                sheet.cell(row=next_row, column=agg_initial_col + j).border = top_thick_border

        next_row = next_row + out.shape[0]

    # Fix some missing borders
    after_last_column = 9 + 2 * len(y_vars)
    for i in range(1, next_row):
        sheet.cell(row=i, column=after_last_column).border = left_thin_border
    for j in range(1, after_last_column):
        sheet.cell(row=next_row, column=j).border = top_thick_border


# temp_file = temp_output()
# wb.save(temp_file)
#
# os.startfile(temp_file)
#
# # PHASE 3: CONTINUOUS VARIABLES
# # TODO:
# #  - means, medians - per group and total (better in two rows)
#
# wb = Workbook()
# sheet = wb.active
# sheet.title = 'Sheet1'
sheet = wb.create_sheet(title=f'Sheet3')

y_vars = y_vars_1 + y_vars_2
x_vars = x_vars_3

for x in x_vars:
    if x in df_meta.missing_ranges:
        print(x, df_meta.missing_ranges[x])
        assert len(df_meta.missing_ranges[x]) == 1
        assert df_meta.missing_ranges[x][0]['lo'] == df_meta.missing_ranges[x][0]['hi']
        df[x] = df[x].replace(df_meta.missing_ranges[x][0]['lo'], np.nan)



# HEADER
sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
sheet.column_dimensions[get_column_letter(1)].width = 12
sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
sheet.column_dimensions[get_column_letter(2)].width = 60
sheet.cell(row=1, column=3, value="TYPE").alignment = horizontal_center
sheet.column_dimensions[get_column_letter(3)].width = 12

next_col = 4
for j, (y, y_value) in enumerate(y_vars):
    y_name = df_meta.column_names_to_labels[y]
    y_label = df_meta.variable_value_labels[y][float(y_value)] if y in df_meta.variable_value_labels else 0
    y_name_label = '' if y_name is None else f' [{y_name} = {y_label}]'
    cell = sheet.cell(row=1, column=next_col + j, value=f'{y} = {y_value}{y_name_label}')
    cell.alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + j)].width = 18

next_col = next_col + len(y_vars)
sheet.cell(row=1, column=next_col, value='TOTAL').alignment = horizontal_center_wrap
sheet.column_dimensions[get_column_letter(next_col)].width = 12

sheet.cell(row=1, column=next_col + 1, value='MINIMUM').alignment = horizontal_center_wrap
sheet.cell(row=1, column=next_col + 2, value='MAXIMUM').alignment = horizontal_center_wrap
sheet.column_dimensions[get_column_letter(next_col + 1)].width = 12
sheet.column_dimensions[get_column_letter(next_col + 2)].width = 12
sheet.cell(row=1, column=next_col + 3, value='ANOVA').alignment = horizontal_center_wrap
sheet.column_dimensions[get_column_letter(next_col + 3)].width = 12
sheet.column_dimensions[get_column_letter(next_col + 4)].width = 12
sheet.merge_cells(f'{get_column_letter(next_col + 3)}1:{get_column_letter(next_col + 4)}1')

# SUBHEADER
next_col = 4
tot_weight = df[w].sum()
for j, (y, y_value) in enumerate(y_vars):
    group_weight = df[df[y] == y_value][w].sum()
    cell = sheet.cell(row=2, column=next_col + j, value=group_weight / tot_weight)
    cell.alignment = horizontal_center_wrap
    cell.number_format = '0.0%'

cell = sheet.cell(row=2, column=next_col + len(y_vars), value=1.)
cell.alignment = horizontal_center_wrap
cell.number_format = '0.0%'

next_col = len(y_vars) + 7
for j in [0, 1]:
    cell = sheet.cell(row=2, column=next_col + j, value='Signifikance' if j else 'F')
    cell.alignment = horizontal_center_wrap
cell.number_format = '0.00'

# FONTS AND BORDERS
for j in range(1, len(y_vars) + 9):
    for i in [1, 2]:
        sheet.cell(row=i, column=j).font = Font(bold=True)
for j in range(4, len(y_vars) + 9):
    for i in [1, 2]:
        sheet.cell(row=i, column=j).border = all_thin_border

# CALCULATION
next_row = 3
for i, x in enumerate(x_vars):  # [:5]):
    print(f'Processing {x}')
    x_name = df_meta.column_names_to_labels[x]

    # VAR NAMES ETC
    cell = sheet.cell(row=next_row + 2 * i, column=1, value=x)
    cell.alignment = center_wrap
    sheet.merge_cells(f'A{next_row + 2 * i}:A{next_row + 2 * i + 1}')

    cell = sheet.cell(row=next_row + 2 * i, column=2, value=x_name)
    cell.alignment = center_wrap
    sheet.merge_cells(f'B{next_row + 2 * i}:B{next_row + 2 * i + 1}')

    sheet.cell(row=next_row + 2 * i, column=3, value='průměr').alignment = center_wrap
    sheet.cell(row=next_row + 2 * i + 1, column=3, value='medián').alignment = center_wrap

    # PER GROUP MEANS AND MEDIANS
    next_col = 4
    for j, (y, y_value) in enumerate(y_vars):
        df_part = df[df[y] == y_value]

        # mean
        om = OMean.compute(x=df_part[x], w=df_part[w])
        cell = sheet.cell(row=next_row + 2 * i, column=next_col + j, value=om.mean)
        cell.number_format = '0.00'
        cell.alignment = horizontal_center

        # median
        dsw = DescrStatsW(df_part[x], weights=df_part[w])
        wmed = dsw.quantile([0.5])[0.5]
        cell = sheet.cell(row=next_row + 2 * i + 1, column=next_col + j, value=wmed)
        cell.number_format = '0.00'
        cell.alignment = horizontal_center

    rng_end = next_col + len(y_vars) - 1
    sheet.conditional_formatting.add(f'{get_column_letter(next_col)}{next_row + 2 * i}:'
                                     f'{get_column_letter(rng_end)}{next_row + 2 * i}', color_scale_rule_cont)
    sheet.conditional_formatting.add(f'{get_column_letter(next_col)}{next_row + 2 * i + 1}:'
                                     f'{get_column_letter(rng_end)}{next_row + 2 * i + 1}', color_scale_rule_cont)

    # TOTAL MEANS AND MEDIANS
    next_col = next_col + len(y_vars)
    om = OMean.compute(x=df[x], w=df[w])
    cell = sheet.cell(row=next_row + 2 * i, column=next_col, value=om.mean)
    cell.number_format = '0.00'
    cell.alignment = horizontal_center

    dsw = DescrStatsW(df[x], weights=df[w])
    wmed = dsw.quantile([0.5])[0.5]
    cell = sheet.cell(row=next_row + 2 * i + 1, column=next_col, value=wmed)
    cell.number_format = '0.00'
    cell.alignment = horizontal_center

    # MINIMUM AND MAXIMUM
    cell = sheet.cell(row=next_row + 2 * i, column=next_col + 1, value=df[x].min())
    cell.number_format = '0.00'
    cell.alignment = Alignment(vertical='center', horizontal='center')
    cell = sheet.cell(row=next_row + 2 * i, column=next_col + 2, value=df[x].max())
    cell.number_format = '0.00'
    cell.alignment = Alignment(vertical='center', horizontal='center')
    sheet.merge_cells(f'{get_column_letter(next_col + 1)}{next_row + 2 * i}:'
                      f'{get_column_letter(next_col + 1)}{next_row + 2 * i + 1}')
    sheet.merge_cells(f'{get_column_letter(next_col + 2)}{next_row + 2 * i}:'
                      f'{get_column_letter(next_col + 2)}{next_row + 2 * i + 1}')

    # BORDERS
    for j in [1, 2, next_col + 1, next_col + 2]:
        sheet.cell(row=next_row + 2 * i, column=j).border = Border(top=thick_side, bottom=thick_side)
        sheet.cell(row=next_row + 2 * i + 1, column=j).border = Border(top=thick_side, bottom=thick_side)
    sheet.cell(row=next_row + 2 * i, column=3).border = Border(top=thick_side, bottom=thin_side)
    sheet.cell(row=next_row + 2 * i + 1, column=3).border = Border(top=thin_side, bottom=thick_side)
    for j in range(4, 7 + len(y_vars)):
        sheet.cell(row=next_row + 2 * i, column=j).border = Border(top=thick_side, bottom=thin_side, left=thin_side,
                                                                   right=thin_side)
        sheet.cell(row=next_row + 2 * i + 1, column=j).border = Border(top=thin_side, bottom=thick_side, left=thin_side,
                                                                       right=thin_side)
    an_i, an_j = next_row + 2 * i, 7 + len(y_vars)
    sheet.cell(row=an_i, column=an_j).border = Border(top=thick_side, left=thin_side)
    sheet.cell(row=an_i + 1, column=an_j).border = Border(bottom=thick_side, left=thin_side)
    sheet.cell(row=an_i, column=an_j + 1).border = Border(top=thick_side, right=thin_side)
    sheet.cell(row=an_i + 1, column=an_j + 1).border = Border(bottom=thick_side, right=thin_side)


temp_file = temp_output()
wb.save(temp_file)

os.startfile(temp_file)


# df
# y_vars
# df[list({y for y, _ in y_vars})].show()

# FIXME
#   - df_meta.variable_value_labels['nar1_1'] - some answers are not present in dataset - the same for ['nar2_12']
#   - missing -9 value 'urgb'


# the rest should be 3, 5, 7 scales with neutral in the middle - easy
# conditional formatting is row-based here


