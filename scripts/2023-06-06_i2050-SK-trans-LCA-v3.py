# region # IMPORTS
import pandas as pd
import numpy as np
import re
import matplotlib as mpl
mpl.use('Agg')
import matplotlib.pyplot as plt
plt.rcParams['figure.figsize'] = 12, 6
import importlib
from omoment import OMeanVar, OMean
import reportree as rt
import pyreadstat
import stata_setup
from libs.utils import *
from libs.rt_content import *
from libs.extensions import *
logger = create_logger(__name__)
# endregion

# region # TO DO
#   - Excelovská tabulka nikoli pouze s průměry, ale stejně tak i informace o missing (kolik jich bylo)
#   - POL_MEAN -> taky hodně missing, možná by stálo za to zohlednit -> dají vědět co s tím
#   - stačí 4-7 tříd pro obě varianty
#   - poslat tabulky, ideálně komplet proměnné

# So what are the final variants?
#   - 5-7 classes
#
# 1. FLAG_INF = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD
#               ordinal: GDF_R + INF_01_R + INF_02_R + CZPROB_01_R + POL_MISS_CAT
#               binary: GDF_99 + INF_01_99 + INF_02_99 + CZPROB_01_99
# 2. FLAG = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD
#           ordinal: GDF_R + CZPROB_01_R + POL_MISS_CAT
#           binary: GDF_99 + CZPROB_01_99
# 3. NOM_INF = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD
#              nominal: GDF + INF_01 + INF_02 + CZPROB_01
#              ordinal: POL_MISS_CAT
# 4. NOM = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD
#          nominal: GDF + CZPROB_01
#          ordinal: POL_MISS_CAT
# 5. SCALE_INF = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD + GDF_R_STD + INF_01_R_STD +
#                       INF_02_R_STD + CZPROB_01_R_STD + POL_MISS_STD
#                binary: GDF_99 + INF_01_99 + INF_02_99 + CZPROB_01_99
# 6. SCALE = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD + GDF_R_STD + CZPROB_01_R_STD +
#                   POL_MISS_STD
#            binary: GDF_99 + CZPROB_01_99

# Ok, I exported all of these from LatentGOLD. Now I can just redo the python stuff and I should be done.
#   - quick reports should be ok
#   - fix the issue in table gen
#   - and check how 9 / 99 aka missing are handled in the tables!

# endregion

# region # DATA LOADING
data_root = 'D:/projects/jan-krajhanzl/2023-05-28_SK-LCA-2023_slovenska-transformace'
w_col = 'w'
c_range = range(5, 8)
cl_cols = [x for n in c_range for x in [f'c{n}_max'] + [f'c{n}_{i}' for i in range(1, n + 1)]]

min_file = 'Transformacia SK 05_26 FA+CA_SK_TP_min3.sav'
full_file = 'Transformacia SK 05_26 FA+CA_SK.sav'

all_variants = [
    'FLAG_INF',
    'FLAG',
    'NOM_INF',
    'NOM',
    'SCALE_INF',
    'SCALE'
]

def load_class_file(variant):
    df, df_meta = pyreadstat.read_sav(os.path.join(data_root, min_file), encoding='utf-8')
    for n in c_range:
        cl, _ = pyreadstat.read_sav(f'{data_root}/{variant}/c{n}.sav', encoding='utf-8')
        cl['IDENT'] = cl['IDENT'].astype('int')
        to_rename = {
            'clu#': f'c{n}_max',
            **{f'clu#{i}': f'c{n}_{i}' for i in range(1, n + 1)}
        }
        cl = cl.rename(columns=to_rename)
        cl[f'c{n}_max'] = cl[f'c{n}_max'].astype('int')
        cl = cl[['IDENT'] + [v for k, v in to_rename.items()]].copy()
        df = cl if df is None else pd.merge(df, cl)
    return df, df_meta


def load_crit_file(variant):
    crit = pd.read_excel(f'{data_root}/crit.xlsx', sheet_name=variant,
                         skiprows=3)
    crit = crit.rename(columns={'Unnamed: 1': 'N'}).drop(columns=['Unnamed: 0'])
    crit['N'] = crit['N'].str.replace('-Cluster', '').astype('int')
    crit = crit.sort_values('N').reset_index(drop=True)
    return crit


def load_full_file(variant):
    df, df_meta = load_class_file(variant)
    full, full_meta = pyreadstat.read_sav(os.path.join(data_root, full_file), encoding='utf-8')
    keep_cols = ['IDENT'] + [c for c in full.columns if c not in df.columns]
    for n in c_range:
        for i in range(1, n + 1):
            df[f'c{n}_{i}_w'] = df[f'c{n}_{i}'] * df[w_col]
            cl_cols.append(f'c{n}_{i}_w')
    df = pd.merge(df, full[keep_cols])
    df_labels = {**df_meta.column_names_to_labels, **full_meta.column_names_to_labels}
    df_measure = {**df_meta.variable_measure, **full_meta.variable_measure}
    return df, df_labels, df_measure

# endregion

# region # QUICK REPORT
# So I already have the files, but I guess I need also appropriate cols
# for each variant?

def single_report(df, crit):
    rep = []
    crit_plots = []
    for c in crit.columns:
        if c != 'N':
            fig, ax = plt.subplots(figsize=(6, 4))
            sns.lineplot(data=crit, x='N', y=c, marker='o')
            ax.set(title=c)
            fig.tight_layout()
            crit_plots.append(fig)

    crit_chart = rt.Leaf(crit_plots, title='Criteria', num_cols=3)
    rep.append(crit_chart)

    # remove_vars = ['IDENT', w_col] + cl_cols
    demo_vars = ['SEX', 'AGECAT', 'REG', 'EDU', 'VMB', 'INCOME']
    suffixes = ['_99', '_R', '_STD', 'POL_MEAN']
    plot_vars = demo_vars.copy()
    for c in df.columns:
        for s in suffixes:
            if c.endswith(s):
                plot_vars.append(c)
    # plot_vars = [c for c in df.columns if c not in remove_vars]
    oms = {x: OMeanVar.compute(df[x], df[w_col]) for x in plot_vars}

    for n in c_range:
        c_frac = df.groupby(f'c{n}_max')[w_col].sum() / df[w_col].sum()
        foo = pd.DataFrame()
        # for x in flag_vars:
        for x in plot_vars:
            means = OMean.of_groupby(df, f'c{n}_max', x, w_col).apply(lambda x: x.mean)
            # if x not in bin_vars:
            #     means = (means - oms[x].mean) / oms[x].std_dev
            means = (means - oms[x].mean) / oms[x].std_dev
            foo[x] = means
        plots = []
        for i in range(1, n + 1):
            fig, ax = plt.subplots(figsize=(8, 5))
            sns.barplot(x=foo.T[i], y=foo.columns, ax=ax)
            ax.set(title=f'Class {i} ({100 * c_frac[i]:.1f} %)')
            fig.tight_layout()
            plots.append(fig)
        rep.append(rt.Leaf(plots, num_cols=2, title=f'{n} classes'))
    return rep

# variant = 'FLAG_INF'
# variant = 'SCALE_INF'
# df, _ = load_class_file(variant)
# crit = load_crit_file(variant)
# rt.Branch(single_report(df, crit), title=variant).show()

full_rep = []
for variant in all_variants:
    logger.info(f'Quick report for {variant}')
    df, _ = load_class_file(variant)
    crit = load_crit_file(variant)
    full_rep.append(rt.Branch(single_report(df, crit), title=variant))
rt.Branch(full_rep, title='Slovenská transformace 2023').show()

# endregion


# region # TABLE GENERATION
# - df, df_labels should contain all I need
# - df_measure: check if it contains 99 or 9 and the second max is much smaller

def table_for_frame(df, df_labels, df_measure):
    # I need to detect missing vars
    miss_cols = []  # list of cols with missing values
    # miss_value = {}  # value that should be treated as missing -> do I care? or shall I just replace it?
    for c, m in df_measure.items():
        if m in ['ordinal', 'nominal']:
            c_max = df[c].max()
            c_2nd_max = df[c][df[c] != c_max].max()
            if c_max in [9., 99.] and (c_max - c_2nd_max) > 1.5:
                # print('missing:', c, c_max)
                # TODO: replace miss by nans and add to list
                miss_cols.append(c)
                df[c] = df[c].replace(c_max, np.nan)

    out_frames = {}
    total_weight = df[w_col].sum()

    # output_cols
    no_output_col = ['IDENT', 'w'] + list(df.dtypes[(df.dtypes != np.float_) & (df.dtypes != np.int_)].index)
    output_cols = [c for c in df.columns if c not in cl_cols + no_output_col]

    for nclass in c_range:
        print(f'nclass = {nclass}')

        out_var = pd.Series({c: c for c in output_cols})
        out_label = pd.Series({c: df_labels[c] for c in output_cols})
        out_mean = pd.Series(dtype='float64')
        out_std_dev = pd.Series(dtype='float64')
        out_miss_w = pd.Series(dtype='float64')
        out_min = pd.Series(dtype='float64')
        out_max = pd.Series(dtype='float64')

        out_class_means = {}
        out_class_miss_w = {}
        out_maxclass_means = {}
        out_maxclass_miss_w = {}
        for i in range(1, nclass + 1):
            out_class_means[i] = pd.Series(dtype='float64')
            out_class_miss_w[i] = pd.Series(dtype='float64')
            out_maxclass_means[i] = pd.Series(dtype='float64')
            out_maxclass_miss_w[i] = pd.Series(dtype='float64')

        class_w = df[[f'c{nclass}_{i}_w' for i in range(1, nclass + 1)]].sum()
        class_w.index = np.arange(1, nclass + 1)
        maxclass_w = df.groupby(f'c{nclass}_max')[w_col].sum()

        for c in output_cols:
            omv = OMeanVar.compute(x=df[c], w=df[w_col])
            out_mean[c] = omv.mean
            out_std_dev[c] = omv.std_dev
            out_miss_w[c] = 1. - omv.weight / total_weight
            out_min[c] = df[c].min()
            out_max[c] = df[c].max()

            for i in range(1, nclass + 1):
                try:
                    om = OMean.compute(x=df['HEAT3_06'], w=df[f'c{nclass}_{i}_w'])
                except ZeroDivisionError:
                    om = OMean()
                out_class_means[i][c] = om.mean
                out_class_miss_w[i][c] = 1. - om.weight / class_w[i]

                # out_class_means[i][c] = nanaverage(df[[c, f'c{nclass}_{i}_w']], weights=f'c{nclass}_{i}_w')[0]

            foo = OMean.of_groupby(df, g=f'c{nclass}_max', x=c, w=w_col)
            # foo = df.groupby(f'c{nclass}_max')[[c, 'w']].apply(nanaverage, weights='w')[c]
            for i in range(1, nclass + 1):
                out_maxclass_means[i][c] = foo[i].mean if i in foo.index else np.nan
                out_maxclass_miss_w[i][c] = (1. - foo[i].weight / maxclass_w[i]) if i in foo.index else np.nan

        out_empty = out_var.apply(lambda _: '')

        outs = {
            'var': out_var,
            'label': out_label,
            'mean': out_mean,
            'std_dev': out_std_dev,
            'miss_w': out_miss_w,
            'min': out_min,
            'max': out_max,
            ' ': out_empty,
        }

        output = pd.DataFrame(outs)

        for i in range(1, nclass + 1):
            output[f'c{i}_mean'] = out_maxclass_means[i]
        output['c_min'] = output[[f'c{i}_mean' for i in range(1, nclass + 1)]].min(axis=1)
        output['c_max'] = output[[f'c{i}_mean' for i in range(1, nclass + 1)]].max(axis=1)
        output['c_diff'] = output['c_max'] - output['c_min']
        output['c_pct'] = output['c_diff'] / (output['max'] - output['min'])
        output['c_pct_x100'] = 100 * output['c_pct']
        for i in range(1, nclass + 1):
            output[f'c{i}_miss_w'] = out_maxclass_miss_w[i]

        output['  '] = out_empty
        for i in range(1, nclass + 1):
            output[f'cpp{i}_mean'] = out_class_means[i]
        output['cpp_min'] = output[[f'cpp{i}_mean' for i in range(1, nclass + 1)]].min(axis=1)
        output['cpp_max'] = output[[f'cpp{i}_mean' for i in range(1, nclass + 1)]].max(axis=1)
        output['cpp_diff'] = output['cpp_max'] - output['cpp_min']
        output['cpp_pct'] = output['cpp_diff'] / (output['max'] - output['min'])
        output['cpp_pct_x100'] = 100 * output['cpp_pct']
        for i in range(1, nclass + 1):
            output[f'cpp{i}_miss_w'] = out_class_miss_w[i]

        # add sizes of classes (relative)
        out_size = output.iloc[0].copy().rename('SIZE').apply(lambda _: '')
        out_size['var'] = 'SIZE'
        out_size['label'] = 'Relativní velikost tříd'
        for i in range(1, nclass + 1):
            out_size[f'cpp{i}_mean'] = class_w[i] / total_weight
            out_size[f'c{i}_mean'] = maxclass_w[i] / total_weight
        output = pd.concat([pd.DataFrame(out_size).T, output], axis=0)

        # output.to_csv(f'output/lca-ord/c{nclass}.csv', index=False, encoding='utf8')
        out_frames[nclass] = output

    return out_frames


out_variants = {}
out_file = 'SK-lca-results.xlsx'
sheet_name = '{}_{}c'
for variant in all_variants:
    logger.info(f'Processing {variant}')
    df, df_labels, df_measure = load_full_file(variant)
    out_variants[variant] = table_for_frame(df, df_labels, df_measure)

logger.info(f'Writing {out_file}')
with pd.ExcelWriter(os.path.join(data_root, out_file)) as writer:
    for variant, tables in out_variants.items():
        for n_class, sheet in tables.items():
            sheet.to_excel(writer, sheet_name=sheet_name.format(variant, n_class),
                           index=False)
logger.info('Done')

# variant = 'SCALE_INF'
# df, df_labels, df_measure = load_full_file(variant)
# foo = table_for_frame(df, df_labels, df_measure)

# endregion

# region # QUICK REPORT
ord_vars = ['GDF', 'INF_01', 'INF_02', 'BELIEF']
cont_vars = ['GDKNWall_MEAN', 'POL_MEAN', 'NAR_ANTI', 'NAR_SOLU']
flag_vars = [x + '_R' for x in ord_vars] + [x + '_99' for x in ord_vars] + [x + '_STD' for x in cont_vars]
bin_vars = [x + '_99' for x in ord_vars]
w_col = 'w'

oms = {x: OMeanVar.compute(df[x], df[w_col]) for x in flag_vars}

crit = pd.read_excel(f'{data_root}/{v}/crit.xlsx', skiprows=3)
crit = crit.rename(columns={'Unnamed: 1': 'N'}).drop(columns=['Unnamed: 0'])
crit['N'] = crit['N'].str.replace('-Cluster', '').astype('int')
crit = crit.sort_values('N').reset_index(drop=True)

# prepare plots with criteria
rep = []

crit_plots = []
for c in crit.columns:
    if c != 'N':
        fig, ax = plt.subplots(figsize=(6, 4))
        sns.lineplot(data=crit, x='N', y=c, marker='o')
        ax.set(title=c)
        fig.tight_layout()
        crit_plots.append(fig)

crit_chart = rt.Leaf(crit_plots, title='Criteria', num_cols=3)
rep.append(crit_chart)

for n in range(2, 10):
    c_frac = df.groupby(f'c{n}_max')[w_col].sum() / df[w_col].sum()
    foo = pd.DataFrame()
    for x in flag_vars:
        means = OMean.of_groupby(df, f'c{n}_max', x, w_col).apply(lambda x: x.mean)
        if x not in bin_vars:
            means = (means - oms[x].mean) / oms[x].std_dev
        foo[x] = means
    plots = []
    for i in range(1, n + 1):
        fig, ax = plt.subplots(figsize=(8, 5))
        sns.barplot(x=foo.T[i], y=foo.columns, ax=ax)
        ax.set(title=f'Class {i} ({100 * c_frac[i]:.1f} %)')
        fig.tight_layout()
        plots.append(fig)
    rep.append(rt.Leaf(plots, num_cols=2, title=f'{n} classes'))

branch_flag = rt.Branch(rep, title=f'Variant {v}')
branch_flag.show()
# endregion

# region # VARIANT NOM
v = 'nom'
df, df_meta = pyreadstat.read_sav(f'{data_root}/Transformacia SK 05_26 FA+CA_SK_TP_min2.sav', encoding='utf-8')
# df = None

for n in range(2, 10):
    cl, _ = pyreadstat.read_sav(f'{data_root}/{v}/c{n}.sav', encoding='utf-8')

    cl['IDENT'] = cl['IDENT'].astype('int')
    to_rename = {
        'clu#': f'c{n}_max',
        **{f'clu#{i}': f'c{n}_{i}' for i in range(1, n + 1)}
    }
    cl = cl.rename(columns=to_rename)
    cl[f'c{n}_max'] = cl[f'c{n}_max'].astype('int')
    cl = cl[['IDENT'] + [v for k, v in to_rename.items()]].copy()
    df = cl if df is None else pd.merge(df, cl)

nom_vars = ord_vars + [x + '_STD' for x in cont_vars]
bin_vars = []
w_col = 'w'

oms = {x: OMeanVar.compute(df[x], df[w_col]) for x in nom_vars}

crit = pd.read_excel(f'{data_root}/{v}/crit.xlsx', skiprows=3)
crit = crit.rename(columns={'Unnamed: 1': 'N'}).drop(columns=['Unnamed: 0'])
crit['N'] = crit['N'].str.replace('-Cluster', '').astype('int')
crit = crit.sort_values('N').reset_index(drop=True)

# prepare plots with criteria
rep = []

crit_plots = []
for c in crit.columns:
    if c != 'N':
        fig, ax = plt.subplots(figsize=(6, 4))
        sns.lineplot(data=crit, x='N', y=c, marker='o')
        ax.set(title=c)
        fig.tight_layout()
        crit_plots.append(fig)

crit_chart = rt.Leaf(crit_plots, title='Criteria', num_cols=3)
rep.append(crit_chart)

for n in range(2, 10):
    c_frac = df.groupby(f'c{n}_max')[w_col].sum() / df[w_col].sum()
    foo = pd.DataFrame()
    for x in nom_vars:
        means = OMean.of_groupby(df, f'c{n}_max', x, w_col).apply(lambda x: x.mean)
        if x not in bin_vars:
            means = (means - oms[x].mean) / oms[x].std_dev
        foo[x] = means
    plots = []
    for i in range(1, n + 1):
        fig, ax = plt.subplots(figsize=(8, 5))
        sns.barplot(x=foo.T[i], y=foo.columns, ax=ax)
        ax.set(title=f'Class {i} ({100 * c_frac[i]:.1f} %)')
        fig.tight_layout()
        plots.append(fig)
    rep.append(rt.Leaf(plots, num_cols=2, title=f'{n} classes'))

branch_nom = rt.Branch(rep, title=f'Variant {v}')

rt.Branch([branch_flag, branch_nom], title='SK Transformace 2023, LCA').show()
# branch_flag.show()
# endregion


# ========================
# generate table for GDF



def table_for_gdf(df, df_labels, df_measure):
    # I need to detect missing vars
    miss_cols = []  # list of cols with missing values
    # miss_value = {}  # value that should be treated as missing -> do I care? or shall I just replace it?
    for c, m in df_measure.items():
        if m in ['ordinal', 'nominal'] and (c != 'GDF'):
            c_max = df[c].max()
            c_2nd_max = df[c][df[c] != c_max].max()
            if c_max in [9., 99.] and (c_max - c_2nd_max) > 1.5:
                # print('missing:', c, c_max)
                # TODO: replace miss by nans and add to list
                miss_cols.append(c)
                df[c] = df[c].replace(c_max, np.nan)

    total_weight = df[w_col].sum()
    gdf_values = [0., 1., 2., 3., 99.]

    # output_cols
    no_output_col = ['IDENT', 'w'] + list(df.dtypes[(df.dtypes != np.float_) & (df.dtypes != np.int_)].index)
    output_cols = [c for c in df.columns if c not in cl_cols + no_output_col + ['GDF']]

    out_var = pd.Series({c: c for c in output_cols})
    out_label = pd.Series({c: df_labels[c] for c in output_cols})
    out_mean = pd.Series(dtype='float64')
    out_std_dev = pd.Series(dtype='float64')
    out_miss_w = pd.Series(dtype='float64')
    out_min = pd.Series(dtype='float64')
    out_max = pd.Series(dtype='float64')

    out_maxclass_means = {}
    out_maxclass_miss_w = {}
    for i in gdf_values:
        out_maxclass_means[i] = pd.Series(dtype='float64')
        out_maxclass_miss_w[i] = pd.Series(dtype='float64')

    maxclass_w = df.groupby('GDF')[w_col].sum()

    for c in output_cols:
        omv = OMeanVar.compute(x=df[c], w=df[w_col])
        out_mean[c] = omv.mean
        out_std_dev[c] = omv.std_dev
        out_miss_w[c] = 1. - omv.weight / total_weight
        out_min[c] = df[c].min()
        out_max[c] = df[c].max()

        foo = OMean.of_groupby(df, g='GDF', x=c, w=w_col)
        for i in gdf_values:
            out_maxclass_means[i][c] = foo[i].mean if i in foo.index else np.nan
            out_maxclass_miss_w[i][c] = (1. - foo[i].weight / maxclass_w[i]) if i in foo.index else np.nan

    out_empty = out_var.apply(lambda _: '')

    outs = {
        'var': out_var,
        'label': out_label,
        'mean': out_mean,
        'std_dev': out_std_dev,
        'miss_w': out_miss_w,
        'min': out_min,
        'max': out_max,
        ' ': out_empty,
    }

    output = pd.DataFrame(outs)

    for i in gdf_values:
        output[f'gdf{i}_mean'] = out_maxclass_means[i]
    output['gdf_min'] = output[[f'gdf{i}_mean' for i in gdf_values]].min(axis=1)
    output['gdf_max'] = output[[f'gdf{i}_mean' for i in gdf_values]].max(axis=1)
    output['gdf_diff'] = output['gdf_max'] - output['gdf_min']
    output['gdf_pct'] = output['gdf_diff'] / (output['max'] - output['min'])
    output['gdf_pct_x100'] = 100 * output['gdf_pct']
    for i in gdf_values:
        output[f'gdf{i}_miss_w'] = out_maxclass_miss_w[i]

    # add sizes of classes (relative)
    out_size = output.iloc[0].copy().rename('SIZE').apply(lambda _: '')
    out_size['var'] = 'SIZE'
    out_size['label'] = 'Relativní velikost tříd'
    for i in gdf_values:
        out_size[f'gdf{i}_mean'] = maxclass_w[i] / total_weight
    output = pd.concat([pd.DataFrame(out_size).T, output], axis=0)
    return output


foo = load_full_file('FLAG')
foo[0].show()
foo[0]['GDF'].value_counts()
# create cols gdf_0, 1, 2, 3, 99
# and use them as max classes
# I can do it just right away, right?


out_file = 'SK-gdf-results.xlsx'
logger.info(f'Processing GDF')
df, df_labels, df_measure = load_full_file('FLAG')
out_gdf = table_for_gdf(df, df_labels, df_measure)

logger.info(f'Writing GDF')
with pd.ExcelWriter(os.path.join(data_root, out_file)) as writer:
        out_gdf.to_excel(writer, sheet_name='GDF', index=False)
logger.info('Done')


# Can I also join the data here?
df, df_meta = load_class_file('NOM')
df.columns

# NOM_5c
# 4. NOM = scale: GDNKWall_MEAN_STD + POL_MEAN_STD + NAR_ANTI_STD + NAR_SOLU_STD
#          nominal: GDF + CZPROB_01
#          ordinal: POL_MISS_CAT

full, full_meta = pyreadstat.read_sav(os.path.join(data_root, full_file), encoding='utf-8')

lca_cols = ['GDKNWall_MEAN_STD', 'POL_MEAN_STD', 'NAR_ANTI_STD', 'NAR_SOLU_STD', 'GDF', 'CZPROB_01', 'POL_MISS_CAT']
keep_cols = ['IDENT', 'c5_max', *[f'c5_{i}' for i in range(1, 6)], *[c for c in lca_cols if c not in full.columns]]

df = pd.merge(full, df[keep_cols], on='IDENT')
df_labels = {**full_meta.column_names_to_labels, **{k: v for k, v in df_meta.column_names_to_labels.items() if k in df.columns}}
df_measure = {**full_meta.variable_measure, **{k: v for k, v in df_meta.variable_measure.items() if k in df.columns}}

full_meta.column_names_to_labels = df_labels
full_meta.variable_measure = df_measure

pyreadstat.write_sav(
    df=df,
    dst_path=os.path.join(data_root, 'nom_5c_merged.sav'),
    column_labels=full_meta.column_names_to_labels,
    row_compress=True,
    variable_value_labels=full_meta.variable_value_labels,
    variable_display_width=full_meta.variable_display_width,
    variable_measure=full_meta.variable_measure,
    variable_format=full_meta.original_variable_types,
)



pyreadstat.write_sav(
    df=df,
    dst_path=os.path.join(data_root, 'nom_5c_merged.sav'),
    column_labels=df_labels,
    variable_value_labels=full_meta.variable_value_labels,
    variable_measure=df_measure,
)


# ========================================
# TABLE COLORING

# Ensure the necessary libraries are installed
# !pip install pandas openpyxl matplotlib seaborn

# import the libraries
import pandas as pd
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Color

def to_excel_rgb(rgb):
    """
    Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl.

    Args:
        rgb (tuple): A tuple of RGB values on a 0-1 scale.

    Returns:
        str: RGB value on a 0-255 scale, formatted for use in openpyxl.
    """
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))

# Create dataframe
data = {
    'A': [1, 5, 10],
    'B': [2, 6, 8],
    'C': [3, 7, 1]
}
df = pd.DataFrame(data)

# Save dataframe to excel
df.to_excel('D:/temp/output.xlsx', index=False)

# Load workbook
book = load_workbook('D:/temp/output.xlsx')
writer = pd.ExcelWriter('D:/temp/output.xlsx', engine='openpyxl')
writer.book = book

# normalize data for color mapping
norm = plt.Normalize(df.values.min(), df.values.max())
colors = plt.cm.Reds(norm(df.values)) # replace "Reds" with the color map you prefer

# Get workbook active sheet
sheet = writer.sheets['Sheet1']

for i in range(df.shape[0]):
    for j in range(df.shape[1]):
        cell = sheet.cell(row=i+2, column=j+2)
        color = to_excel_rgb(colors[i,j])
        fill = PatternFill(start_color=color,
                           end_color=color, fill_type="solid")
        cell.fill = fill

writer.save()




