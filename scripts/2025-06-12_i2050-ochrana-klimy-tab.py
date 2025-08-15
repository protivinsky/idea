import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *


data_root = Path("/home/thomas/projects/jan-krajhanzl/2025-06-11_lca-o-ochrane-klimy")
data_file = "25144-01-SK_DATA_ALL-FINAL-FINAL_VAHY_20250603_TP_classes.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file)
# df.shape

all_vars_df = pd.read_excel(f"{data_root}/Premenné pre porovnanie.xlsx").dropna().copy()
all_vars_df.columns = ["names"]
all_vars_df["measure"] = all_vars_df["names"].apply(
    lambda x: df_meta.variable_measure[x] if x in df_meta.variable_measure else None
)
all_vars = list(all_vars_df.values[:, 0])
#
# df_meta.variable_measure[all_vars[0]]
# all_vars_df["measure"].value_counts()
# df_meta.variable_value_labels[all_vars[-1]]

# FIXES:
for c in ["POL_MISS"] + [f"OS_{i}" for i in range(1, 4)] + ["CARS"]:
    df_meta.variable_measure[c] = "scale"

nom_fixes = ["SEX", "EDU", "VMB", "SET", "INT", "POJ", "INC", "INC_PCT", "INC_MEDP", "EL21",
             "HOME", "HEAT1", "EDU2", "ETS2A", "ETS2B"]

for k, v in df_meta.variable_measure.items():
    if v == "nominal" and  k not in nom_fixes:
        df_meta.variable_measure[k] = "ordinal"

xs = [x for x in all_vars if df_meta.variable_measure[x] in ["nominal", "ordinal"]]
cont_xs = [x for x in all_vars if df_meta.variable_measure[x] == "scale"]

for i in range(1, 6):
    df_meta.variable_value_labels[f"SQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 9):
    df_meta.variable_value_labels[f"SOC_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 10):
    df_meta.variable_value_labels[f"EQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}


# xs = [ "GDF", "POL_MISS", "NAR_1", "NAR_2", "NAR_3", "NAR_4", "NAR_5", "NAR_6", "NAR_7", "NAR_8",
#       "NAR_9", "NAR_10", "NAR_11", "NAR_12", "NAR_13", "NAR_14", "NAR_15", "NAR_16", "NAR_17",
#       "NAR_18", "NAR_19", "NAR_20", "NAR_21", "NAR_22", "NAR_23", "NAR_24", "NAR_25", "NAR_26",
#       "NAR_27", "NAR_28", "NAR_29", "NAR_30", "NAR_31", "NAR_32", "NAR_33", "NAR_34", "NAR_35",
#       "NAR_36", "NAR_37", "NAR_38", "NAR_39", "POL_1", "POL_2", "POL_3", "POL_4", "POL_5",
#       "POL_6", "POL_7", "POL_8", "POL_9", "POL_10", "POL_11", "POL_12", "ENV2", "BELIEF",
#       "URGB", "SKPROB_1", "SKPROB_2", "SKPROB_3", "SKPROB_4", "SKPROB_5", "SKPROB_6",
#       "KNOW_1_all", "KNOW_2_all", "KNOW_3_all", "KNOW_4_all", "KNOW_5_all", "ENE_1", "ENE_2",
#       "ENE_3", "ENE_4", "ENE_5", "AID", "VE_A", "VE_B", "VE_C", "BRAND_1", "BRAND_2", "BRAND_3",
#       "BRAND_4", "BRAND_5", "BRAND_6", "BRAND_7", "OSV_1", "OSV_2", "OSV_3", "OSV_4", "OSV_5",
#       "OSV_6", "OSV_7", "OSV_8", "GDGD_GD", "GDGD_ZD", "GDGD_OK", "GDGD_ZM", "GDGD_ST", "GDGD_U",
#       "GDGD_RES", "GDGD_REG", "SQ_1", "SQ_2", "SQ_3", "SQ_4", "SQ_5", "SEX", "EDU", "VMB", "SET",
#       "INT", "OS_1", "OS_2", "OS_3", "ECS", "ECON", "VYD", "POJ", "INC", "INC_PCT", "INC_MEDP",
#       "SOC_1", "SOC_2", "SOC_3", "SOC_4", "SOC_5", "SOC_6", "SOC_7", "SOC_8", "MEDT_1", "MEDT_2",
#       "MEDT_3", "MEDT_4", "MEDT_5", "MEDT_6", "MEDT_7", "MEDT_8", "MEDT_9", "MEDT_10", "MEDT_11",
#       "MEDT_12", "MEDT_13", "MEDT_14", "MEDT_15", "MEDT_16", "MEDT_17", "MEDT_18", "MEDT_19",
#       "MEDT_20", "MEDT_21", "MEDT_22", "MEDT_23", "MEDT_24", "MEDT_25", "MEDT_26", "MEDT_27",
#       "NEED_1", "NEED_2", "NEED_3", "NEED_4", "NEED_5", "NEED_6", "NEED_7", "NEED_8", "NEED_9",
#       "NEED_10", "NEED_11", "NEED_12", "NEED_13", "NEED_14", "NEED_15", "NEED_16", "NEED_17",
#       "NEED_18", "NEED_19", "NEED_20", "NEED_21", "NEED_22", "NEED_23", "NEED_24", "NEED_25",
#       "NEED_26", "NEED_27", "NEED_28", "NEED_29", "IDENT_1", "IDENT_2", "IDENT_3", "IDENT_4",
#       "IDENT_5", "IDENT_6", "IDENT_7", "IDENT_8", "IDENT_9", "IDENT_10", "PO_1", "PO_2", "PO_3",
#       "PO_4", "PO_5", "PO_6", "PO_7", "PO_8", "PO_9", "PO_10", "PO_11", "PO_12", "PO_13", "PO_14",
#       "PO_15", "PO_16", "PO_17", "PO_18", "PO_19", "PO_20", "PO_21", "PO_22", "EL", "EL21",
#       "ELECT", "ELECT2", "ELECT_FIN", "VOTE_1", "VOTE_2", "VOTE_3", "VOTE_4", "VOTE_5", "VOTE_6",
#       "VOTE_7", "VOTE_8", "VOTE_9", "VOTE_10", "VOTE_11", "VOTE_12", "VOTE_13", "VOTE_14",
#       "TRP_1", "TRP_2", "TRP_3", "TRP_4", "TRP_5", "TRP_6", "TRP_7", "TRP_8", "TRP_9", "TRP_10",
#       "BIL_1", "BIL_2", "BIL_3", "HOME", "HEAT1", "HEAT3_1_FIN", "HEAT3_2_FIN", "HEAT3_3_FIN",
#       "HEAT3_4_FIN", "HEAT3_5_FIN", "HEAT3_6_FIN", "INS", "EQ_1", "EQ_2", "EQ_3", "EQ_4", "EQ_5",
#       "EQ_6", "EQ_7", "EQ_8", "EQ_9", "MOB_1", "MOB_2", "MOB_3", "MOB_4", "MOB_5", "MOB_6",
#       "MOB_7", "MOB_8", "MOB_9", "MOB_10", "MOB_11", "MOB_12", "EDU2", "TPZ", "TPZ_B_1",
#       "TPZ_B_2", "TPZ_B_3", "TPZ_B_4", "TPZ_B_5", "TPZ_B_6", "TPZ_B_7", "TPZ_B_8", "TPN",
#       "BEV_1", "BEV_2", "BEV_3", "BEV_4", "BEV_5", "BEV_6", "BEV_7", "BEV_8", "BEV_9", "BEV_DV",
#       "CARS", "ETS2A", "ETS2B", "MBAR_1", "MBAR_2", "MBAR_3", "MBAR_4", "MBAR_5", "MBAR_6",
#       "MBAR_7", "MBAR_8", "MBAR_9", "MBAR_10", "MBAR_11", "TNAR_1", "TNAR_2", "TNAR_3",
#       "TNAR_4", "TNAR_5", "TNAR_6", "TNAR_7", "TNAR_8", "TNAR_9", "TNAR_10", "TNAR_11",
#       "TNAR_12", "TNAR_13", "TNAR_14", "TNAR_15", "TMOT_1", "TMOT_2", "TMOT_3", "TMOT_4",
#       "TMOT_5", "TPPOL_1", "TPPOL_2", "TPPOL_3", "TPPOL_4", "TPPOL_5", "TPPOL_6", "TPPOL_7",
#       "TPPOL_8", "TPPOL_9", "TPPOL_10", "TPPOL_11" ]
#
# cont_xs


w = "vahy"
# agg_omitted_vars = ['POHLAVI', 'AGECAT', 'AGECAT2', 'EDU', 'EDUCAT', 'INT', 'ECS', 'KRAJ', 'OKR', 'OBEC', 'VMB', 'VMBCAT', 'SET']
# agg_omitted_vars = ["SEX", "EDU", "VMB", "SET", "INT"]
agg_omitted_vars = []

ys = [f"c{i}_max" for i in range(2, 10)]


def temp_output():
    return f"/home/thomas/tmp/output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def to_excel_rgb(rgb):
    """Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl."""
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))


# STYLES ETC
color_scale_rule = ColorScaleRule(start_type="min", start_color="FFEF9C", end_type="max", end_color="63BE7B")
color_scale_rule_3 = ColorScaleRule(
    start_type="min",
    start_color="F8696B",
    mid_type="percent",
    mid_value=50,
    mid_color="FFEB84",
    end_type="max",
    end_color="63BE7B",
)
color_scale_rule_cont = ColorScaleRule(
    start_type="min",
    start_color="F8696B",
    mid_type="percent",
    mid_value=50,
    mid_color="FCFCFF",
    end_type="max",
    end_color="5A8AC6",
)
color_scale_rule_cont_miss = ColorScaleRule(start_type="min", start_color="FCFCFF", end_type="max", end_color="F8696B")
# 3-graded: red F8696B, yellow FFEB84, green 63BE7B
# 3-graded for scale vars: red F8696B, white FCFCFF, blue 5A8AC6
center_wrap = Alignment(vertical="center", wrap_text=True)
horizontal_center = Alignment(horizontal="center")
both_center = Alignment(horizontal="center", vertical="center")
horizontal_center_wrap = Alignment(horizontal="center", wrap_text=True)
gray_font = Font(color="808080")
border_color = "666666"
thin_side = Side(style="thin", color=border_color)
thick_side = Side(style="medium", color=border_color)
thin_border = Border(bottom=thin_side)
thick_border = Border(bottom=thick_side)
value_cell_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)
value_cell_thick_border = Border(left=thin_side, right=thin_side, bottom=thick_side)
all_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)
all_thin_upper_thick_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thick_side)
top_thick_border = Border(top=thick_side)
left_thin_border = Border(left=thin_side)
right_thin_border = Border(right=thin_side)
right_thin_top_thick_border = Border(right=thin_side, top=thick_side)
left_thin_top_thick_border = Border(left=thin_side, top=thick_side)

wb = Workbook()
sheet = wb.active

if False:
    temp_file = temp_output()
    wb.save(temp_file)
    # os.system(f'xdg_open {temp_file} >/dev/null 2>&1 &')

add_tests = False
tests_gap = 2 if add_tests else 0

for sheet_num, yss in enumerate(ys):
    y_vars = np.sort(df[yss].unique())
    if sheet_num:
        # sheet = wb.create_sheet(title=sheet_titles[yss])
        sheet = wb.create_sheet(title=yss)
    else:
        # sheet.title = sheet_titles[yss]
        sheet.title = yss

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 36
    sheet.cell(row=1, column=3, value="VALUE").alignment = horizontal_center
    sheet.cell(row=1, column=4, value="LABEL").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(4)].width = 24

    for y_offset in (5, 5 + tests_gap + len(y_vars) + 1):
        tot_w = df[w].sum()
        y_vars_w = df.groupby(yss)[w].sum() / tot_w
        for j, y in enumerate(y_vars):
            y_label = f"Třída {y} ({100 * y_vars_w[y]:.1f} %)" if y != 99 else "Missing"
            cell = sheet.cell(row=1, column=y_offset + j, value=y_label)
            cell.alignment = horizontal_center_wrap
            sheet.column_dimensions[get_column_letter(y_offset + j)].width = 18

        sheet.cell(row=1, column=y_offset + len(y_vars), value="TOTAL").alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(y_offset + len(y_vars))].width = 12

    if add_tests:
        sheet.column_dimensions[get_column_letter(6 + len(y_vars))].width = 12
        sheet.column_dimensions[get_column_letter(7 + len(y_vars))].width = 12
        sheet.cell(row=1, column=6 + len(y_vars), value="Kruskal-Wallis test").alignment = horizontal_center_wrap
        sheet.merge_cells(f"{get_column_letter(6 + len(y_vars))}1:{get_column_letter(7 + len(y_vars))}1")
        sheet.cell(row=2, column=6 + len(y_vars), value="Chui-Square").alignment = horizontal_center_wrap
        sheet.cell(row=2, column=7 + len(y_vars), value="Signifikance").alignment = horizontal_center_wrap

    col_freq_desc = "Sloupcové relativní četnosti (formátování dle sloupců)"
    sheet.cell(row=2, column=5, value=col_freq_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f"E2:{get_column_letter(5 + len(y_vars))}2")

    col_agg_desc = "Součty kladných a záporných odpovědí (formátování dle řádků)"
    sheet.cell(row=2, column=6 + tests_gap + len(y_vars), value=col_agg_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(
        f"{get_column_letter(6 + tests_gap + len(y_vars))}2:{get_column_letter(6 + tests_gap + 2 * len(y_vars))}2"
    )

    total_columns = 5 + tests_gap + 2 * len(y_vars) + 1
    merged_columns = 2
    for j in range(1, total_columns + 1):
        if j > 4:
            sheet.cell(row=1, column=j).border = all_thin_border
        sheet.cell(row=2, column=j).border = value_cell_thick_border if j > 4 else thick_border

    for j in range(1, 2 * len(y_vars) + 7 + tests_gap):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)

    # PROCEED COLUMN-WISE
    next_row = 3
    xs_remove = ["NUM_1", "NUM_2", "NUM_3", "OKR", "OBEC"]
    for x in [x for x in xs if x not in xs_remove]:  # [:7]:
        print(f"Processing {yss} -> {x}")
        x_name = df_meta.column_names_to_labels[x]
        out_index = sorted(set([*df_meta.variable_value_labels[x].keys(), *df[x].unique()]))
        out = pd.DataFrame(index=out_index)
        for y in y_vars:
            sum_per_group = df[df[yss] == y].groupby(x, dropna=False)[w].sum()
            y_total = df[df[yss] == y][w].sum()
            pct_per_group = sum_per_group / y_total
            out[f"{yss}_{y}"] = pct_per_group

        sum_total = df.groupby(x, dropna=False)[w].sum()
        pct_total = sum_total / df[w].sum()
        out["total"] = pct_total
        out = out.fillna(0.0)
        out = out.sort_index()

        cell = sheet.cell(row=next_row, column=1, value=x)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f"A{next_row}:A{next_row + len(out.index) - 1}")

        cell = sheet.cell(row=next_row, column=2, value=x_name)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f"B{next_row}:B{next_row + len(out.index) - 1}")

        for i, x_value in enumerate(out.index):
            sheet.cell(row=next_row + i, column=3, value=str(x_value)).border = thin_border
            if np.isnan(x_value):
                if len(df_meta.variable_value_labels[x]) == 1:
                    x_label = f"MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]"
                else:
                    x_label = "MISSING"
            elif x_value in df_meta.variable_value_labels[x]:
                x_label = df_meta.variable_value_labels[x][x_value]
            else:
                x_label = str(x_value)
            sheet.cell(row=next_row + i, column=4, value=x_label).border = thin_border
        sheet.cell(row=next_row + i, column=3).border = thick_border
        sheet.cell(row=next_row + i, column=4).border = thick_border

        # COLUMN-PERCENTAGES
        for j, out_col in enumerate(out.columns):
            y_out = out[out_col]
            norm = plt.Normalize(y_out.values.min(), y_out.values.max())
            colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
            for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
                # lighten colors
                cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
                # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
                cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
                # cell.fill = fill
                cell.number_format = "0.0%"
                cell.alignment = horizontal_center
                cell.border = value_cell_thin_border if i < len(y_out) - 1 else value_cell_thick_border
                if out_col == "total":
                    cell.font = gray_font
            if out_col != "total":
                col_letter = get_column_letter(5 + j)
                rng = f"{col_letter}{next_row}:{col_letter}{next_row + i}"
                sheet.conditional_formatting.add(rng, color_scale_rule)

        # KRUSKAL WALLIS TEST
        if add_tests:
            for i in range(out.shape[0]):
                sheet.cell(row=next_row + i, column=5 + out.shape[1]).border = (
                    left_thin_border if i else left_thin_top_thick_border
                )
                sheet.cell(row=next_row + i, column=6 + out.shape[1]).border = (
                    right_thin_border if i else right_thin_top_thick_border
                )

        # POSITIVE-NEGATIVE AGGREGATIONS
        agg_initial_col = 5 + tests_gap + out.shape[1]
        is_ordinal = x in df_meta.variable_measure and df_meta.variable_measure[x] == "ordinal"
        if x not in agg_omitted_vars and is_ordinal:
            # this is a bit hacky, but should work
            value_list = list(df_meta.variable_value_labels[x])
            value_list = [int(x) for x in value_list if int(x) not in [85, 88, 99]]
            agg_count = len(value_list) // 2
            agg_gap = len(value_list) % 2
            # agg_count = (out.shape[0] - 1) // 2
            agg_start = out.iloc[:agg_count].sum(axis=0)
            agg_end = out.iloc[agg_count + agg_gap : 2 * agg_count + agg_gap].sum(axis=0)
            for j, (sum_start, sum_end) in enumerate(zip(agg_start, agg_end)):
                for offset, sum_value in [(0, sum_start), (agg_count + agg_gap, sum_end)]:
                    cell = sheet.cell(row=next_row + offset, column=agg_initial_col + j, value=sum_value)
                    cell.number_format = "0.0%"
                    cell.alignment = both_center
                    cell.border = all_thin_border if offset else all_thin_upper_thick_border
                    if j == out.shape[1] - 1:
                        cell.font = gray_font

                sheet.cell(row=next_row + offset - 1, column=agg_initial_col + j).border = all_thin_border
                for i in range(2 * agg_count + agg_gap, out.shape[0]):
                    sheet.cell(row=next_row + i, column=agg_initial_col + j).border = all_thin_border

            # conditional formatting
            col_letter_start = get_column_letter(agg_initial_col)
            col_letter_end = get_column_letter(agg_initial_col + out.shape[1] - 2)
            rng_start = f"{col_letter_start}{next_row}:{col_letter_end}{next_row}"
            rng_end = (
                f"{col_letter_start}{next_row + agg_count + agg_gap}:{col_letter_end}{next_row + agg_count + agg_gap}"
            )
            sheet.conditional_formatting.add(rng_start, color_scale_rule_3)
            sheet.conditional_formatting.add(rng_end, color_scale_rule_3)

            # merge if necessary
            if agg_count > 1:
                for j in range(out.shape[1]):
                    col_letter = get_column_letter(agg_initial_col + j)
                    sheet.merge_cells(f"{col_letter}{next_row}:{col_letter}{next_row + agg_count - 1}")
                    sheet.merge_cells(
                        f"{col_letter}{next_row + agg_count + agg_gap}:{col_letter}{next_row + 2 * agg_count + agg_gap - 1}"
                    )

        else:  # only fix borders here
            for j in range(out.shape[1]):
                # bd = all_thin_border if i else all_thin_upper_thick_border
                sheet.cell(row=next_row, column=agg_initial_col + j).border = top_thick_border

        next_row = next_row + out.shape[0]

    # Fix some missing borders
    after_last_column = 7 + tests_gap + 2 * len(y_vars)
    for i in range(1, next_row):
        sheet.cell(row=i, column=after_last_column).border = left_thin_border
    for j in range(1, after_last_column):
        sheet.cell(row=next_row, column=j).border = top_thick_border

    # NOTE: continue with continuous variables here
    y_vars = np.sort(df[yss].unique())
    sheet = wb.create_sheet(title=f"{yss}_cont")

    for x in cont_xs:
        if x in df_meta.missing_ranges:
            print(x, df_meta.missing_ranges[x])
            # assert len(df_meta.missing_ranges[x]) == 1
            for miss_range in df_meta.missing_ranges[x]:
                assert miss_range["lo"] == miss_range["hi"]
                df[x] = df[x].replace(miss_range["lo"], np.nan)

    # for c in fix_missing_vars:
    #     df[c] = df[c].fillna(0)

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(1)].width = 12
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 60
    sheet.cell(row=1, column=3, value="TYPE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(3)].width = 12

    next_col = 4
    tot_w = df[w].sum()
    y_vars_w = df.groupby(yss)[w].sum() / tot_w
    for j, y in enumerate(y_vars):
        y_label = f"Třída {y} ({100 * y_vars_w[y]:.1f} %)" if y != 99 else "Missing"
        cell = sheet.cell(row=1, column=next_col + j, value=y_label)
        cell.alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(next_col + j)].width = 18

    next_col = next_col + len(y_vars)
    sheet.cell(row=1, column=next_col, value="TOTAL").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col)].width = 12

    sheet.cell(row=1, column=next_col + 1, value="MINIMUM").alignment = horizontal_center_wrap
    sheet.cell(row=1, column=next_col + 2, value="MAXIMUM").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 1)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 2)].width = 12
    sheet.cell(row=1, column=next_col + 3, value="ANOVA").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 3)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 4)].width = 12
    sheet.merge_cells(f"{get_column_letter(next_col + 3)}1:{get_column_letter(next_col + 4)}1")

    # SUBHEADER
    next_col = 4
    tot_weight = df[w].sum()
    for j, y_value in enumerate(y_vars):
        group_weight = df[df[yss] == y_value][w].sum()
        cell = sheet.cell(row=2, column=next_col + j, value=group_weight / tot_weight)
        cell.alignment = horizontal_center_wrap
        cell.number_format = "0.0%"

    cell = sheet.cell(row=2, column=next_col + len(y_vars), value=1.0)
    cell.alignment = horizontal_center_wrap
    cell.number_format = "0.0%"

    next_col = len(y_vars) + 7
    for j in [0, 1]:
        cell = sheet.cell(row=2, column=next_col + j, value="Signifikance" if j else "F")
        cell.alignment = horizontal_center_wrap
    cell.number_format = "0.00"

    # FONTS AND BORDERS
    for j in range(1, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)
    for j in range(4, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).border = all_thin_border

    # CALCULATION
    next_row = 3
    for i, x in enumerate(cont_xs):  # [:5]):
        print(f"Processing {x}")
        x_name = df_meta.column_names_to_labels[x]

        # VAR NAMES ETC
        cell = sheet.cell(row=next_row + 3 * i, column=1, value=x)
        cell.alignment = center_wrap
        sheet.merge_cells(f"A{next_row + 3 * i}:A{next_row + 3 * i + 2}")

        cell = sheet.cell(row=next_row + 3 * i, column=2, value=x_name)
        cell.alignment = center_wrap
        sheet.merge_cells(f"B{next_row + 3 * i}:B{next_row + 3 * i + 2}")

        sheet.cell(row=next_row + 3 * i, column=3, value="průměr").alignment = center_wrap
        sheet.cell(row=next_row + 3 * i + 1, column=3, value="medián").alignment = center_wrap
        sheet.cell(row=next_row + 3 * i + 2, column=3, value="missing").alignment = center_wrap

        # PER GROUP MEANS AND MEDIANS
        next_col = 4
        for j, y_value in enumerate(y_vars):
            df_part = df[df[yss] == y_value]

            if len(df_part.dropna(subset=[x])) > 0:
                # mean
                om = OMean.compute(x=df_part[x], w=df_part[w])
                wmean = om.mean
                # median
                dsw = DescrStatsW(df_part[x], weights=df_part[w])
                wmed = dsw.quantile([0.5])[0.5]
            else:
                wmean = np.nan
                wmed = np.nan

            part_miss = df_part[np.isnan(df_part[x])][w].sum()
            part_weight = df_part[w].sum()
            wmiss = part_miss / part_weight

            cell = sheet.cell(row=next_row + 3 * i, column=next_col + j, value=wmean)
            cell.number_format = "0.00"
            cell.alignment = horizontal_center

            cell = sheet.cell(row=next_row + 3 * i + 1, column=next_col + j, value=wmed)
            cell.number_format = "0.00"
            cell.alignment = horizontal_center

            cell = sheet.cell(row=next_row + 3 * i + 2, column=next_col + j, value=wmiss)
            cell.number_format = "0.0%"
            cell.alignment = horizontal_center

        rng_end = next_col + len(y_vars) - 1
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i}:{get_column_letter(rng_end)}{next_row + 3 * i}",
            color_scale_rule_cont,
        )
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i + 1}:{get_column_letter(rng_end)}{next_row + 3 * i + 1}",
            color_scale_rule_cont,
        )
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i + 2}:{get_column_letter(rng_end)}{next_row + 3 * i + 2}",
            color_scale_rule_cont_miss,
        )

        # TOTAL MEANS AND MEDIANS
        next_col = next_col + len(y_vars)
        om = OMean.compute(x=df[x], w=df[w])
        cell = sheet.cell(row=next_row + 3 * i, column=next_col, value=om.mean)
        cell.number_format = "0.00"
        cell.alignment = horizontal_center

        dsw = DescrStatsW(df[x], weights=df[w])
        wmed = dsw.quantile([0.5])[0.5]
        cell = sheet.cell(row=next_row + 3 * i + 1, column=next_col, value=wmed)
        cell.number_format = "0.00"
        cell.alignment = horizontal_center

        tot_miss = df[np.isnan(df[x])][w].sum()
        tot_weight = df[w].sum()
        cell = sheet.cell(row=next_row + 3 * i + 2, column=next_col, value=tot_miss / tot_weight)
        cell.number_format = "0.0%"
        cell.alignment = horizontal_center

        # MINIMUM AND MAXIMUM
        cell = sheet.cell(row=next_row + 3 * i, column=next_col + 1, value=df[x].min())
        cell.number_format = "0.00"
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell = sheet.cell(row=next_row + 3 * i, column=next_col + 2, value=df[x].max())
        cell.number_format = "0.00"
        cell.alignment = Alignment(vertical="center", horizontal="center")
        sheet.merge_cells(
            f"{get_column_letter(next_col + 1)}{next_row + 3 * i}:"
            f"{get_column_letter(next_col + 1)}{next_row + 3 * i + 2}"
        )
        sheet.merge_cells(
            f"{get_column_letter(next_col + 2)}{next_row + 3 * i}:"
            f"{get_column_letter(next_col + 2)}{next_row + 3 * i + 2}"
        )

        # BORDERS
        for j in [1, 2, next_col + 1, next_col + 2]:
            sheet.cell(row=next_row + 3 * i, column=j).border = Border(top=thick_side, bottom=thick_side)
            sheet.cell(row=next_row + 3 * i + 2, column=j).border = Border(top=thick_side, bottom=thick_side)
        sheet.cell(row=next_row + 3 * i, column=3).border = Border(top=thick_side, bottom=thin_side)
        sheet.cell(row=next_row + 3 * i + 1, column=3).border = Border(top=thin_side, bottom=thin_side)
        sheet.cell(row=next_row + 3 * i + 2, column=3).border = Border(top=thin_side, bottom=thick_side)
        for j in range(4, 7 + len(y_vars)):
            sheet.cell(row=next_row + 3 * i, column=j).border = Border(
                top=thick_side, bottom=thin_side, left=thin_side, right=thin_side
            )
            sheet.cell(row=next_row + 3 * i + 1, column=j).border = Border(
                top=thin_side, bottom=thin_side, left=thin_side, right=thin_side
            )
            sheet.cell(row=next_row + 3 * i + 2, column=j).border = Border(
                top=thin_side, bottom=thick_side, left=thin_side, right=thin_side
            )
        an_i, an_j = next_row + 3 * i, 7 + len(y_vars)
        sheet.cell(row=an_i, column=an_j).border = Border(top=thick_side, left=thin_side)
        sheet.cell(row=an_i + 1, column=an_j).border = Border(left=thin_side)
        sheet.cell(row=an_i + 2, column=an_j).border = Border(bottom=thick_side, left=thin_side)
        sheet.cell(row=an_i, column=an_j + 1).border = Border(top=thick_side, right=thin_side)
        sheet.cell(row=an_i + 1, column=an_j + 1).border = Border(right=thin_side)
        sheet.cell(row=an_i + 2, column=an_j + 1).border = Border(bottom=thick_side, right=thin_side)


temp_file = temp_output()
wb.save(temp_file)
