import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.linear_model import LogisticRegression
from xgboost import XGBClassifier, plot_importance
from sklearn.metrics import (accuracy_score, roc_auc_score, confusion_matrix,
                             classification_report, roc_curve)


data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-11-14.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file)
filt_var = "filter_$"
df = df[df[filt_var] == 1].copy()

xs = [
    'POHLAVI', 'AGECAT', 'AGECAT2', 'EDU', 'EDUCAT', 'INT', 'ECS', 'KRAJ', 'OKR', 'OBEC', 'VMB', 'VMBCAT', 'SET', 'NUM_1',
    'NUM_2', 'NUM_3', 'MEDT_1', 'MEDT_2', 'MEDT_3', 'MEDT_4', 'MEDT_5', 'MEDT_6', 'MEDT_7', 'MEDT_8', 'MEDT_9', 'MEDT_10', 'MEDT_11', 'MEDT_12',
    'MEDT_13', 'MEDT_14', 'MEDT_15', 'MEDT_16', 'MEDT_17', 'MEDT_18', 'MEDT_19', 'MEDT_20', 'MEDT_21', 'MEDT_22', 'MEDT_23', 'MEDT_24',
    'MEDT_25', 'NEED_1', 'NEED_2', 'NEED_3', 'NEED_4', 'NEED_5', 'NEED_6', 'NEED_7', 'NEED_8', 'NEED_9', 'NEED_10', 'NEED_11', 'NEED_12',
    'NEED_13', 'NEED_14', 'NEED_15', 'NEED_16', 'NEED_17', 'NEED_18', 'NEED_19', 'NEED_20', 'NEED_21', 'NEED_22', 'NEED_23', 'NEED_24',
    'PO_1', 'PO_2', 'PO_3', 'PO_4', 'PO_5', 'PO_6', 'PO_7', 'PO_8', 'PO_9', 'PO_10', 'PO_11', 'PO_12', 'PO_13', 'PO_14', 'PO_15', 'PO_16', 'PO_17', 'PO_18',
    'PO_19', 'PO_20', 'IDENT_1', 'IDENT_2', 'IDENT_3', 'IDENT_4', 'IDENT_5', 'IDENT_6', 'IDENT_7', 'IDENT_8', 'IDENT_9', 'EL', 'EL21', 'ELECT',
    'ELECT2', 'VOTE_1', 'VOTE_2', 'VOTE_3', 'VOTE_4', 'VOTE_5', 'VOTE_6', 'VOTE_7', 'VOTE_8', 'VOTE_9', 'VOTE_10', 'VOTE_11', 'VOTE_12',
    'VOTE_13', 'VOTE_14', 'VOTE_15', 'TRP_1', 'TRP_2', 'TRP_3', 'TRP_4', 'TRP_5', 'TRP_6', 'TRP_7', 'TRP_8', 'TRP_9', 'TRP_10', 'NOST_1', 'NOST_2',
    'NOST_3', 'NOST_4', 'NOST_5', 'ENV2', 'BELIEF', 'KNOW_1', 'KNOW_2', 'KNOW_3', 'KNOW_4', 'KNOW_5', 'KNOW_6', 'CZPROB_1', 'CZPROB_2',
    'CZPROB_3', 'CZPROB_4', 'CZPROB_5', 'URGB', 'URGB_REC', 'SECT_1', 'SECT_2', 'SECT_3', 'SECT_4', 'SECT_5', 'ENE_1', 'ENE_2', 'ENE_3', 'ENE_4',
    'ENE_5', 'ENE_6', 'ENE_7', 'ENE_8', 'ENC', 'VE_A', 'VE_B', 'VE_C', 'POL_1', 'POL_2', 'POL_3', 'POL_4', 'POL_5', 'POL_6', 'POL_7', 'POL_8', 'POL_9',
    'POL_10', 'BRAND_1', 'BRAND_2', 'BRAND_3', 'BRAND_4', 'BRAND_5', 'BRAND_6', 'OSV_1', 'OSV_2', 'OSV_3', 'OSV_4', 'OSV_5', 'OSV_6', 'OSV_7',
    'OSV_8', 'GDF', 'GDGD_GD', 'GDGD_ZD', 'GDGD_N', 'GDGD_OK', 'GDGD_ZM', 'GDGD_ST', 'GDGD_U', 'TRACT_1', 'TRACT_2', 'TRACT_3', 'TRACT_4',
    'TRACT_5', 'TRACT_6', 'TRACT_7', 'TRACT_8', 'TRACT_9', 'TRACT_10', 'TRACT_11', 'TRACT_12', 'NAR_1', 'NAR_2', 'NAR_3', 'NAR_4', 'NAR_5',
    'NAR_6', 'NAR_7', 'NAR_8', 'NAR_9', 'NAR_10', 'NAR_11', 'NAR_12', 'NAR_13', 'NAR_14', 'NAR_15', 'NAR_16', 'NAR_17', 'NAR_18', 'NAR_19',
    'NAR_20', 'NAR_21', 'NAR_22', 'NAR_23', 'NAR_24', 'NAR_25', 'NAR_26', 'NAR_27', 'NAR_28', 'NAR_29', 'NAR_30', 'NAR_31', 'NAR_32', 'NAR_33',
    'NAR_34', 'NAR_35', 'NAR_36', 'NAR_37', 'NAR_38', 'NAR_39', 'NAR_40', 'NAR_41', 'NAR_42', 'NAR_43', 'NAR_44', 'NAR_45', 'NAR_46', 'EQ_1',
    'EQ_2', 'EQ_3', 'EQ_4', 'EQ_5', 'EQ_6', 'EQ_7', 'EQ_8', 'EQ_9', 'EQ_10', 'EQ_11', 'EQ_12', 'EQ_13', 'INS', 'MOB_1', 'MOB_2', 'MOB_3', 'MOB_4', 'MOB_5',
    'MOB_6', 'MOB_7', 'MOB_8', 'MOB_9', 'HEAT1', 'HEAT2_1', 'HEAT2_2', 'HEAT2_3', 'HEAT2_4', 'HEAT2_5', 'HEAT2_6', 'HEAT3_1', 'HEAT3_2',
    'HEAT3_3', 'HEAT3_4', 'HEAT3_5', 'HEAT3_6', 'HEAT4', 'TPZ', 'TPN', 'BEV_1', 'BEV_2', 'BEV_3', 'BEV_4', 'BEV_5', 'BEV_6', 'BEV_7', 'BEV_8', 'BEV_9',
    'BEV_10', 'BEV_DV', 'CARS_1_4', 'CARS_1_5', 'CARS_2_4', 'CARS_2_5', 'CARS_3_4', 'CARS_3_5', 'CARS_4_4', 'CARS_4_5', 'CARS_5_4',
    'CARS_5_5', 'CPRE_1', 'CPRE_2', 'CPRE_3', 'MBAR_1', 'MBAR_2', 'MBAR_3', 'MBAR_4', 'MBAR_5', 'MBAR_6', 'MBAR_7', 'MBAR_8', 'MBAR_9',
    'MBAR_10', 'MBAR_11', 'BBAR_1', 'BBAR_2', 'BBAR_3', 'BBAR_4', 'BBAR_5', 'BBAR_6', 'BBAR_7', 'BBAR_8',
    'EBAR_1', 'EBAR_2', 'EBAR_3', 'TPPOL_1', 'TPPOL_2', 'TPPOL_3', 'TPPOL_4', 'TPPOL_5', 'TPPOL_6', 'TPPOL_7', 'TPPOL_8',
    'TPPOL_9', 'TPPOL_10', 'TPS_1', 'TPS_2', 'INC', 'EB_FILTR', 'VYD', 'POJ', 'TYP_1', 'TYP_2', 'TYP_3', 'TYP_4', 'TYP_5', 'TYP_6', 'TYP_7', 'TYP_8',
    'EQ_1_2', 'HEAT3_1_FIN', 'HEAT3_2_FIN', 'HEAT3_3_FIN', 'HEAT3_4_FIN', 'HEAT3_5_FIN', 'HEAT3_6_FIN', 'ETS2A', 'ETS2B', 'INC_SILC_EQ_KAT', 'HOUS_KAT',
    'BUF_KAT', 'PKM_KAT', 'PKM_NUM_SILC_KAT', 'CKMD_KAT', 'CKMD_NUM_SILC_KAT', 'BUIT_KAT', 'BUIT_NUM_SILC_KAT', 'CVAR_KAT', 'CVAR_NUM_SILC_KAT', 'MKM_KAT',
    'MKM_NUM_SILC_KAT', 'MKM_2_KAT', 'MKM_2_NUM_SILC_KAT', 'BUMT_KAT', 'BUMT_NUM_SILC_KAT', 'BKM_KAT', 'BKM_NUM_SILC_KAT', 'MKMM_KAT', 'MKMM_NUM_SILC_KAT',
    'KM_1_1_KAT', 'KM_1_2_KAT', 'KM_1_3_KAT', 'KM_1_4_KAT', 'KM_1_5_KAT', 'KM_2_1_KAT', 'KM_2_2_KAT', 'KM_2_3_KAT', 'KM_2_4_KAT', 'KM_2_5_KAT',
    'KM_3_1_KAT', 'KM_3_2_KAT', 'KM_3_3_KAT', 'KM_3_4_KAT', 'KM_3_5_KAT', 'KM_4_KAT', 'KM_5_KAT', 'KM_6_KAT', 'KM_7_KAT', 'KM_8_KAT', 'KM_9_KAT',
    'KM_MEAN_KAT', 'CARS_1_1_KAT', 'CARS_2_1_KAT', 'CARS_3_1_KAT', 'CARS_4_1_KAT', 'CARS_5_1_KAT', 'CARS_1_2_KAT', 'CARS_2_2_KAT', 'CARS_3_2_KAT',
    'CARS_4_2_KAT', 'CARS_5_2_KAT', 'CARS_1_3_KAT', 'CARS_2_3_KAT', 'CARS_3_3_KAT', 'CARS_4_3_KAT', 'CARS_5_3_KAT', 'CARS_1_6_KAT', 'CARS_2_6_KAT',
    'CARS_3_6_KAT', 'CARS_4_6_KAT', 'CARS_5_6_KAT'
]

cont_xs = [
    'NUM_1', 'NUM_2', 'NUM_3', 'EB_FIN', 'INC_SILC_EQ_FIN', 'HOUS', 'HOUS_NUM_SILC', 'BUF', 'BUF_NUM_SILC', 'MFAM', 'PKM', 'PKM_NUM_SILC', 'CKMD',
    'CKMD_NUM_SILC', 'BUIT', 'BUIT_NUM_SILC', 'CVAR', 'CVAR_NUM_SILC', 'MKM', 'MKM_NUM_SILC', 'MKM_2', 'MKM_2_NUM_SILC', 'BUMT', 'BUMT_NUM_SILC',
    'BKM', 'BKM_NUM_SILC', 'PKM_BKM_NUM_SILC', 'MKMM', 'MKMM_NUM_SILC', 'KM_1_1', 'KM_1_2', 'KM_1_3', 'KM_1_4', 'KM_1_5', 'KM_2_1', 'KM_2_2', 'KM_2_3',
    'KM_2_4', 'KM_2_5', 'KM_3_1', 'KM_3_2', 'KM_3_3', 'KM_3_4', 'KM_3_5', 'KM_4', 'KM_5', 'KM_6', 'KM_7', 'KM_8', 'KM_9', 'KM_MEAN', 'CARS1', 'CARS_1_1',
    'CARS_2_1', 'CARS_3_1', 'CARS_4_1', 'CARS_5_1', 'CARS_1_2', 'CARS_2_2', 'CARS_3_2', 'CARS_4_2', 'CARS_5_2', 'CARS_1_3', 'CARS_2_3', 'CARS_3_3',
    'CARS_4_3', 'CARS_5_3', 'CARS_1_6', 'CARS_2_6', 'CARS_3_6', 'CARS_4_6', 'CARS_5_6', 'CARS_1', 'CARS_3', 'CARS_4', 'CARS_5', 'CARS_6', 'CARS_7',
    'CARS_2', 'CARS_8', 'NAR_CLICON', 'NAR_GDTHREAT', 'NAR_ACTION', 'NAR_PROACT', 'POL_MEAN', 'POL_MISS', 'NAR_MISS', 'CZPROB_MEAN', 'SECT_MEAN',
    'BRAND_MEAN', 'OSV_MEAN', 'TRACT_MEAN', 'MEDT_SELF', 'MEDT_NEWS', 'MEDT_TECH', 'MEDT_CLICK', 'MEDT_ART', 'MEDT_GARD', 'PO_RUS', 'PO_PROSP',
    'PO_RIGHTS', 'IDENT_LIB', 'IDENT_RIGHT', 'TRP_AUT', 'TRP_DEM', 'ENE_RENEW', 'ENE_FOSS', 'ENE_VAR', 'NEED_SECU', 'NEED_ENVI', 'NEED_DEMO',
    'NEED_FREED', 'VOTE_COAL', 'VOTE_ANTI', 'VOTE_LIB', 'POL_CLIM', 'POL_EMOB'
]


w = df['weights']
agg_omitted_vars = ['POHLAVI', 'AGECAT', 'AGECAT2', 'EDU', 'EDUCAT', 'INT', 'ECS', 'KRAJ', 'OKR', 'OBEC', 'VMB', 'VMBCAT', 'SET']

ys = ["dc7_log"]

df[ys].value_counts()
df["target"] = (df[ys] == 7.0).astype(int)
y = df["target"]

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df[xs + cont_xs], y, w, test_size=0.2, random_state=1234, stratify=y
)

# Train XGBoost model
xgb_model = XGBClassifier(
    n_estimators=100,
    max_depth=3,
    learning_rate=0.1,
    objective='binary:logistic',
    n_jobs=-1,
    random_state=42,
    use_label_encoder=False,
    eval_metric='logloss'
)

xgb_model.fit(
    X_train.values,
    y_train.values,
    sample_weight=w_train.values,
    eval_set=[(X_test.values, y_test.values)],
    sample_weight_eval_set=[w_test.values],
    verbose=False
)

# Feature importance
importances = xgb_model.feature_importances_
importances
indices = np.argsort(importances)[::-1]
selected_features = X_train.columns[indices[:20]]
print(f"Selected features: {list(selected_features)}")

predictors = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'MKM_2_KAT', 'MEDT_5', 'TPN']

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df[selected_features], y, w, test_size=0.2, random_state=1234, stratify=y
)

# Train XGBoost model
xgb_model = XGBClassifier(
    n_estimators=100,
    max_depth=3,
    learning_rate=0.1,
    objective='binary:logistic',
    n_jobs=-1,
    random_state=42,
    use_label_encoder=False,
    eval_metric='logloss'
)

xgb_model.fit(
    X_train.values,
    y_train.values,
    sample_weight=w_train.values,
    eval_set=[(X_test.values, y_test.values)],
    sample_weight_eval_set=[w_test.values],
    verbose=False
)

# Feature importance
importances = xgb_model.feature_importances_
importances
indices = np.argsort(importances)[::-1]
selected_features = X_train.columns[indices[:10]]
print(f"Selected features: {list(selected_features)}")

predictors = ['TPZ', 'TPN', 'MKM_2_KAT', 'KM_7_KAT', 'KM_9_KAT', 'MKM_NUM_SILC', 'MEDT_5', 'CKMD_NUM_SILC', 'CVAR_NUM_SILC', 'HEAT4']

data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-11-14.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file, user_missing=True)
filt_var = "filter_$"
df = df[df[filt_var] == 1].copy()

xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'MKM_2_KAT', 'MEDT_5', 'TPN']
w = df['weights']
df["target"] = (df["dc7_log"] == 7.0).astype(int)
y = df["target"]

# fix variables
df["CPRE_2"].value_counts(dropna=False)
df["CPRE_2_FIX"] = df["CPRE_2"].replace({3.0: 4.0, 4.0: 5.0}).fillna(3.0)
df["CPRE_2_FIX"].value_counts(dropna=False)

df["KM_7_KAT_MISS"] = df["KM_7_KAT"].isna().astype(int)
df["KM_7_KAT_FIX"] = df["KM_7_KAT"].fillna(df["KM_7_KAT"].median())
df["MKM_2_KAT_MISS"] = df["MKM_2_KAT"].isna().astype(int)
df["MKM_2_KAT_FIX"] = df["MKM_2_KAT"].fillna(df["MKM_2_KAT"].median())

xs = ['CPRE_2_FIX', 'TPZ', 'KM_7_KAT_FIX', 'MKM_2_KAT_FIX', 'MEDT_5',
      'TPN', 'KM_7_KAT_MISS', 'MKM_2_KAT_MISS']
xs = ['CPRE_2_FIX', 'TPZ', 'KM_7_KAT_FIX', 'MKM_2_KAT_FIX', 'KM_7_KAT_MISS', 'MKM_2_KAT_MISS']

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df[xs], y, w, test_size=0.2, random_state=1234, stratify=y
)

log_reg = LogisticRegression(solver="liblinear", max_iter=5000)
log_reg.fit(X_train, y_train, sample_weight=w_train)

coefficients = pd.DataFrame(log_reg.coef_, columns=xs)
print(coefficients)

# Evaluate the model
y_pred_log = log_reg.predict(X_train)
accuracy_log = accuracy_score(y_train, y_pred_log)
print(f'Accuracy of logistic regression: {accuracy_log:.4f}')

pd.DataFrame({'true': y_train, 'pred': y_pred_log}).value_counts().unstack().fillna(0).astype(int)


predictors = ['TPZ', 'TPN', 'MKM_2_KAT', 'KM_7_KAT', 'KM_9_KAT', 'MKM_NUM_SILC', 'MEDT_5', 'CKMD_NUM_SILC', 'CVAR_NUM_SILC', 'HEAT4']

xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'MKM_2_KAT', 'MEDT_5', 'TPN']
xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'MKM_2_KAT']
xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'TPN']
xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'MEDT_5']
xs = ['TPZ', 'TPN', 'MKM_2_KAT', 'KM_7_KAT', 'KM_9_KAT', 'MKM_NUM_SILC', 'MEDT_5', 'CKMD_NUM_SILC', 'CVAR_NUM_SILC', 'HEAT4']
xs = ['TPZ', 'TPN', 'KM_7_KAT', 'KM_9_KAT']
xs = ['TPZ', 'TPN', 'MKM_2_KAT', 'KM_7_KAT']
w = df['weights']
df["target"] = (df["dc7_log"] == 7.0).astype(int)
y = df["target"]

df2 = df.dropna(subset=xs)
df2.shape
df2["target"].value_counts()
w = df2['weights']
y = df2["target"]


X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df2[xs], y, w, test_size=0.2, random_state=42, stratify=y
)

log_reg = LogisticRegression(solver="liblinear", max_iter=5000)
log_reg.fit(X_train, y_train, sample_weight=w_train)

coefficients = pd.DataFrame(log_reg.coef_, columns=xs)
print(coefficients)
print(log_reg.intercept_)

# In [190]: print(coefficients)
#        TPZ       TPN  MKM_2_KAT  KM_7_KAT
# 0  1.03901  0.733066    0.02172  0.268105
#
# In [191]: print(log_reg.intercept_)
# [-8.1525627]

def sigmoid(x):
    return 1 / (1 + np.exp(-x))

sigmoid(-0.2)

# Evaluate the model
y_pred_log = log_reg.predict(X_test)
accuracy_log = accuracy_score(y_test, y_pred_log)
print(f'Accuracy of logistic regression: {accuracy_log:.4f}')

pd.DataFrame({'true': y_test, 'pred': y_pred_log}).value_counts().unstack().fillna(0).astype(int)

# This set is ok!
xs = ['TPZ', 'TPN', 'MKM_2_KAT', 'KM_7_KAT']
# xs = ['CPRE_2', 'TPZ', 'KM_7_KAT', 'TPN']


for i, x in enumerate(xs):
    print()
    print(i + 1, x, df_meta.column_names_to_labels[x])
    print(f"{x}: {df_meta.variable_value_labels[x]}")
    print(df[x].value_counts(dropna=False))
    print()

for i, x in enumerate(xs):
    print(x, df_meta.column_names_to_labels[x])


# Function to select features using Logistic Regression with L1 penalty
def select_features(X_train, y_train, weights_train, num_features=4):
    # Try different values of regularization strength C
    Cs = np.logspace(-4, 4, 100)
    for C in Cs:
        clf = LogisticRegression(
            penalty='l1',
            C=C,
            solver='liblinear',
            max_iter=1000
        )
        clf.fit(X_train, y_train, sample_weight=weights_train)
        coef = clf.coef_.ravel()
        num_nonzero = np.sum(np.abs(coef) > 1e-5)
        if num_nonzero <= num_features:
            selected_features = X_train.columns[np.abs(coef) > 1e-5]
            print(f'C={C}, selected features: {list(selected_features)}')
            return selected_features, clf
    # If not found, return the model with smallest number of features
    clf = LogisticRegression(
        penalty='l1',
        C=Cs[-1],
        solver='liblinear',
        max_iter=1000
    )
    clf.fit(X_train, y_train, sample_weight=weights_train)
    coef = clf.coef_.ravel()
    selected_features = X_train.columns[np.abs(coef) > 1e-5]
    return selected_features, clf

# Select features and get the classifier
selected_features, clf = select_features(X_train, y_train, w_train, num_features=4)

# Use only the selected features
X_train_selected = X_train[selected_features]
X_test_selected = X_test[selected_features]

# Refit the model on selected features
clf.fit(X_train_selected, y_train, sample_weight=weights_train)

# Predict on test set
y_pred = clf.predict(X_test_selected)
y_pred_proba = clf.predict_proba(X_test_selected)[:, 1]

# Evaluate performance
accuracy = accuracy_score(y_test, y_pred, sample_weight=weights_test)
roc_auc = roc_auc_score(y_test, y_pred_proba, sample_weight=weights_test)
conf_matrix = confusion_matrix(y_test, y_pred, sample_weight=weights_test)
class_report = classification_report(y_test, y_pred, sample_weight=weights_test)

print(f'Accuracy: {accuracy}')
print(f'ROC AUC: {roc_auc}')
print('Confusion Matrix:')
print(conf_matrix)
print('Classification Report:')
print(class_report)





