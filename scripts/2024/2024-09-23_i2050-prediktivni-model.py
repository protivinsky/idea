import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *

data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-09-10.sav"

# df, df_meta = pyreadstat.read_sav(data_root / data_file, user_missing=True)
df, df_meta = pyreadstat.read_sav(data_root / data_file)

df["BELIEF"].value_counts(dropna=False)
df["KNOW_2"].value_counts(dropna=False)
df.shape
first_idx = df.columns.get_loc("ENV2")
last_idx = df.columns.get_loc("NAR_46") + 1
feat_cols = df.columns[first_idx:last_idx]
resp_col = "c7_max_alt"
w_col = "weights"

cat_columns = [c for c in df.columns if c.startswith("KNOW_") and len(c) == 6]
cat_columns

for c in cat_columns:
    df[c] = df[c].fillna(0).astype(int)

x_cols = [c for c in feat_cols if c not in cat_columns]


# Split data into features and target
df.value_counts(resp_col, dropna=False)
df = df.dropna(subset=[resp_col])
y = df[resp_col].astype('category').cat.codes
weights = df[w_col]

# Split into training and test sets
X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df[feat_cols], y, weights, test_size=0.2, random_state=42, stratify=y
)

# Create LightGBM datasets
train_data = lgb.Dataset(X_train, label=y_train, weight=w_train, categorical_feature=cat_columns)
test_data = lgb.Dataset(X_test, label=y_test, weight=w_test, categorical_feature=cat_columns)

# Set model parameters
params = {
    'objective': 'multiclass',
    'num_class': 7,  # Number of classes
    'metric': 'multi_logloss',
    'verbosity': -1,
    'seed': 42
}

# Train the model
model = lgb.train(
    params,
    train_data,
    valid_sets=[test_data],
    num_boost_round=1000,
)

model

# Make predictions
y_pred = model.predict(X_test)
y_pred_classes = y_pred.argmax(axis=1)

pd.DataFrame({'true': y_test, 'pred': y_pred_classes}).value_counts().unstack().fillna(0)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred_classes)
print(f'Accuracy: {accuracy:.4f}')

# Feature importance
importance_df = pd.DataFrame({
    'feature': model.feature_name(),
    'importance': model.feature_importance()
}).sort_values(by='importance', ascending=False)

importance_df

# Select top 20 features
top_features = importance_df.head(20)['feature'].tolist()
top_features

# Retrain the model with top features
X_train_top = X_train[top_features]
X_test_top = X_test[top_features]

train_data_top = lgb.Dataset(X_train_top, label=y_train, weight=w_train)
test_data_top = lgb.Dataset(X_test_top, label=y_test, weight=w_test)

model_top = lgb.train(
    params,
    train_data_top,
    valid_sets=[test_data_top],
    num_boost_round=1000,
)

# Evaluate the new model
y_pred_top = model_top.predict(X_test_top)
y_pred_classes_top = y_pred_top.argmax(axis=1)
accuracy_top = accuracy_score(y_test, y_pred_classes_top)
print(f'Accuracy with top features: {accuracy_top:.4f}')

pd.DataFrame({'true': y_test, 'pred': y_pred_classes_top}).value_counts().unstack().fillna(0)

# so this is doable and I am getting sensible model out.
# TODO:
# - use only GDF, POL_X, NAR_X
# - try it also with missing flag: GDF as 3 and move the rest, others as 3 w/o shift
# Options + accuracies:
# - full model: 0.84
# - top 20, no missing: 0.80
# - top 20, w missing: 0.78
# - tree: 0.61
# - log reg, top 20, w missing: 0.82
# - log reg, top 20, no missing: 0.70
# - log reg, top 15, w missing: 0.79
# - log reg, top 15, no missing: 0.72


starts = ["GDF", "POL_", "NAR_"]
sel_cols = [c for c in feat_cols if any(c.startswith(s) for s in starts)]
sel_cols

df2 = df[sel_cols + [resp_col, w_col]].copy()
for c in sel_cols:
    df2[c + "_missing"] = df2[c].isna().astype(int)
df2["GDF"].value_counts(dropna=False)
df2["GDF"] = df2["GDF"].replace({3: 4, 4: 5, np.nan: 3})
for c in sel_cols[1:]:
    df2[c] = df2[c].replace({np.nan: 3})

x_cols = [c for c in df2.columns if c not in [resp_col, w_col]]
x_cols

y = df2[resp_col].astype('category').cat.codes
weights = df2[w_col]

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df2[x_cols], y, weights, test_size=0.2, random_state=42, stratify=y
)

# Create LightGBM datasets
train_data = lgb.Dataset(X_train, label=y_train, weight=w_train)
test_data = lgb.Dataset(X_test, label=y_test, weight=w_test)

# Set model parameters
params = {
    'objective': 'multiclass',
    'num_class': 7,  # Number of classes
    'metric': 'multi_logloss',
    'verbosity': -1,
    'seed': 42
}

# Train the model
model = lgb.train(
    params,
    train_data,
    valid_sets=[test_data],
    num_boost_round=1000,
)

model

# Make predictions
y_pred = model.predict(X_test)
y_pred_classes = y_pred.argmax(axis=1)

pd.DataFrame({'true': y_test, 'pred': y_pred_classes}).value_counts().unstack().fillna(0)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred_classes)
print(f'Accuracy: {accuracy:.4f}')

# Feature importance
importance_df = pd.DataFrame({
    'feature': model.feature_name(),
    'importance': model.feature_importance()
}).sort_values(by='importance', ascending=False)

importance_df

# Select top 20 features
top_features = importance_df.head(15)['feature'].tolist()

top_features_w_missing = top_features + [f + "_missing" for f in top_features]
top_features_w_missing

from sklearn.tree import DecisionTreeClassifier, export_text

# Train a decision tree with limited depth
tree_clf = DecisionTreeClassifier(max_depth=10, random_state=42)
tree_clf.fit(X_train[top_features_w_missing], y_train, sample_weight=w_train)

# Visualize the tree
tree_rules = export_text(tree_clf, feature_names=top_features_w_missing)
print(tree_rules)

# Evaluate the model
y_pred_tree = tree_clf.predict(X_test[top_features_w_missing])
accuracy_tree = accuracy_score(y_test, y_pred_tree)
print(f'Accuracy of decision tree: {accuracy_tree:.4f}')

pd.DataFrame({'true': y_test, 'pred': y_pred_tree}).value_counts().unstack().fillna(0)

starts = ["GDF", "POL_", "NAR_"]
sel_cols = [c for c in feat_cols if any(c.startswith(s) for s in starts)]

df2 = df[sel_cols + [resp_col, w_col]].copy()
# for c in sel_cols:
#     df2[c + "_missing"] = df2[c].isna().astype(int)
# df2["GDF"].value_counts(dropna=False)
# df2["GDF"] = df2["GDF"].replace({3: 4, 4: 5, np.nan: 3})
# for c in sel_cols[1:]:
#     df2[c] = df2[c].replace({np.nan: 3})

x_cols = [c for c in df2.columns if c not in [resp_col, w_col]]

y = df2[resp_col].astype('category').cat.codes
weights = df2[w_col]

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df2[x_cols], y, weights, test_size=0.2, random_state=42, stratify=y
)

# Create LightGBM datasets
train_data = lgb.Dataset(X_train, label=y_train, weight=w_train)
test_data = lgb.Dataset(X_test, label=y_test, weight=w_test)

# Set model parameters
params = {
    'objective': 'multiclass',
    'num_class': 7,  # Number of classes
    'metric': 'multi_logloss',
    'verbosity': -1,
    'seed': 42
}

# Train the model
model = lgb.train(
    params,
    train_data,
    valid_sets=[test_data],
    num_boost_round=1000,
)

# Make predictions
y_pred = model.predict(X_test)
y_pred_classes = y_pred.argmax(axis=1)

pd.DataFrame({'true': y_test, 'pred': y_pred_classes}).value_counts().unstack().fillna(0)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred_classes)
print(f'Accuracy: {accuracy:.4f}')

# Feature importance
importance_df = pd.DataFrame({
    'feature': model.feature_name(),
    'importance': model.feature_importance()
}).sort_values(by='importance', ascending=False)

importance_df
importance_df

# Select top 20 features
top_features = importance_df.head(20)['feature'].tolist()


# Retrain the model with top features
X_train_top = X_train[top_features]
X_test_top = X_test[top_features]

train_data_top = lgb.Dataset(X_train_top, label=y_train, weight=w_train)
test_data_top = lgb.Dataset(X_test_top, label=y_test, weight=w_test)

model_top = lgb.train(
    params,
    train_data_top,
    valid_sets=[test_data_top],
    num_boost_round=1000,
)

# Evaluate the new model
y_pred_top = model_top.predict(X_test_top)
y_pred_classes_top = y_pred_top.argmax(axis=1)
accuracy_top = accuracy_score(y_test, y_pred_classes_top)
print(f'Accuracy with top features: {accuracy_top:.4f}')

pd.DataFrame({'true': y_test, 'pred': y_pred_classes_top}).value_counts().unstack().fillna(0)


from sklearn.linear_model import LogisticRegression

df2 = df[sel_cols + [resp_col, w_col]].copy()
for c in sel_cols:
    df2[c + "_missing"] = df2[c].isna().astype(int)
df2["GDF"].value_counts(dropna=False)
df2["GDF"] = df2["GDF"].replace({np.nan: 2})
for c in sel_cols[1:]:
    df2[c] = df2[c].replace({np.nan: 3})

x_cols = [c for c in df2.columns if c not in [resp_col, w_col]]

y = df2[resp_col].astype('category').cat.codes
weights = df2[w_col]

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df2[x_cols], y, weights, test_size=0.2, random_state=1234, stratify=y
)
top_features_w_missing = top_features + [f + "_missing" for f in top_features]
# len(top_features_w_missing)

# Train logistic regression
log_reg = LogisticRegression(multi_class='multinomial', solver='lbfgs', max_iter=1000)
log_reg.fit(X_train[top_features_w_missing], y_train, sample_weight=w_train)

# Coefficients
coefficients = pd.DataFrame(log_reg.coef_, columns=top_features_w_missing)
print(coefficients)

# Evaluate the model
y_pred_log = log_reg.predict(X_test[top_features_w_missing])
accuracy_log = accuracy_score(y_test, y_pred_log)
print(f'Accuracy of logistic regression: {accuracy_log:.4f}')

pd.DataFrame({'true': y_test, 'pred': y_pred_log}).value_counts().unstack().fillna(0).astype(int)

# Train logistic regression
log_reg = LogisticRegression(multi_class='multinomial', solver='lbfgs', max_iter=1000)
log_reg.fit(X_train[top_features], y_train, sample_weight=w_train)

# Coefficients
coefficients = pd.DataFrame(log_reg.coef_, columns=top_features)
print(coefficients)

# Evaluate the model
y_pred_log = log_reg.predict(X_test[top_features])
accuracy_log = accuracy_score(y_test, y_pred_log)
print(f'Accuracy of logistic regression: {accuracy_log:.4f}')

np.exp(log_reg.intercept_ + log_reg.coef_ @ )

nom = np.exp(log_reg.intercept_ + np.dot(X_test[top_features_w_missing], log_reg.coef_.T))
den = np.sum(np.exp(log_reg.intercept_ + np.dot(X_test[top_features_w_missing], log_reg.coef_.T)), axis=1)

my_pred = np.argmax(nom / den[:, None], axis=1)
y_pred_log

np.ex


log_reg.coef_.shape
log_reg.intercept_
X_test[top_features].


