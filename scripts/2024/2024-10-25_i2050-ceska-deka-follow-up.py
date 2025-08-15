import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.linear_model import LogisticRegression
import reportree as rt
from omoment import OMeanVar, OMean, OCov
from libs.extensions import *

data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-09-10.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file)
w_col = "weights"


df.groupby(["GDF", "PO_20"])[w_col].sum().unstack()
df.groupby("GDF")[w_col].sum()
df.groupby("PO_20")[w_col].sum()

df["green_deal"] = df["GDF"].replace({np.nan: 2})

foo = df[df["GDF"].isin([1, 3, 4])].copy()
foo["green_deal"] = foo["GDF"].replace({1: 0, 3: 1, 4: 1})
foo["kapitulace"] = (foo["PO_20"] > 3).astype(int)

np.average(foo["kapitulace"], weights=foo[w_col])

OMean.of_groupby(data=foo, g="green_deal", x="kapitulace" , w=w_col)

np.average((df["PO_20"] > 3).astype(int), weights=df[w_col])

# TODO: Calculate the entropy and verify its correctness
data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
c7, c7_meta = pyreadstat.read_sav(data_root / "lca_main" / "c07.sav")
w_col = "weights"

c7.columns
n_classes = 7
probs_cols = [f"clu#{i}" for i in range(1, 8)]

# entropy
c7["entropy"] = 0
for c in probs_cols:
    c7["entropy"] -= c7[c] * np.log(c7[c])

c7["entropy"].describe()

1 - (c7["entropy"] * c7[w_col]).sum() / (np.log(n_classes) * c7[w_col].sum())
1 - c7["entropy"].sum() / (np.log(n_classes) * len(c7))

len(c7)


data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-09-10.sav"

selected_nar = [30, 11, 7, 19, 10, 15, 13, 6, 16, 3, 14, 2, 39, 24]
selected = ["GDF"] + [f"NAR_{nar}" for nar in selected_nar]

df, df_meta = pyreadstat.read_sav(data_root / data_file, user_missing=True)
resp_col = "c7_max_alt"
w_col = "weights"

df = df.dropna(subset=[resp_col]).copy()
df2 = df.copy()
for c in selected:
    df2[c + "_missing"] = (df2[c].isna() | (df2[c] == 99)).astype(int)
df2["GDF"] = df2["GDF"].replace({np.nan: 2, 99: 2})
for c in selected:
    df2[c] = df2[c].replace({np.nan: 3, 99: 3})
selected_missing = [c + "_missing" for c in selected]

y = df2[resp_col].astype('category').cat.codes
weights = df2[w_col]

np.sum(~np.isfinite(df2[w_col]))


selected_missing

ocov = OCov.of_frame(data=df2, x=selected + selected_missing, w=w_col)
ocov.cov

ocov.corr
eigenvalues, eigenvectors = np.linalg.eig(ocov.corr)
# Sort the eigenvalues in descending order
sorted_indices = np.argsort(eigenvalues)[::-1]
sorted_eigenvalues = eigenvalues[sorted_indices]
sorted_eigenvectors = eigenvectors[:, sorted_indices]

eigenvalues


sorted_eigenvalues / sorted_eigenvalues.sum()

df3 = (df2[selected + selected_missing].copy() - ocov.mean) / np.sqrt(ocov.cov.diagonal())

proj = pd.DataFrame(df3.values @ eigenvectors, columns=[f"comp_{i}" for i in range(1, 31)])
proj[resp_col] = df2[resp_col].astype('category').cat.codes.values
proj[w_col] = df2[w_col].values

class_labels = {
    0: "Vlažní",
    1: "Sympatizující",
    2: "Odmítající",
    3: "Nejistí",
    4: "Angažovaní pro Zelenou dohodu",
    5: "Angažovaní bez Zelené dohody",
    6: "Popírající",
}

proj["class"] = proj[resp_col].map(class_labels)
proj

title = "PCA projection of the data"
doc = rt.Doc(title=title)

for i in range(2, 10):
    fig, ax = plt.subplots(figsize=(10, 5))
    sns.scatterplot(data=proj, x="comp_3", y=f"comp_{i}", hue="class")
    doc.figure_as_b64(fig)

doc.show()


proj.columns
proj[proj.columns[:30]].cov().values.diagonal()
eigenvalues
foo = OCov.of_frame(data=proj, x=proj.columns[:30], w=w_col)
foo.cov.diagonal()
eigenvalues
np.sum(np.isfinite(proj), axis=0)
df2.shape
proj.shape
proj[w_col]

foo.cov




df3[w_col] = df2[w_col]
foo = OCov.of_frame(data=df3, x=selected + selected_missing, w=w_col)
foo.mean
foo.cov.diagonal()

df3 / np.sqrt(ocov.cov.diagonal())


ocov2 = OCov.of_frame(data=df2, x=selected + selected_missing)
eigenvalues, eigenvectors = np.linalg.eig(ocov2.corr)
# Sort the eigenvalues in descending order
sorted_indices = np.argsort(eigenvalues)[::-1]
sorted_eigenvalues = eigenvalues[sorted_indices]
sorted_eigenvectors = eigenvectors[:, sorted_indices]

sorted_eigenvalues / sorted_eigenvalues.sum()





from sklearn.decomposition import PCA
from sklearn.preprocessing import StandardScaler
import sklearn

sklearn.__version__

X = df2[selected + selected_missing]
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)
w = df2[w_col].values

pca = PCA()
pca.fit(X)
# unscaled - main component is about 50 % of variance
pca.explained_variance_ratio_

pca = PCA()
pca.fit(X_scaled)
# scaled - main component is 27 % of variance
# scaled version should be generally preferable for PCA
pca.explained_variance_ratio_

import plotext as pltx

pltx.plotsize(100, 30)
pltx.theme('clear')

pltx.bar(pca.explained_variance_ratio_)
pltx.show()

