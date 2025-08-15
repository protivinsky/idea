import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *


# NOTE: calculate the colored and tabulated data
# =====================================================

data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-11-14.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file, user_missing=True)
filt_var = "filter_$"
df = df[df[filt_var] == 1].copy()

xs = [
    'POHLAVI', 'AGECAT', 'AGECAT2', 'EDU', 'EDUCAT', 'INT', 'ECS', 'KRAJ', 'OKR', 'OBEC', 'VMB', 'VMBCAT', 'SET', 'NUM_1',
    'NUM_2', 'NUM_3', 'MEDT_1', 'MEDT_2', 'MEDT_3', 'MEDT_4', 'MEDT_5', 'MEDT_6', 'MEDT_7', 'MEDT_8', 'MEDT_9', 'MEDT_10', 'MEDT_11', 'MEDT_12',
    'MEDT_13', 'MEDT_14', 'MEDT_15', 'MEDT_16', 'MEDT_17', 'MEDT_18', 'MEDT_19', 'MEDT_20', 'MEDT_21', 'MEDT_22', 'MEDT_23', 'MEDT_24',
    'MEDT_25', 'NEED_1', 'NEED_2', 'NEED_3', 'NEED_4', 'NEED_5', 'NEED_6', 'NEED_7', 'NEED_8', 'NEED_9', 'NEED_10', 'NEED_11', 'NEED_12',
    'NEED_13', 'NEED_14', 'NEED_15', 'NEED_16', 'NEED_17', 'NEED_18', 'NEED_19', 'NEED_20', 'NEED_21', 'NEED_22', 'NEED_23', 'NEED_24',
    'PO_1', 'PO_2', 'PO_3', 'PO_4', 'PO_5', 'PO_6', 'PO_7', 'PO_8', 'PO_9', 'PO_10', 'PO_11', 'PO_12', 'PO_13', 'PO_14', 'PO_15', 'PO_16', 'PO_17', 'PO_18',
    'PO_19', 'PO_20', 'IDENT_1', 'IDENT_2', 'IDENT_3', 'IDENT_4', 'IDENT_5', 'IDENT_6', 'IDENT_7', 'IDENT_8', 'IDENT_9', 'EL', 'EL21', 'ELECT',
    'ELECT2', 'VOTE_1', 'VOTE_2', 'VOTE_3', 'VOTE_4', 'VOTE_5', 'VOTE_6', 'VOTE_7', 'VOTE_8', 'VOTE_9', 'VOTE_10', 'VOTE_11', 'VOTE_12',
    'VOTE_13', 'VOTE_14', 'VOTE_15', 'TRP_1', 'TRP_2', 'TRP_3', 'TRP_4', 'TRP_5', 'TRP_6', 'TRP_7', 'TRP_8', 'TRP_9', 'TRP_10', 'NOST_1', 'NOST_2',
    'NOST_3', 'NOST_4', 'NOST_5', 'ENV2', 'BELIEF', 'KNOW_1', 'KNOW_2', 'KNOW_3', 'KNOW_4', 'KNOW_5', 'KNOW_6', 'CZPROB_1', 'CZPROB_2',
    'CZPROB_3', 'CZPROB_4', 'CZPROB_5', 'URGB', 'URGB_REC', 'SECT_1', 'SECT_2', 'SECT_3', 'SECT_4', 'SECT_5', 'ENE_1', 'ENE_2', 'ENE_3', 'ENE_4',
    'ENE_5', 'ENE_6', 'ENE_7', 'ENE_8', 'ENC', 'VE_A', 'VE_B', 'VE_C', 'POL_1', 'POL_2', 'POL_3', 'POL_4', 'POL_5', 'POL_6', 'POL_7', 'POL_8', 'POL_9',
    'POL_10', 'BRAND_1', 'BRAND_2', 'BRAND_3', 'BRAND_4', 'BRAND_5', 'BRAND_6', 'OSV_1', 'OSV_2', 'OSV_3', 'OSV_4', 'OSV_5', 'OSV_6', 'OSV_7',
    'OSV_8', 'GDF', 'GDGD_GD', 'GDGD_ZD', 'GDGD_N', 'GDGD_OK', 'GDGD_ZM', 'GDGD_ST', 'GDGD_U', 'TRACT_1', 'TRACT_2', 'TRACT_3', 'TRACT_4',
    'TRACT_5', 'TRACT_6', 'TRACT_7', 'TRACT_8', 'TRACT_9', 'TRACT_10', 'TRACT_11', 'TRACT_12', 'NAR_1', 'NAR_2', 'NAR_3', 'NAR_4', 'NAR_5',
    'NAR_6', 'NAR_7', 'NAR_8', 'NAR_9', 'NAR_10', 'NAR_11', 'NAR_12', 'NAR_13', 'NAR_14', 'NAR_15', 'NAR_16', 'NAR_17', 'NAR_18', 'NAR_19',
    'NAR_20', 'NAR_21', 'NAR_22', 'NAR_23', 'NAR_24', 'NAR_25', 'NAR_26', 'NAR_27', 'NAR_28', 'NAR_29', 'NAR_30', 'NAR_31', 'NAR_32', 'NAR_33',
    'NAR_34', 'NAR_35', 'NAR_36', 'NAR_37', 'NAR_38', 'NAR_39', 'NAR_40', 'NAR_41', 'NAR_42', 'NAR_43', 'NAR_44', 'NAR_45', 'NAR_46', 'EQ_1',
    'EQ_2', 'EQ_3', 'EQ_4', 'EQ_5', 'EQ_6', 'EQ_7', 'EQ_8', 'EQ_9', 'EQ_10', 'EQ_11', 'EQ_12', 'EQ_13', 'INS', 'MOB_1', 'MOB_2', 'MOB_3', 'MOB_4', 'MOB_5',
    'MOB_6', 'MOB_7', 'MOB_8', 'MOB_9', 'HEAT1', 'HEAT2_1', 'HEAT2_2', 'HEAT2_3', 'HEAT2_4', 'HEAT2_5', 'HEAT2_6', 'HEAT3_1', 'HEAT3_2',
    'HEAT3_3', 'HEAT3_4', 'HEAT3_5', 'HEAT3_6', 'HEAT4', 'TPZ', 'TPN', 'BEV_1', 'BEV_2', 'BEV_3', 'BEV_4', 'BEV_5', 'BEV_6', 'BEV_7', 'BEV_8', 'BEV_9',
    'BEV_10', 'BEV_DV', 'CARS_1_4', 'CARS_1_5', 'CARS_2_4', 'CARS_2_5', 'CARS_3_4', 'CARS_3_5', 'CARS_4_4', 'CARS_4_5', 'CARS_5_4',
    'CARS_5_5', 'CPRE_1', 'CPRE_2', 'CPRE_3', 'MBAR_1', 'MBAR_2', 'MBAR_3', 'MBAR_4', 'MBAR_5', 'MBAR_6', 'MBAR_7', 'MBAR_8', 'MBAR_9',
    'MBAR_10', 'MBAR_11', 'BBAR_1', 'BBAR_2', 'BBAR_3', 'BBAR_4', 'BBAR_5', 'BBAR_6', 'BBAR_7', 'BBAR_8',
    'EBAR_1', 'EBAR_2', 'EBAR_3', 'TPPOL_1', 'TPPOL_2', 'TPPOL_3', 'TPPOL_4', 'TPPOL_5', 'TPPOL_6', 'TPPOL_7', 'TPPOL_8',
    'TPPOL_9', 'TPPOL_10', 'TPS_1', 'TPS_2', 'INC', 'EB_FILTR', 'VYD', 'POJ', 'TYP_1', 'TYP_2', 'TYP_3', 'TYP_4', 'TYP_5', 'TYP_6', 'TYP_7', 'TYP_8',
    'EQ_1_2', 'HEAT3_1_FIN', 'HEAT3_2_FIN', 'HEAT3_3_FIN', 'HEAT3_4_FIN', 'HEAT3_5_FIN', 'HEAT3_6_FIN', 'ETS2A', 'ETS2B', 'INC_SILC_EQ_KAT', 'HOUS_KAT',
    'BUF_KAT', 'PKM_KAT', 'PKM_NUM_SILC_KAT', 'CKMD_KAT', 'CKMD_NUM_SILC_KAT', 'BUIT_KAT', 'BUIT_NUM_SILC_KAT', 'CVAR_KAT', 'CVAR_NUM_SILC_KAT', 'MKM_KAT',
    'MKM_NUM_SILC_KAT', 'MKM_2_KAT', 'MKM_2_NUM_SILC_KAT', 'BUMT_KAT', 'BUMT_NUM_SILC_KAT', 'BKM_KAT', 'BKM_NUM_SILC_KAT', 'MKMM_KAT', 'MKMM_NUM_SILC_KAT',
    'KM_1_1_KAT', 'KM_1_2_KAT', 'KM_1_3_KAT', 'KM_1_4_KAT', 'KM_1_5_KAT', 'KM_2_1_KAT', 'KM_2_2_KAT', 'KM_2_3_KAT', 'KM_2_4_KAT', 'KM_2_5_KAT',
    'KM_3_1_KAT', 'KM_3_2_KAT', 'KM_3_3_KAT', 'KM_3_4_KAT', 'KM_3_5_KAT', 'KM_4_KAT', 'KM_5_KAT', 'KM_6_KAT', 'KM_7_KAT', 'KM_8_KAT', 'KM_9_KAT',
    'KM_MEAN_KAT', 'CARS_1_1_KAT', 'CARS_2_1_KAT', 'CARS_3_1_KAT', 'CARS_4_1_KAT', 'CARS_5_1_KAT', 'CARS_1_2_KAT', 'CARS_2_2_KAT', 'CARS_3_2_KAT',
    'CARS_4_2_KAT', 'CARS_5_2_KAT', 'CARS_1_3_KAT', 'CARS_2_3_KAT', 'CARS_3_3_KAT', 'CARS_4_3_KAT', 'CARS_5_3_KAT', 'CARS_1_6_KAT', 'CARS_2_6_KAT',
    'CARS_3_6_KAT', 'CARS_4_6_KAT', 'CARS_5_6_KAT'
]

cont_xs = [
    'NUM_1', 'NUM_2', 'NUM_3', 'EB_FIN', 'INC_SILC_EQ_FIN', 'HOUS', 'HOUS_NUM_SILC', 'BUF', 'BUF_NUM_SILC', 'MFAM', 'PKM', 'PKM_NUM_SILC', 'CKMD',
    'CKMD_NUM_SILC', 'BUIT', 'BUIT_NUM_SILC', 'CVAR', 'CVAR_NUM_SILC', 'MKM', 'MKM_NUM_SILC', 'MKM_2', 'MKM_2_NUM_SILC', 'BUMT', 'BUMT_NUM_SILC',
    'BKM', 'BKM_NUM_SILC', 'PKM_BKM_NUM_SILC', 'MKMM', 'MKMM_NUM_SILC', 'KM_1_1', 'KM_1_2', 'KM_1_3', 'KM_1_4', 'KM_1_5', 'KM_2_1', 'KM_2_2', 'KM_2_3',
    'KM_2_4', 'KM_2_5', 'KM_3_1', 'KM_3_2', 'KM_3_3', 'KM_3_4', 'KM_3_5', 'KM_4', 'KM_5', 'KM_6', 'KM_7', 'KM_8', 'KM_9', 'KM_MEAN', 'CARS1', 'CARS_1_1',
    'CARS_2_1', 'CARS_3_1', 'CARS_4_1', 'CARS_5_1', 'CARS_1_2', 'CARS_2_2', 'CARS_3_2', 'CARS_4_2', 'CARS_5_2', 'CARS_1_3', 'CARS_2_3', 'CARS_3_3',
    'CARS_4_3', 'CARS_5_3', 'CARS_1_6', 'CARS_2_6', 'CARS_3_6', 'CARS_4_6', 'CARS_5_6', 'CARS_1', 'CARS_3', 'CARS_4', 'CARS_5', 'CARS_6', 'CARS_7',
    'CARS_2', 'CARS_8', 'NAR_CLICON', 'NAR_GDTHREAT', 'NAR_ACTION', 'NAR_PROACT', 'POL_MEAN', 'POL_MISS', 'NAR_MISS', 'CZPROB_MEAN', 'SECT_MEAN',
    'BRAND_MEAN', 'OSV_MEAN', 'TRACT_MEAN', 'MEDT_SELF', 'MEDT_NEWS', 'MEDT_TECH', 'MEDT_CLICK', 'MEDT_ART', 'MEDT_GARD', 'PO_RUS', 'PO_PROSP',
    'PO_RIGHTS', 'IDENT_LIB', 'IDENT_RIGHT', 'TRP_AUT', 'TRP_DEM', 'ENE_RENEW', 'ENE_FOSS', 'ENE_VAR', 'NEED_SECU', 'NEED_ENVI', 'NEED_DEMO',
    'NEED_FREED', 'VOTE_COAL', 'VOTE_ANTI', 'VOTE_LIB', 'POL_CLIM', 'POL_EMOB'
]


w = 'weights'
agg_omitted_vars = ['POHLAVI', 'AGECAT', 'AGECAT2', 'EDU', 'EDUCAT', 'INT', 'ECS', 'KRAJ', 'OKR', 'OBEC', 'VMB', 'VMBCAT', 'SET']

ys = ["dc7_log"]

# for variant in ['log', 'cap']:
#     for n in range(2, 10):
#         cl, _ = pyreadstat.read_sav(f'{data_root}/doprava_lca_C_{variant}_z/c{n}.sav')
#         to_rename = {
#             'clu#': f'dc{n}_{variant}',
#             # **{f'clu#{i}': f'c{n}_{i}' for i in range(1, n + 1)}
#         }
#         cl = cl.rename(columns=to_rename)
#         # cl[f'c{n}_max'] = cl[f'c{n}_max'].astype('int')
#         cl = cl[['RESPID'] + [v for k, v in to_rename.items()]].copy()
#         df = cl if df is None else pd.merge(df, cl)
#
# pyreadstat.write_sav(
#     df=df,
#     dst_path=data_root / "data dekarbonizace_24-11-07_lca.sav",
#     column_labels=df_meta.column_names_to_labels,
#     variable_value_labels=df_meta.variable_value_labels,
#     variable_display_width=df_meta.variable_display_width,
#     variable_measure=df_meta.variable_measure,
#     variable_format=df_meta.original_variable_types,
# )

# ys = [f'dc{i}_{v}' for v in ['log', 'cap'] for i in range(2, 10)]

def temp_output():
    return f'/home/thomas/tmp/output_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

def to_excel_rgb(rgb):
    """ Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl. """
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0]*255), int(rgb[1]*255), int(rgb[2]*255))


# STYLES ETC
color_scale_rule = ColorScaleRule(start_type='min', start_color='FFEF9C', end_type='max', end_color='63BE7B')
color_scale_rule_3 = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percent', mid_value=50,
                                    mid_color='FFEB84', end_type='max', end_color='63BE7B')
color_scale_rule_cont = ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percent', mid_value=50,
                                       mid_color='FCFCFF', end_type='max', end_color='5A8AC6')
# 3-graded: red F8696B, yellow FFEB84, green 63BE7B
# 3-graded for scale vars: red F8696B, white FCFCFF, blue 5A8AC6
center_wrap = Alignment(vertical='center', wrap_text=True)
horizontal_center = Alignment(horizontal='center')
both_center = Alignment(horizontal='center', vertical='center')
horizontal_center_wrap = Alignment(horizontal='center', wrap_text=True)
gray_font = Font(color='808080')
border_color = '666666'
thin_side = Side(style='thin', color=border_color)
thick_side = Side(style='medium', color=border_color)
thin_border = Border(bottom=thin_side)
thick_border = Border(bottom=thick_side)
value_cell_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)
value_cell_thick_border = Border(left=thin_side, right=thin_side, bottom=thick_side)
all_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)
all_thin_upper_thick_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thick_side)
top_thick_border = Border(top=thick_side)
left_thin_border = Border(left=thin_side)
right_thin_border = Border(right=thin_side)
right_thin_top_thick_border = Border(right=thin_side, top=thick_side)
left_thin_top_thick_border = Border(left=thin_side, top=thick_side)

wb = Workbook()
sheet = wb.active

if False:
    temp_file = temp_output()
    wb.save(temp_file)
    # os.system(f'xdg_open {temp_file} >/dev/null 2>&1 &')

# sheet_titles = {
#     'c6_max': '6 tříd, základní',
#     'c7_max_alt': '7 tříd, alternativní'
# }

add_tests = False
tests_gap = 2 if add_tests else 0

for sheet_num, yss in enumerate(ys):
    y_vars = np.sort(df[yss].unique())
    if sheet_num:
        # sheet = wb.create_sheet(title=sheet_titles[yss])
        sheet = wb.create_sheet(title=yss)
    else:
        # sheet.title = sheet_titles[yss]
        sheet.title = yss

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 36
    sheet.cell(row=1, column=3, value="VALUE").alignment = horizontal_center
    sheet.cell(row=1, column=4, value="LABEL").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(4)].width = 24

    for y_offset in (5, 5 + tests_gap + len(y_vars) + 1):
        for j, y in enumerate(y_vars):
            y_label = f'Třída {y}' if y != 99 else 'Missing'
            cell = sheet.cell(row=1, column=y_offset + j, value=y_label)
            cell.alignment = horizontal_center_wrap
            sheet.column_dimensions[get_column_letter(y_offset + j)].width = 18

        sheet.cell(row=1, column=y_offset + len(y_vars), value='TOTAL').alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(y_offset + len(y_vars))].width = 12

    if add_tests:
        sheet.column_dimensions[get_column_letter(6 + len(y_vars))].width = 12
        sheet.column_dimensions[get_column_letter(7 + len(y_vars))].width = 12
        sheet.cell(row=1, column=6 + len(y_vars), value='Kruskal-Wallis test').alignment = horizontal_center_wrap
        sheet.merge_cells(f'{get_column_letter(6 + len(y_vars))}1:{get_column_letter(7 + len(y_vars))}1')
        sheet.cell(row=2, column=6 + len(y_vars), value='Chui-Square').alignment = horizontal_center_wrap
        sheet.cell(row=2, column=7 + len(y_vars), value='Signifikance').alignment = horizontal_center_wrap

    col_freq_desc = 'Sloupcové relativní četnosti (formátování dle sloupců)'
    sheet.cell(row=2, column=5, value=col_freq_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f'E2:{get_column_letter(5 + len(y_vars))}2')

    col_agg_desc = 'Součty kladných a záporných odpovědí (formátování dle řádků)'
    sheet.cell(row=2, column=6 + tests_gap + len(y_vars), value=col_agg_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f'{get_column_letter(6 + tests_gap + len(y_vars))}2:{get_column_letter(6 + tests_gap + 2 * len(y_vars))}2')

    total_columns = 5 + tests_gap + 2 * len(y_vars) + 1
    merged_columns = 2
    for j in range(1, total_columns + 1):
        if j > 4:
            sheet.cell(row=1, column=j).border = all_thin_border
        sheet.cell(row=2, column=j).border = value_cell_thick_border if j > 4 else thick_border

    for j in range(1, 2 * len(y_vars) + 7 + tests_gap):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)

    # PROCEED COLUMN-WISE
    next_row = 3
    xs_remove = ['NUM_1', 'NUM_2', 'NUM_3', 'OKR', 'OBEC']
    for x in [x for x in xs if x not in xs_remove]:  # [:7]:
        print(f'Processing {yss} -> {x}')
        x_name = df_meta.column_names_to_labels[x]
        out_index = sorted(set([*df_meta.variable_value_labels[x].keys(), *df[x].unique()]))
        out = pd.DataFrame(index=out_index)
        for y in y_vars:
            sum_per_group = df[df[yss] == y].groupby(x, dropna=False)[w].sum()
            y_total = df[df[yss] == y][w].sum()
            pct_per_group = sum_per_group / y_total
            out[f'{yss}_{y}'] = pct_per_group

        sum_total = df.groupby(x, dropna=False)[w].sum()
        pct_total = sum_total / df[w].sum()
        out['total'] = pct_total
        out = out.fillna(0.)
        out = out.sort_index()

        cell = sheet.cell(row=next_row, column=1, value=x)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f'A{next_row}:A{next_row + len(out.index) - 1}')

        cell = sheet.cell(row=next_row, column=2, value=x_name)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f'B{next_row}:B{next_row + len(out.index) - 1}')

        for i, x_value in enumerate(out.index):
            sheet.cell(row=next_row + i, column=3, value=str(x_value)).border = thin_border
            if np.isnan(x_value):
                if len(df_meta.variable_value_labels[x]) == 1:
                    x_label = f'MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]'
                else:
                    x_label = "MISSING"
            elif x_value in df_meta.variable_value_labels[x]:
                x_label = df_meta.variable_value_labels[x][x_value]
            else:
                x_label = str(x_value)
            sheet.cell(row=next_row + i, column=4, value=x_label).border = thin_border
        sheet.cell(row=next_row + i, column=3).border = thick_border
        sheet.cell(row=next_row + i, column=4).border = thick_border

        # COLUMN-PERCENTAGES
        for j, out_col in enumerate(out.columns):
            y_out = out[out_col]
            norm = plt.Normalize(y_out.values.min(), y_out.values.max())
            colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
            for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
                # lighten colors
                cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
                # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
                cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
                # cell.fill = fill
                cell.number_format = '0.0%'
                cell.alignment = horizontal_center
                cell.border = value_cell_thin_border if i < len(y_out) - 1 else value_cell_thick_border
                if out_col == 'total':
                    cell.font = gray_font
            if out_col != 'total':
                col_letter = get_column_letter(5 + j)
                rng = f'{col_letter}{next_row}:{col_letter}{next_row + i}'
                sheet.conditional_formatting.add(rng, color_scale_rule)

        # KRUSKAL WALLIS TEST
        if add_tests:
            for i in range(out.shape[0]):
                sheet.cell(row=next_row + i,
                        column=5 + out.shape[1]).border = left_thin_border if i else left_thin_top_thick_border
                sheet.cell(row=next_row + i,
                        column=6 + out.shape[1]).border = right_thin_border if i else right_thin_top_thick_border

        # POSITIVE-NEGATIVE AGGREGATIONS
        agg_initial_col = 5 + tests_gap + out.shape[1]
        is_ordinal = x in df_meta.variable_measure and df_meta.variable_measure[x] == 'ordinal'
        if x not in agg_omitted_vars and is_ordinal:
            # this is a bit hacky, but should work
            value_list = list(df_meta.variable_value_labels[x])
            value_list = [int(x) for x in value_list if int(x) not in [88, 99]]
            agg_count = len(value_list) // 2
            agg_gap = len(value_list) % 2
            # agg_count = (out.shape[0] - 1) // 2
            agg_start = out.iloc[:agg_count].sum(axis=0)
            agg_end = out.iloc[agg_count + agg_gap: 2 * agg_count + agg_gap].sum(axis=0)
            for j, (sum_start, sum_end) in enumerate(zip(agg_start, agg_end)):
                for offset, sum_value in [(0, sum_start), (agg_count + agg_gap, sum_end)]:
                    cell = sheet.cell(row=next_row + offset, column=agg_initial_col + j, value=sum_value)
                    cell.number_format = '0.0%'
                    cell.alignment = both_center
                    cell.border = all_thin_border if offset else all_thin_upper_thick_border
                    if j == out.shape[1] - 1:
                        cell.font = gray_font

                sheet.cell(row=next_row + offset - 1, column=agg_initial_col + j).border = all_thin_border
                for i in range(2 * agg_count + agg_gap, out.shape[0]):
                    sheet.cell(row=next_row + i, column=agg_initial_col + j).border = all_thin_border

            # conditional formatting
            col_letter_start = get_column_letter(agg_initial_col)
            col_letter_end = get_column_letter(agg_initial_col + out.shape[1] - 2)
            rng_start = f'{col_letter_start}{next_row}:{col_letter_end}{next_row}'
            rng_end = f'{col_letter_start}{next_row + agg_count + agg_gap}:{col_letter_end}{next_row + agg_count + agg_gap}'
            sheet.conditional_formatting.add(rng_start, color_scale_rule_3)
            sheet.conditional_formatting.add(rng_end, color_scale_rule_3)

            # merge if necessary
            if agg_count > 1:
                for j in range(out.shape[1]):
                    col_letter = get_column_letter(agg_initial_col + j)
                    sheet.merge_cells(f'{col_letter}{next_row}:{col_letter}{next_row + agg_count - 1}')
                    sheet.merge_cells(
                        f'{col_letter}{next_row + agg_count + agg_gap}:{col_letter}{next_row + 2 * agg_count + agg_gap - 1}')

        else:  # only fix borders here
            for j in range(out.shape[1]):
                # bd = all_thin_border if i else all_thin_upper_thick_border
                sheet.cell(row=next_row, column=agg_initial_col + j).border = top_thick_border

        next_row = next_row + out.shape[0]

    # Fix some missing borders
    after_last_column = 7 + tests_gap + 2 * len(y_vars)
    for i in range(1, next_row):
        sheet.cell(row=i, column=after_last_column).border = left_thin_border
    for j in range(1, after_last_column):
        sheet.cell(row=next_row, column=j).border = top_thick_border


    # NOTE: continue with continuous variables here
    y_vars = np.sort(df[yss].unique())
    sheet = wb.create_sheet(title=f'{yss}_cont')

    for x in cont_xs:
        if x in df_meta.missing_ranges:
            print(x, df_meta.missing_ranges[x])
            # assert len(df_meta.missing_ranges[x]) == 1
            for miss_range in df_meta.missing_ranges[x]:
                assert miss_range['lo'] == miss_range['hi']
                df[x] = df[x].replace(miss_range['lo'], np.nan)


    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(1)].width = 12
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 60
    sheet.cell(row=1, column=3, value="TYPE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(3)].width = 12

    next_col = 4
    for j, y in enumerate(y_vars):
        y_label = f'Třída {y}' if y != 99 else 'Missing'
        cell = sheet.cell(row=1, column=next_col + j, value=y_label)
        cell.alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(next_col + j)].width = 18

    next_col = next_col + len(y_vars)
    sheet.cell(row=1, column=next_col, value='TOTAL').alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col)].width = 12

    sheet.cell(row=1, column=next_col + 1, value='MINIMUM').alignment = horizontal_center_wrap
    sheet.cell(row=1, column=next_col + 2, value='MAXIMUM').alignment = horizontal_center_wrap
    sheet.cell(row=1, column=next_col + 3, value='MISSING').alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 1)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 2)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 3)].width = 12
    sheet.cell(row=1, column=next_col + 4, value='ANOVA').alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 4)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 5)].width = 12
    sheet.merge_cells(f'{get_column_letter(next_col + 4)}1:{get_column_letter(next_col + 5)}1')

    # SUBHEADER
    next_col = 4
    tot_weight = df[w].sum()
    for j, y_value in enumerate(y_vars):
        group_weight = df[df[yss] == y_value][w].sum()
        cell = sheet.cell(row=2, column=next_col + j, value=group_weight / tot_weight)
        cell.alignment = horizontal_center_wrap
        cell.number_format = '0.0%'

    cell = sheet.cell(row=2, column=next_col + len(y_vars), value=1.)
    cell.alignment = horizontal_center_wrap
    cell.number_format = '0.0%'

    next_col = len(y_vars) + 8
    for j in [0, 1]:
        cell = sheet.cell(row=2, column=next_col + j, value='Signifikance' if j else 'F')
        cell.alignment = horizontal_center_wrap
    cell.number_format = '0.00'

    # FONTS AND BORDERS
    for j in range(1, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)
    for j in range(4, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).border = all_thin_border

    # CALCULATION
    next_row = 3
    for i, x in enumerate(cont_xs):  # [:5]):
        print(f'Processing {x}')
        x_name = df_meta.column_names_to_labels[x]

        # VAR NAMES ETC
        cell = sheet.cell(row=next_row + 2 * i, column=1, value=x)
        cell.alignment = center_wrap
        sheet.merge_cells(f'A{next_row + 2 * i}:A{next_row + 2 * i + 1}')

        cell = sheet.cell(row=next_row + 2 * i, column=2, value=x_name)
        cell.alignment = center_wrap
        sheet.merge_cells(f'B{next_row + 2 * i}:B{next_row + 2 * i + 1}')

        sheet.cell(row=next_row + 2 * i, column=3, value='průměr').alignment = center_wrap
        sheet.cell(row=next_row + 2 * i + 1, column=3, value='medián').alignment = center_wrap

        # PER GROUP MEANS AND MEDIANS
        next_col = 4
        for j, y_value in enumerate(y_vars):
            df_part = df[df[yss] == y_value]

            if len(df_part.dropna(subset=[x])) > 0:
                # mean
                om = OMean.compute(x=df_part[x], w=df_part[w])
                wmean = om.mean
                # median
                dsw = DescrStatsW(df_part[x], weights=df_part[w])
                wmed = dsw.quantile([0.5])[0.5]
            else:
                wmean = np.nan
                wmed = np.nan

            cell = sheet.cell(row=next_row + 2 * i, column=next_col + j, value=wmean)
            cell.number_format = '0.00'
            cell.alignment = horizontal_center

            cell = sheet.cell(row=next_row + 2 * i + 1, column=next_col + j, value=wmed)
            cell.number_format = '0.00'
            cell.alignment = horizontal_center

        rng_end = next_col + len(y_vars) - 1
        sheet.conditional_formatting.add(f'{get_column_letter(next_col)}{next_row + 2 * i}:'
                                        f'{get_column_letter(rng_end)}{next_row + 2 * i}', color_scale_rule_cont)
        sheet.conditional_formatting.add(f'{get_column_letter(next_col)}{next_row + 2 * i + 1}:'
                                        f'{get_column_letter(rng_end)}{next_row + 2 * i + 1}', color_scale_rule_cont)

        # TOTAL MEANS AND MEDIANS
        next_col = next_col + len(y_vars)
        om = OMean.compute(x=df[x], w=df[w])
        cell = sheet.cell(row=next_row + 2 * i, column=next_col, value=om.mean)
        cell.number_format = '0.00'
        cell.alignment = horizontal_center

        dsw = DescrStatsW(df[x], weights=df[w])
        wmed = dsw.quantile([0.5])[0.5]
        cell = sheet.cell(row=next_row + 2 * i + 1, column=next_col, value=wmed)
        cell.number_format = '0.00'
        cell.alignment = horizontal_center

        # MINIMUM AND MAXIMUM
        cell = sheet.cell(row=next_row + 2 * i, column=next_col + 1, value=df[x].min())
        cell.number_format = '0.00'
        cell.alignment = Alignment(vertical='center', horizontal='center')
        cell = sheet.cell(row=next_row + 2 * i, column=next_col + 2, value=df[x].max())
        cell.number_format = '0.00'
        cell.alignment = Alignment(vertical='center', horizontal='center')
        sheet.merge_cells(f'{get_column_letter(next_col + 1)}{next_row + 2 * i}:'
                        f'{get_column_letter(next_col + 1)}{next_row + 2 * i + 1}')
        sheet.merge_cells(f'{get_column_letter(next_col + 2)}{next_row + 2 * i}:'
                        f'{get_column_letter(next_col + 2)}{next_row + 2 * i + 1}')
        tot_miss = df[np.isnan(df[x])][w].sum()
        tot_weight = df[w].sum()
        cell = sheet.cell(row=next_row + 2 * i, column=next_col + 3, value=tot_miss / tot_weight)
        cell.number_format = '0.0%'
        cell.alignment = Alignment(vertical='center', horizontal='center')
        sheet.merge_cells(f'{get_column_letter(next_col + 3)}{next_row + 2 * i}:'
                        f'{get_column_letter(next_col + 3)}{next_row + 2 * i + 1}')

        # BORDERS
        for j in [1, 2, next_col + 1, next_col + 2]:
            sheet.cell(row=next_row + 2 * i, column=j).border = Border(top=thick_side, bottom=thick_side)
            sheet.cell(row=next_row + 2 * i + 1, column=j).border = Border(top=thick_side, bottom=thick_side)
        sheet.cell(row=next_row + 2 * i, column=3).border = Border(top=thick_side, bottom=thin_side)
        sheet.cell(row=next_row + 2 * i + 1, column=3).border = Border(top=thin_side, bottom=thick_side)
        for j in range(4, 8 + len(y_vars)):
            sheet.cell(row=next_row + 2 * i, column=j).border = Border(top=thick_side, bottom=thin_side, left=thin_side,
                                                                    right=thin_side)
            sheet.cell(row=next_row + 2 * i + 1, column=j).border = Border(top=thin_side, bottom=thick_side, left=thin_side,
                                                                        right=thin_side)
        an_i, an_j = next_row + 2 * i, 8 + len(y_vars)
        sheet.cell(row=an_i, column=an_j).border = Border(top=thick_side, left=thin_side)
        sheet.cell(row=an_i + 1, column=an_j).border = Border(bottom=thick_side, left=thin_side)
        sheet.cell(row=an_i, column=an_j + 1).border = Border(top=thick_side, right=thin_side)
        sheet.cell(row=an_i + 1, column=an_j + 1).border = Border(bottom=thick_side, right=thin_side)
        sheet.cell(row=an_i + 1, column=an_j + 1).border = Border(bottom=thick_side, right=thin_side)


temp_file = temp_output()
wb.save(temp_file)


