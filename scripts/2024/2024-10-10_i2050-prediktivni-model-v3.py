import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.linear_model import LogisticRegression
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *

data_root = Path("/home/thomas/projects/jan-krajhanzl/2024-09-06_ceska-dekarbonizace-24")
data_file = "data dekarbonizace_24-09-10.sav"

selected_nar = [30, 11, 7, 19, 10, 15, 13, 6, 16, 3, 14, 2, 39, 24]
len(selected_nar)
selected = ["GDF"] + [f"NAR_{nar}" for nar in selected_nar]
selected

df, df_meta = pyreadstat.read_sav(data_root / data_file)
resp_col = "c7_max_alt"
w_col = "weights"

df2 = df.dropna(subset=[resp_col]).copy()
for c in selected:
    df2[c + "_missing"] = df2[c].isna().astype(int)
df2["GDF"].value_counts(dropna=False)
df2["GDF"] = df2["GDF"].replace({np.nan: 2})
for c in selected:
    df2[c] = df2[c].replace({np.nan: 3})
selected_missing = [c + "_missing" for c in selected]

y = df2[resp_col].astype('category').cat.codes
weights = df2[w_col]

y.value_counts()
df2[resp_col].value_counts()

X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
    df2[selected + selected_missing], y, weights, test_size=0.2, random_state=1234, stratify=y
)

# X_train, X_test, y_train, y_test, w_train, w_test = train_test_split(
#     df2[list(feat_cols)], y, weights, test_size=0.2, random_state=1234, stratify=y
# )

# Create LightGBM datasets
train_data = lgb.Dataset(X_train, label=y_train, weight=w_train)
test_data = lgb.Dataset(X_test, label=y_test, weight=w_test)

# Set model parameters
params = {
    'objective': 'multiclass',
    'num_class': 7,  # Number of classes
    'metric': 'multi_logloss',
    'verbosity': -1,
    'seed': 42
}

# Train the model
model = lgb.train(
    params,
    train_data,
    valid_sets=[test_data],
    num_boost_round=1000,
)

# Make predictions
y_pred = model.predict(X_test)
y_pred_classes = y_pred.argmax(axis=1)

pd.DataFrame({'true': y_test, 'pred': y_pred_classes}).value_counts().unstack().fillna(0).astype(int)

# Evaluate the model - accuracy 0.7677
accuracy = accuracy_score(y_test, y_pred_classes)
print(f'Accuracy: {accuracy:.4f}')


# NOTE: finalni model
log_reg = LogisticRegression(multi_class='multinomial', solver='lbfgs', max_iter=5000)
log_reg.fit(X_train, y_train, sample_weight=w_train)

# Coefficients
coefficients = pd.DataFrame(log_reg.coef_, columns=selected + selected_missing)
print(coefficients)

# Evaluate the model - accuracy 0.8031
y_pred_log = log_reg.predict(X_test)
accuracy_log = accuracy_score(y_test, y_pred_log)
print(f'Accuracy of logistic regression: {accuracy_log:.4f}')

pred = pd.DataFrame({'true': y_test, 'pred': y_pred_log})
conf = pred.value_counts().unstack().fillna(0).astype(int)
conf.values.diagonal().sum() / conf.sum().sum()

pred = pd.DataFrame({'true': y_test, 'pred': y_pred_log, 'w': w_test})
conf = pred.groupby(["true", "pred"])["w"].sum().unstack().fillna(0)
conf.values.diagonal().sum() / conf.sum().sum()

class_labels = {
    0: "Vlažní",
    1: "Sympatizující",
    2: "Odmítající",
    3: "Nejistí",
    4: "Angažovaní pro Zelenou dohodu",
    5: "Angažovaní bez Zelené dohody",
    6: "Popírající",
}

conf_w_labels = conf.rename(columns=class_labels, index=class_labels)
conf_w_labels.round(1).to_csv(data_root / "model" / "confusion_matrix.csv")

log_reg
log_reg.intercept_
coef_df = pd.DataFrame(log_reg.coef_.T)

coef_df.loc[-1] = log_reg.intercept_
coef_df
coef_df.round(3).to_csv(data_root / "model" / "logistic_regression_coefs.csv", index=False)

int_df = pd.DataFrame(columns=range(7))
int_df log_reg.intercept_


import pickle
with open(data_root / "model" / "logistic_regression.pkl", "wb") as f:
    pickle.dump(log_reg, f)


sample = X_test.iloc[:20].reset_index(drop=True)
sample
sample.to_csv("logistic_regression_sample.csv", index=False)

foo = {"coef": [list(x) for x in list(log_reg.coef_)], "intercept": list(log_reg.intercept_)}
import json
with open(data_root / "model" / "logistic_regression.json", "w") as f:
    json.dump(foo, f)

df, df_meta = pyreadstat.read_sav(data_root / data_file)
resp_col = "c7_max_alt"
sample = df.dropna(subset=[resp_col]).sample(50, random_state=42).copy()
sample["class"] = (sample[resp_col] - 1).astype(int)
sample = sample[["class"] + selected].reset_index(drop=True).fillna(99).astype(int)

sample.to_csv(data_root / "model" / "logistic_regression_sample.csv", index=False)

# df, df_meta = pyreadstat.read_sav(data_root / data_file, user_missing=True)
# df["GDF"].value_counts(dropna=False)
# df["NAR_30"].value_counts(dropna=False)

sample = pd.read_csv(data_root / "model" / "logistic_regression_sample.csv")
with open(data_root / "model" / "logistic_regression.pkl", "rb") as f:
    log_reg = pickle.load(f)

import pickle
import pandas as pd
from sklearn.metrics import accuracy_score
from sklearn.linear_model import LogisticRegression

sample = pd.read_csv("logistic_regression_sample.csv")
with open("logistic_regression.pkl", "rb") as f:
    log_reg = pickle.load(f)

log_reg_vars = list(sample.columns)

for c in log_reg_vars:
    sample[c + "_missing"] = (sample[c] == 99).astype(int)
sample["GDF"] = sample["GDF"].replace({99: 2})
for c in log_reg_vars:
    sample[c] = sample[c].replace({99: 3})
log_reg_missing_vars = [c + "_missing" for c in log_reg_vars]

class_prediction = log_reg.predict(sample[log_reg_vars + log_reg_missing_vars])
class_prediction = log_reg.predict(df2[log_reg_vars])

# Evaluate the model - accuracy 0.8031
accuracy = accuracy_score(sample["class"], class_prediction)
print(f'Accuracy of logistic regression: {accuracy:.4f}')

# Confusion matrix
pd.DataFrame({'true': sample["class"], 'pred': class_prediction}).value_counts().unstack().fillna(0).astype(int)


questions = pd.DataFrame({"variable": log_reg_vars})
questions["text"] = questions["variable"].map(df_meta.column_names_to_labels)
questions["labels"] = questions["variable"].map(df_meta.variable_value_labels)

questions.to_csv(data_root / "model" / "questions.csv", index=False)






pd.isna(sample).sum()


df2.groupby(["GDF", "PO_20"])[w_col].sum().unstack()
df2.groupby("GDF")[w_col].sum()
df2.groupby("PO_20")[w_col].sum()

from scipy.stats import spearmanr

df3 = df2.dropna(subset=["GDF", "PO_20"])
# x and y are your ordinal variables
corr, p_value = spearmanr(df3["GDF"], df3["PO_20"])
print(f"Spearman correlation: {corr}, p-value: {p_value}")

import numpy as np

def weighted_corr(x, y, w):
    """Compute the weighted Pearson correlation between two variables."""
    def weighted_cov(x, y, w):
        """Compute the weighted covariance between two variables."""
        w_mean_x = np.average(x, weights=w)
        w_mean_y = np.average(y, weights=w)
        cov = np.sum(w * (x - w_mean_x) * (y - w_mean_y)) / np.sum(w)
        return cov

    cov_xy = weighted_cov(x, y, w)
    std_x = np.sqrt(weighted_cov(x, x, w))
    std_y = np.sqrt(weighted_cov(y, y, w))
    return cov_xy / (std_x * std_y)

# x and y are your continuous variables, w are the weights
corr = weighted_corr(df3["GDF"], df3["PO_20"], df3[w_col])
print(f"Weighted Pearson correlation: {corr}")

contingency_table = pd.crosstab(df2["GDF"], df2["PO_20"], values=df2[w_col], aggfunc='sum')
print(contingency_table)

from scipy.stats import chi2_contingency

chi2, p, dof, expected = chi2_contingency(contingency_table)
print(f"Chi-squared Statistic: {chi2}")
print(f"P-value: {p}")
print(f"Degrees of Freedom: {dof}")
print("Expected Frequencies:")
print(expected)


contingency_table = pd.crosstab(df2["GDF"], df2["PO_20"])
print(contingency_table)



df["POL_MISS"].value_counts(dropna=False)
len(df["NAR_MISS"].value_counts(dropna=False))

