import os
from pathlib import Path
from datetime import datetime
import pyreadstat
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color, Alignment, Font, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.stats.weightstats import DescrStatsW
import reportree as rt
from omoment import OMeanVar, OMean
from libs.extensions import *

data_root = Path("/home/thomas/projects/jan-krajhanzl/2025-06-11_lca-o-ochrane-klimy")
data_file = "25144-01-SK_DATA_ALL-FINAL-FINAL_VAHY_20250603.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file)

w_col = 'vahy'

cont_vars = ['NAR_CLICON', 'NAR_GDTHREAT', 'NAR_STRAT', 'NAR_TECH', 'POL_MEAN']
other_vars = ['GDF', 'POL_MISS', 'NAR_MISS']

gdf_map_cat = {0.0: 0, 1.0: 1, 2.0: 2, 3.0: 3, np.nan: 4 }
df['GDF_cat'] = df['GDF'].map(gdf_map_cat)
df['GDF_cat'].value_counts()

df['POL_MISS'].value_counts()
pd.cut(df['POL_MISS'], bins=[-1, 0, 1, 3, 12], labels=False).value_counts()
df['POL_MISS_rec'] = pd.cut(df['POL_MISS'], bins=[-1, 0, 1, 3, 12], labels=False)

df['NAR_MISS'].value_counts()
pd.cut(df['NAR_MISS'], bins=[-1, 0, 2, 5, 10, 100], labels=False).value_counts()
df['NAR_MISS_rec'] = pd.cut(df['NAR_MISS'], bins=[-1, 0, 2, 5, 10, 100], labels=False)

df["index"] = df.index
id_col = "index"
id_vars = [w_col, id_col]

new_other_vars = ['GDF_cat', 'POL_MISS_rec', 'NAR_MISS_rec']

df2 = df[cont_vars + new_other_vars + id_vars].dropna().copy()
df2 = df2[[id_col]].reset_index(drop=True).copy()
df2.shape

for n in range(2, 10):
    print(n)
    cl, _ = pyreadstat.read_sav(f'{data_root}/lca/c{n}.sav')
    to_rename = {
        'clu#': f'c{n}_max',
        **{f'clu#{i}': f'c{n}_{i}' for i in range(1, n + 1)}
    }
    cl = cl.rename(columns=to_rename)
    cl[f'c{n}_max'] = cl[f'c{n}_max'].astype('int')
    cl = cl[[v for k, v in to_rename.items()]].copy()
    df2 = pd.concat([df2, cl], axis=1)

df = pd.merge(df, df2, on=id_col, how='left')

for c in ["POL_MISS"] + [f"OS_{i}" for i in range(1, 4)] + ["CARS"]:
    df_meta.variable_measure[c] = "scale"

nom_fixes = ["SEX", "EDU", "VMB", "SET", "INT", "POJ", "INC", "INC_PCT", "INC_MEDP", "EL21",
             "HOME", "HEAT1", "EDU2", "ETS2A", "ETS2B"]

for k, v in df_meta.variable_measure.items():
    if v == "nominal" and  k not in nom_fixes:
        df_meta.variable_measure[k] = "ordinal"

for i in range(1, 6):
    df_meta.variable_value_labels[f"SQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 9):
    df_meta.variable_value_labels[f"SOC_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 10):
    df_meta.variable_value_labels[f"EQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}

pyreadstat.write_sav(
    df=df,
    dst_path=data_root / "25144-01-SK_DATA_ALL-FINAL-FINAL_VAHY_20250603_TP_classes_full.sav",
    column_labels=df_meta.column_names_to_labels,
    variable_value_labels=df_meta.variable_value_labels,
    variable_display_width=df_meta.variable_display_width,
    variable_measure=df_meta.variable_measure,
    variable_format=df_meta.original_variable_types,
)

c_range = range(2, 10)
cl_cols = [x for n in c_range for x in [f'c{n}_max'] + [f'c{n}_{i}' for i in range(1, n + 1)]]

for n in c_range:
    for i in range(1, n + 1):
        df[f'c{n}_{i}_w'] = df[f'c{n}_{i}'] * df[w_col]
        cl_cols.append(f'c{n}_{i}_w')

all_vars_df = pd.read_excel(f"{data_root}/Premenné pre porovnanie.xlsx").dropna().copy()
all_vars_df.columns = ["names"]
all_vars_df["measure"] = all_vars_df["names"].apply(
    lambda x: df_meta.variable_measure[x] if x in df_meta.variable_measure else None
)
all_vars = list(all_vars_df.values[:, 0])

def table_for_frame(df, df_labels, df_measure):
    # I need to detect missing vars
    miss_cols = []  # list of cols with missing values
    # miss_value = {}  # value that should be treated as missing -> do I care? or shall I just replace it?
    for c, m in df_measure.items():
        if m in ['ordinal', 'nominal']:
            c_max = df[c].max()
            c_2nd_max = df[c][df[c] != c_max].max()
            if c_max in [9., 85., 98., 99.] and (c_max - c_2nd_max) > 1.5:
                # print('missing:', c, c_max)
                # TODO: replace miss by nans and add to list
                miss_cols.append(c)
                df[c] = df[c].replace(c_max, np.nan)

    out_frames = {}
    total_weight = df[w_col].sum()

    # output_cols
    no_output_col = [w_col] + list(df.dtypes[(df.dtypes != np.float64) & (df.dtypes != np.int64)].index)
    no_output_col += ["GDF_cat", "NAR_MISS_rec", "POL_MISS_rec", "index"]
    output_cols = [c for c in df.columns if c not in cl_cols + no_output_col]
    output_cols = all_vars

    for nclass in c_range:
        print(f'nclass = {nclass}')

        out_var = pd.Series({c: c for c in output_cols})
        out_label = pd.Series({c: df_labels[c] for c in output_cols})
        out_mean = pd.Series(dtype='float64')
        out_std_dev = pd.Series(dtype='float64')
        out_miss_w = pd.Series(dtype='float64')
        out_min = pd.Series(dtype='float64')
        out_max = pd.Series(dtype='float64')

        out_class_means = {}
        out_class_miss_w = {}
        out_maxclass_means = {}
        out_maxclass_miss_w = {}
        for i in range(1, nclass + 1):
            out_class_means[i] = pd.Series(dtype='float64')
            out_class_miss_w[i] = pd.Series(dtype='float64')
            out_maxclass_means[i] = pd.Series(dtype='float64')
            out_maxclass_miss_w[i] = pd.Series(dtype='float64')

        class_w = df[[f'c{nclass}_{i}_w' for i in range(1, nclass + 1)]].sum()
        class_w.index = np.arange(1, nclass + 1)
        maxclass_w = df.groupby(f'c{nclass}_max')[w_col].sum()

        # for c in output_cols:
        for c in all_vars:
            print(f'Processing {c}')
            try:
                omv = OMeanVar.compute(x=df[c], w=df[w_col])
            except ZeroDivisionError:
                continue
            out_mean[c] = omv.mean
            out_std_dev[c] = omv.std_dev
            out_miss_w[c] = 1. - omv.weight / total_weight
            out_min[c] = df[c].min()
            out_max[c] = df[c].max()

            for i in range(1, nclass + 1):
                try:
                    om = OMean.compute(x=df[c], w=df[f'c{nclass}_{i}_w'])
                except ZeroDivisionError:
                    om = OMean()
                out_class_means[i][c] = om.mean
                out_class_miss_w[i][c] = 1. - om.weight / class_w[i]

            foo = OMean.of_groupby(df, g=f'c{nclass}_max', x=c, w=w_col)
            # foo = df.groupby(f'c{nclass}_max')[[c, 'w']].apply(nanaverage, weights='w')[c]
            for i in range(1, nclass + 1):
                out_maxclass_means[i][c] = foo[i].mean if i in foo.index else np.nan
                out_maxclass_miss_w[i][c] = (1. - foo[i].weight / maxclass_w[i]) if i in foo.index else np.nan

        out_empty = out_var.apply(lambda _: '')

        outs = {
            'var': out_var,
            'label': out_label,
            'mean': out_mean,
            'std_dev': out_std_dev,
            'miss_w': out_miss_w,
            'min': out_min,
            'max': out_max,
            ' ': out_empty,
        }
        output = pd.DataFrame(outs)

        for i in range(1, nclass + 1):
            output[f'c{i}_mean'] = out_maxclass_means[i]
        output['c_min'] = output[[f'c{i}_mean' for i in range(1, nclass + 1)]].min(axis=1)
        output['c_max'] = output[[f'c{i}_mean' for i in range(1, nclass + 1)]].max(axis=1)
        output['c_diff'] = output['c_max'] - output['c_min']
        output['c_pct'] = output['c_diff'] / (output['max'] - output['min'])
        output['c_pct_x100'] = 100 * output['c_pct']
        for i in range(1, nclass + 1):
            output[f'c{i}_miss_w'] = out_maxclass_miss_w[i]

        output['  '] = out_empty
        for i in range(1, nclass + 1):
            output[f'cpp{i}_mean'] = out_class_means[i]
        output['cpp_min'] = output[[f'cpp{i}_mean' for i in range(1, nclass + 1)]].min(axis=1)
        output['cpp_max'] = output[[f'cpp{i}_mean' for i in range(1, nclass + 1)]].max(axis=1)
        output['cpp_diff'] = output['cpp_max'] - output['cpp_min']
        output['cpp_pct'] = output['cpp_diff'] / (output['max'] - output['min'])
        output['cpp_pct_x100'] = 100 * output['cpp_pct']
        for i in range(1, nclass + 1):
            output[f'cpp{i}_miss_w'] = out_class_miss_w[i]

        # add sizes of classes (relative)
        out_size = output.iloc[0].copy().rename('SIZE').apply(lambda _: '')
        out_size['var'] = 'SIZE'
        out_size['label'] = 'Relativní velikost tříd'
        for i in range(1, nclass + 1):
            out_size[f'cpp{i}_mean'] = class_w[i] / total_weight
            out_size[f'c{i}_mean'] = maxclass_w[i] / total_weight
        output = pd.concat([pd.DataFrame(out_size).T, output], axis=0)

        # output.to_csv(f'output/lca-ord/c{nclass}.csv', index=False, encoding='utf8')
        out_frames[nclass] = output

    return out_frames


out_variants = {}
out_file = 'lca-results.xlsx'
sheet_name = 'lca_{}c'
tables = table_for_frame(df, df_meta.column_names_to_labels, df_meta.variable_measure)


with pd.ExcelWriter(os.path.join(data_root, out_file)) as writer:
    for n_class, sheet in tables.items():
        sheet.to_excel(writer, sheet_name=sheet_name.format(n_class), index=False)


# DONE
df_labels = df_meta.column_names_to_labels
df_measure = df_meta.variable_measure

OMeanVar.compute(df["M9W1"], w=df[w_col])


xs = [x for x in all_vars if df_meta.variable_measure[x] in ["nominal", "ordinal"]]
cont_xs = [x for x in all_vars if df_meta.variable_measure[x] == "scale"]


df.shape
df2
pd.
pd.concat([df2.reset_index(drop=True), cl], axis=1)

data_root = Path("/home/thomas/projects/jan-krajhanzl/2025-06-11_lca-o-ochrane-klimy")
data_file = "25144-01-SK_DATA_ALL-FINAL-FINAL_VAHY_20250603_TP_classes.sav"

df, df_meta = pyreadstat.read_sav(data_root / data_file)
# df.shape

all_vars_df = pd.read_excel(f"{data_root}/Premenné pre porovnanie.xlsx").dropna().copy()
all_vars_df.columns = ["names"]
all_vars_df["measure"] = all_vars_df["names"].apply(
    lambda x: df_meta.variable_measure[x] if x in df_meta.variable_measure else None
)
all_vars = list(all_vars_df.values[:, 0])
#
# df_meta.variable_measure[all_vars[0]]
# all_vars_df["measure"].value_counts()
# df_meta.variable_value_labels[all_vars[-1]]

# FIXES:
for c in ["POL_MISS"] + [f"OS_{i}" for i in range(1, 4)] + ["CARS"]:
    df_meta.variable_measure[c] = "scale"

nom_fixes = ["SEX", "EDU", "VMB", "SET", "INT", "POJ", "INC", "INC_PCT", "INC_MEDP", "EL21",
             "HOME", "HEAT1", "EDU2", "ETS2A", "ETS2B"]

for k, v in df_meta.variable_measure.items():
    if v == "nominal" and  k not in nom_fixes:
        df_meta.variable_measure[k] = "ordinal"

xs = [x for x in all_vars if df_meta.variable_measure[x] in ["nominal", "ordinal"]]
cont_xs = [x for x in all_vars if df_meta.variable_measure[x] == "scale"]

for i in range(1, 6):
    df_meta.variable_value_labels[f"SQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 9):
    df_meta.variable_value_labels[f"SOC_{i}"] = {0.0: "Ne", 1.0: "Ano"}
for i in range(1, 10):
    df_meta.variable_value_labels[f"EQ_{i}"] = {0.0: "Ne", 1.0: "Ano"}

w = "vahy"
agg_omitted_vars = []

ys = [f"c{i}_max" for i in range(2, 10)]




def temp_output():
    return f"/home/thomas/tmp/output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def to_excel_rgb(rgb):
    """Convert RGB values from 0-1 scale to 0-255 scale and format it for use in openpyxl."""
    return "{0:02X}{1:02X}{2:02X}".format(int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255))


# STYLES ETC
color_scale_rule = ColorScaleRule(start_type="min", start_color="FFEF9C", end_type="max", end_color="63BE7B")
color_scale_rule_3 = ColorScaleRule(
    start_type="min",
    start_color="F8696B",
    mid_type="percent",
    mid_value=50,
    mid_color="FFEB84",
    end_type="max",
    end_color="63BE7B",
)
color_scale_rule_cont = ColorScaleRule(
    start_type="min",
    start_color="F8696B",
    mid_type="percent",
    mid_value=50,
    mid_color="FCFCFF",
    end_type="max",
    end_color="5A8AC6",
)
color_scale_rule_cont_miss = ColorScaleRule(start_type="min", start_color="FCFCFF", end_type="max", end_color="F8696B")
# 3-graded: red F8696B, yellow FFEB84, green 63BE7B
# 3-graded for scale vars: red F8696B, white FCFCFF, blue 5A8AC6
center_wrap = Alignment(vertical="center", wrap_text=True)
horizontal_center = Alignment(horizontal="center")
both_center = Alignment(horizontal="center", vertical="center")
horizontal_center_wrap = Alignment(horizontal="center", wrap_text=True)
gray_font = Font(color="808080")
border_color = "666666"
thin_side = Side(style="thin", color=border_color)
thick_side = Side(style="medium", color=border_color)
thin_border = Border(bottom=thin_side)
thick_border = Border(bottom=thick_side)
value_cell_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side)
value_cell_thick_border = Border(left=thin_side, right=thin_side, bottom=thick_side)
all_thin_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thin_side)
all_thin_upper_thick_border = Border(left=thin_side, right=thin_side, bottom=thin_side, top=thick_side)
top_thick_border = Border(top=thick_side)
left_thin_border = Border(left=thin_side)
right_thin_border = Border(right=thin_side)
right_thin_top_thick_border = Border(right=thin_side, top=thick_side)
left_thin_top_thick_border = Border(left=thin_side, top=thick_side)

wb = Workbook()
sheet = wb.active

if False:
    temp_file = temp_output()
    wb.save(temp_file)
    # os.system(f'xdg_open {temp_file} >/dev/null 2>&1 &')

add_tests = False
tests_gap = 2 if add_tests else 0

for sheet_num, yss in enumerate(ys):
    y_vars = np.sort(df[yss].unique())
    if sheet_num:
        # sheet = wb.create_sheet(title=sheet_titles[yss])
        sheet = wb.create_sheet(title=yss)
    else:
        # sheet.title = sheet_titles[yss]
        sheet.title = yss

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 36
    sheet.cell(row=1, column=3, value="VALUE").alignment = horizontal_center
    sheet.cell(row=1, column=4, value="LABEL").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(4)].width = 24

    for y_offset in (5, 5 + tests_gap + len(y_vars) + 1):
        tot_w = df[w].sum()
        y_vars_w = df.groupby(yss)[w].sum() / tot_w
        for j, y in enumerate(y_vars):
            y_label = f"Třída {y} ({100 * y_vars_w[y]:.1f} %)" if y != 99 else "Missing"
            cell = sheet.cell(row=1, column=y_offset + j, value=y_label)
            cell.alignment = horizontal_center_wrap
            sheet.column_dimensions[get_column_letter(y_offset + j)].width = 18

        sheet.cell(row=1, column=y_offset + len(y_vars), value="TOTAL").alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(y_offset + len(y_vars))].width = 12

    if add_tests:
        sheet.column_dimensions[get_column_letter(6 + len(y_vars))].width = 12
        sheet.column_dimensions[get_column_letter(7 + len(y_vars))].width = 12
        sheet.cell(row=1, column=6 + len(y_vars), value="Kruskal-Wallis test").alignment = horizontal_center_wrap
        sheet.merge_cells(f"{get_column_letter(6 + len(y_vars))}1:{get_column_letter(7 + len(y_vars))}1")
        sheet.cell(row=2, column=6 + len(y_vars), value="Chui-Square").alignment = horizontal_center_wrap
        sheet.cell(row=2, column=7 + len(y_vars), value="Signifikance").alignment = horizontal_center_wrap

    col_freq_desc = "Sloupcové relativní četnosti (formátování dle sloupců)"
    sheet.cell(row=2, column=5, value=col_freq_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(f"E2:{get_column_letter(5 + len(y_vars))}2")

    col_agg_desc = "Součty kladných a záporných odpovědí (formátování dle řádků)"
    sheet.cell(row=2, column=6 + tests_gap + len(y_vars), value=col_agg_desc).alignment = horizontal_center_wrap
    sheet.merge_cells(
        f"{get_column_letter(6 + tests_gap + len(y_vars))}2:{get_column_letter(6 + tests_gap + 2 * len(y_vars))}2"
    )

    total_columns = 5 + tests_gap + 2 * len(y_vars) + 1
    merged_columns = 2
    for j in range(1, total_columns + 1):
        if j > 4:
            sheet.cell(row=1, column=j).border = all_thin_border
        sheet.cell(row=2, column=j).border = value_cell_thick_border if j > 4 else thick_border

    for j in range(1, 2 * len(y_vars) + 7 + tests_gap):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)

    # PROCEED COLUMN-WISE
    next_row = 3
    xs_remove = ["NUM_1", "NUM_2", "NUM_3", "OKR", "OBEC"]
    for x in [x for x in xs if x not in xs_remove]:  # [:7]:
        print(f"Processing {yss} -> {x}")
        x_name = df_meta.column_names_to_labels[x]
        out_index = sorted(set([*df_meta.variable_value_labels[x].keys(), *df[x].unique()]))
        out = pd.DataFrame(index=out_index)
        for y in y_vars:
            sum_per_group = df[df[yss] == y].groupby(x, dropna=False)[w].sum()
            y_total = df[df[yss] == y][w].sum()
            pct_per_group = sum_per_group / y_total
            out[f"{yss}_{y}"] = pct_per_group

        sum_total = df.groupby(x, dropna=False)[w].sum()
        pct_total = sum_total / df[w].sum()
        out["total"] = pct_total
        out = out.fillna(0.0)
        out = out.sort_index()

        cell = sheet.cell(row=next_row, column=1, value=x)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f"A{next_row}:A{next_row + len(out.index) - 1}")

        cell = sheet.cell(row=next_row, column=2, value=x_name)
        cell.alignment = center_wrap
        cell.border = thick_border
        sheet.merge_cells(f"B{next_row}:B{next_row + len(out.index) - 1}")

        for i, x_value in enumerate(out.index):
            sheet.cell(row=next_row + i, column=3, value=str(x_value)).border = thin_border
            if np.isnan(x_value):
                if len(df_meta.variable_value_labels[x]) == 1:
                    x_label = f"MISSING [{list(df_meta.variable_value_labels[x].values())[0]}]"
                else:
                    x_label = "MISSING"
            elif x_value in df_meta.variable_value_labels[x]:
                x_label = df_meta.variable_value_labels[x][x_value]
            else:
                x_label = str(x_value)
            sheet.cell(row=next_row + i, column=4, value=x_label).border = thin_border
        sheet.cell(row=next_row + i, column=3).border = thick_border
        sheet.cell(row=next_row + i, column=4).border = thick_border

        # COLUMN-PERCENTAGES
        for j, out_col in enumerate(out.columns):
            y_out = out[out_col]
            norm = plt.Normalize(y_out.values.min(), y_out.values.max())
            colors = plt.cm.RdYlGn(norm(y_out.values))  # replace "Reds" with the color map you prefer
            for i, (c, (x_value, value)) in enumerate(zip(colors, y_out.items())):
                # lighten colors
                cc = to_excel_rgb([comp + 0.6 * (1 - comp) for comp in c])
                # fill = PatternFill(start_color=cc, end_color=cc, fill_type="solid")
                cell = sheet.cell(row=next_row + i, column=5 + j, value=value)
                # cell.fill = fill
                cell.number_format = "0.0%"
                cell.alignment = horizontal_center
                cell.border = value_cell_thin_border if i < len(y_out) - 1 else value_cell_thick_border
                if out_col == "total":
                    cell.font = gray_font
            if out_col != "total":
                col_letter = get_column_letter(5 + j)
                rng = f"{col_letter}{next_row}:{col_letter}{next_row + i}"
                sheet.conditional_formatting.add(rng, color_scale_rule)

        # KRUSKAL WALLIS TEST
        if add_tests:
            for i in range(out.shape[0]):
                sheet.cell(row=next_row + i, column=5 + out.shape[1]).border = (
                    left_thin_border if i else left_thin_top_thick_border
                )
                sheet.cell(row=next_row + i, column=6 + out.shape[1]).border = (
                    right_thin_border if i else right_thin_top_thick_border
                )

        # POSITIVE-NEGATIVE AGGREGATIONS
        agg_initial_col = 5 + tests_gap + out.shape[1]
        is_ordinal = x in df_meta.variable_measure and df_meta.variable_measure[x] == "ordinal"
        if x not in agg_omitted_vars and is_ordinal:
            # this is a bit hacky, but should work
            value_list = list(df_meta.variable_value_labels[x])
            value_list = [int(x) for x in value_list if int(x) not in [85, 88, 99]]
            agg_count = len(value_list) // 2
            agg_gap = len(value_list) % 2
            # agg_count = (out.shape[0] - 1) // 2
            agg_start = out.iloc[:agg_count].sum(axis=0)
            agg_end = out.iloc[agg_count + agg_gap : 2 * agg_count + agg_gap].sum(axis=0)
            for j, (sum_start, sum_end) in enumerate(zip(agg_start, agg_end)):
                for offset, sum_value in [(0, sum_start), (agg_count + agg_gap, sum_end)]:
                    cell = sheet.cell(row=next_row + offset, column=agg_initial_col + j, value=sum_value)
                    cell.number_format = "0.0%"
                    cell.alignment = both_center
                    cell.border = all_thin_border if offset else all_thin_upper_thick_border
                    if j == out.shape[1] - 1:
                        cell.font = gray_font

                sheet.cell(row=next_row + offset - 1, column=agg_initial_col + j).border = all_thin_border
                for i in range(2 * agg_count + agg_gap, out.shape[0]):
                    sheet.cell(row=next_row + i, column=agg_initial_col + j).border = all_thin_border

            # conditional formatting
            col_letter_start = get_column_letter(agg_initial_col)
            col_letter_end = get_column_letter(agg_initial_col + out.shape[1] - 2)
            rng_start = f"{col_letter_start}{next_row}:{col_letter_end}{next_row}"
            rng_end = (
                f"{col_letter_start}{next_row + agg_count + agg_gap}:{col_letter_end}{next_row + agg_count + agg_gap}"
            )
            sheet.conditional_formatting.add(rng_start, color_scale_rule_3)
            sheet.conditional_formatting.add(rng_end, color_scale_rule_3)

            # merge if necessary
            if agg_count > 1:
                for j in range(out.shape[1]):
                    col_letter = get_column_letter(agg_initial_col + j)
                    sheet.merge_cells(f"{col_letter}{next_row}:{col_letter}{next_row + agg_count - 1}")
                    sheet.merge_cells(
                        f"{col_letter}{next_row + agg_count + agg_gap}:{col_letter}{next_row + 2 * agg_count + agg_gap - 1}"
                    )

        else:  # only fix borders here
            for j in range(out.shape[1]):
                # bd = all_thin_border if i else all_thin_upper_thick_border
                sheet.cell(row=next_row, column=agg_initial_col + j).border = top_thick_border

        next_row = next_row + out.shape[0]

    # Fix some missing borders
    after_last_column = 7 + tests_gap + 2 * len(y_vars)
    for i in range(1, next_row):
        sheet.cell(row=i, column=after_last_column).border = left_thin_border
    for j in range(1, after_last_column):
        sheet.cell(row=next_row, column=j).border = top_thick_border

    # NOTE: continue with continuous variables here
    y_vars = np.sort(df[yss].unique())
    sheet = wb.create_sheet(title=f"{yss}_cont")

    for x in cont_xs:
        if x in df_meta.missing_ranges:
            print(x, df_meta.missing_ranges[x])
            # assert len(df_meta.missing_ranges[x]) == 1
            for miss_range in df_meta.missing_ranges[x]:
                assert miss_range["lo"] == miss_range["hi"]
                df[x] = df[x].replace(miss_range["lo"], np.nan)

    # for c in fix_missing_vars:
    #     df[c] = df[c].fillna(0)

    # HEADER
    sheet.cell(row=1, column=1, value="VARIABLE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(1)].width = 12
    sheet.cell(row=1, column=2, value="NAME").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(2)].width = 60
    sheet.cell(row=1, column=3, value="TYPE").alignment = horizontal_center
    sheet.column_dimensions[get_column_letter(3)].width = 12

    next_col = 4
    tot_w = df[w].sum()
    y_vars_w = df.groupby(yss)[w].sum() / tot_w
    for j, y in enumerate(y_vars):
        y_label = f"Třída {y} ({100 * y_vars_w[y]:.1f} %)" if y != 99 else "Missing"
        cell = sheet.cell(row=1, column=next_col + j, value=y_label)
        cell.alignment = horizontal_center_wrap
        sheet.column_dimensions[get_column_letter(next_col + j)].width = 18

    next_col = next_col + len(y_vars)
    sheet.cell(row=1, column=next_col, value="TOTAL").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col)].width = 12

    sheet.cell(row=1, column=next_col + 1, value="MINIMUM").alignment = horizontal_center_wrap
    sheet.cell(row=1, column=next_col + 2, value="MAXIMUM").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 1)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 2)].width = 12
    sheet.cell(row=1, column=next_col + 3, value="ANOVA").alignment = horizontal_center_wrap
    sheet.column_dimensions[get_column_letter(next_col + 3)].width = 12
    sheet.column_dimensions[get_column_letter(next_col + 4)].width = 12
    sheet.merge_cells(f"{get_column_letter(next_col + 3)}1:{get_column_letter(next_col + 4)}1")

    # SUBHEADER
    next_col = 4
    tot_weight = df[w].sum()
    for j, y_value in enumerate(y_vars):
        group_weight = df[df[yss] == y_value][w].sum()
        cell = sheet.cell(row=2, column=next_col + j, value=group_weight / tot_weight)
        cell.alignment = horizontal_center_wrap
        cell.number_format = "0.0%"

    cell = sheet.cell(row=2, column=next_col + len(y_vars), value=1.0)
    cell.alignment = horizontal_center_wrap
    cell.number_format = "0.0%"

    next_col = len(y_vars) + 7
    for j in [0, 1]:
        cell = sheet.cell(row=2, column=next_col + j, value="Signifikance" if j else "F")
        cell.alignment = horizontal_center_wrap
    cell.number_format = "0.00"

    # FONTS AND BORDERS
    for j in range(1, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).font = Font(bold=True)
    for j in range(4, len(y_vars) + 9):
        for i in [1, 2]:
            sheet.cell(row=i, column=j).border = all_thin_border

    # CALCULATION
    next_row = 3
    for i, x in enumerate(cont_xs):  # [:5]):
        print(f"Processing {x}")
        x_name = df_meta.column_names_to_labels[x]

        # VAR NAMES ETC
        cell = sheet.cell(row=next_row + 3 * i, column=1, value=x)
        cell.alignment = center_wrap
        sheet.merge_cells(f"A{next_row + 3 * i}:A{next_row + 3 * i + 2}")

        cell = sheet.cell(row=next_row + 3 * i, column=2, value=x_name)
        cell.alignment = center_wrap
        sheet.merge_cells(f"B{next_row + 3 * i}:B{next_row + 3 * i + 2}")

        sheet.cell(row=next_row + 3 * i, column=3, value="průměr").alignment = center_wrap
        sheet.cell(row=next_row + 3 * i + 1, column=3, value="medián").alignment = center_wrap
        sheet.cell(row=next_row + 3 * i + 2, column=3, value="missing").alignment = center_wrap

        # PER GROUP MEANS AND MEDIANS
        next_col = 4
        for j, y_value in enumerate(y_vars):
            df_part = df[df[yss] == y_value]

            if len(df_part.dropna(subset=[x])) > 0:
                # mean
                om = OMean.compute(x=df_part[x], w=df_part[w])
                wmean = om.mean
                # median
                dsw = DescrStatsW(df_part[x], weights=df_part[w])
                wmed = dsw.quantile([0.5])[0.5]
            else:
                wmean = np.nan
                wmed = np.nan

            part_miss = df_part[np.isnan(df_part[x])][w].sum()
            part_weight = df_part[w].sum()
            wmiss = part_miss / part_weight

            cell = sheet.cell(row=next_row + 3 * i, column=next_col + j, value=wmean)
            cell.number_format = "0.00"
            cell.alignment = horizontal_center

            cell = sheet.cell(row=next_row + 3 * i + 1, column=next_col + j, value=wmed)
            cell.number_format = "0.00"
            cell.alignment = horizontal_center

            cell = sheet.cell(row=next_row + 3 * i + 2, column=next_col + j, value=wmiss)
            cell.number_format = "0.0%"
            cell.alignment = horizontal_center

        rng_end = next_col + len(y_vars) - 1
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i}:{get_column_letter(rng_end)}{next_row + 3 * i}",
            color_scale_rule_cont,
        )
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i + 1}:{get_column_letter(rng_end)}{next_row + 3 * i + 1}",
            color_scale_rule_cont,
        )
        sheet.conditional_formatting.add(
            f"{get_column_letter(next_col)}{next_row + 3 * i + 2}:{get_column_letter(rng_end)}{next_row + 3 * i + 2}",
            color_scale_rule_cont_miss,
        )

        # TOTAL MEANS AND MEDIANS
        next_col = next_col + len(y_vars)
        om = OMean.compute(x=df[x], w=df[w])
        cell = sheet.cell(row=next_row + 3 * i, column=next_col, value=om.mean)
        cell.number_format = "0.00"
        cell.alignment = horizontal_center

        dsw = DescrStatsW(df[x], weights=df[w])
        wmed = dsw.quantile([0.5])[0.5]
        cell = sheet.cell(row=next_row + 3 * i + 1, column=next_col, value=wmed)
        cell.number_format = "0.00"
        cell.alignment = horizontal_center

        tot_miss = df[np.isnan(df[x])][w].sum()
        tot_weight = df[w].sum()
        cell = sheet.cell(row=next_row + 3 * i + 2, column=next_col, value=tot_miss / tot_weight)
        cell.number_format = "0.0%"
        cell.alignment = horizontal_center

        # MINIMUM AND MAXIMUM
        cell = sheet.cell(row=next_row + 3 * i, column=next_col + 1, value=df[x].min())
        cell.number_format = "0.00"
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell = sheet.cell(row=next_row + 3 * i, column=next_col + 2, value=df[x].max())
        cell.number_format = "0.00"
        cell.alignment = Alignment(vertical="center", horizontal="center")
        sheet.merge_cells(
            f"{get_column_letter(next_col + 1)}{next_row + 3 * i}:"
            f"{get_column_letter(next_col + 1)}{next_row + 3 * i + 2}"
        )
        sheet.merge_cells(
            f"{get_column_letter(next_col + 2)}{next_row + 3 * i}:"
            f"{get_column_letter(next_col + 2)}{next_row + 3 * i + 2}"
        )

        # BORDERS
        for j in [1, 2, next_col + 1, next_col + 2]:
            sheet.cell(row=next_row + 3 * i, column=j).border = Border(top=thick_side, bottom=thick_side)
            sheet.cell(row=next_row + 3 * i + 2, column=j).border = Border(top=thick_side, bottom=thick_side)
        sheet.cell(row=next_row + 3 * i, column=3).border = Border(top=thick_side, bottom=thin_side)
        sheet.cell(row=next_row + 3 * i + 1, column=3).border = Border(top=thin_side, bottom=thin_side)
        sheet.cell(row=next_row + 3 * i + 2, column=3).border = Border(top=thin_side, bottom=thick_side)
        for j in range(4, 7 + len(y_vars)):
            sheet.cell(row=next_row + 3 * i, column=j).border = Border(
                top=thick_side, bottom=thin_side, left=thin_side, right=thin_side
            )
            sheet.cell(row=next_row + 3 * i + 1, column=j).border = Border(
                top=thin_side, bottom=thin_side, left=thin_side, right=thin_side
            )
            sheet.cell(row=next_row + 3 * i + 2, column=j).border = Border(
                top=thin_side, bottom=thick_side, left=thin_side, right=thin_side
            )
        an_i, an_j = next_row + 3 * i, 7 + len(y_vars)
        sheet.cell(row=an_i, column=an_j).border = Border(top=thick_side, left=thin_side)
        sheet.cell(row=an_i + 1, column=an_j).border = Border(left=thin_side)
        sheet.cell(row=an_i + 2, column=an_j).border = Border(bottom=thick_side, left=thin_side)
        sheet.cell(row=an_i, column=an_j + 1).border = Border(top=thick_side, right=thin_side)
        sheet.cell(row=an_i + 1, column=an_j + 1).border = Border(right=thin_side)
        sheet.cell(row=an_i + 2, column=an_j + 1).border = Border(bottom=thick_side, right=thin_side)


temp_file = temp_output()
wb.save(temp_file)
